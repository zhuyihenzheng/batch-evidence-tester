# -*- coding: utf-8 -*-
"""レビュー指摘の回帰テスト。

このツールは「他のプログラムが正しいか」を判定する立場なので、
偽 OK（本当は NG なのに OK と出る）は最も重い欠陥として扱う。
以下はすべて偽 OK か破壊的操作に関する再現テスト。

標準ライブラリの unittest だけを使う（Anaconda 5.2 に追加インストール不要）。

  python -m unittest discover -s tests -v
"""

import os
import shutil
import sys
import tempfile
import unittest
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from autotest import compare, fsops, logs  # noqa: E402
from autotest.config import ConfigError, Settings, load_cases  # noqa: E402
from autotest.models import NG, OK, REVIEW, SKIP, CaseResult, RunResult, Table  # noqa: E402


def make_table(title, columns, rows, **kw):
    return Table(title=title, columns=columns, rows=rows, **kw)


class TmpDirCase(unittest.TestCase):
    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp(prefix="autotest_ut_"))

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def write_settings(self, paths, extra=None):
        """最小構成の Settings を組み立てる。"""
        raw = {
            "batch": {"exe_path": "dummy.exe"},
            "database": {"server": "s", "database": "d"},
            "paths": paths,
            "log": {"patterns": ["*.log"], "encoding": "utf-8"},
        }
        if extra:
            raw.update(extra)
        return Settings(raw=raw, source=self.tmp / "settings.yaml", project_root=self.tmp)


# =============================================================================
# 1. DB スナップショットの打ち切りが判定に影響してはならない
# =============================================================================
class TestDbTruncationNotUsedForJudgment(unittest.TestCase):
    def test_row_beyond_display_limit_is_still_compared(self):
        """表示上限を超えた行の相違も NG として検出されること。

        表示用の打ち切りと判定用データが同一だと、上限より後ろの異常が
        永久に検査されない（偽 OK）。
        """
        columns = ["ID", "VAL"]
        expected_rows = [[str(i), "ok"] for i in range(1, 502)]
        actual_rows = [[str(i), "ok"] for i in range(1, 501)] + [["501", "BROKEN"]]

        result = compare.compare_db_table(
            "T", make_table("T", columns, actual_rows),
            make_table("T", columns, expected_rows),
            keys=["ID"],
        )
        self.assertEqual(result.verdict, NG, "501 行目の相違が検出されていない（偽 OK）")

    def test_ignoring_every_value_column_is_not_a_pass(self):
        """キー以外を全部 ignore_columns で外したら OK にしないこと。

        「値が合わないから ignore に足す」を続けると、最後には何も照合して
        いないのに緑になる。実際そうなっても気づけないのが一番の問題なので、
        要確認として人の目に触れさせる。
        """
        result = compare.compare_db_table(
            "T", make_table("T", ["ID", "VAL"], [["1", "BROKEN"]]),
            make_table("T", ["ID", "VAL"], [["1", "ok"]]),
            keys=["ID"], ignore_columns=["VAL"])
        self.assertEqual(result.verdict, REVIEW, "値を一切見ていないのに合格になっている")
        self.assertIn("ignore_columns", result.detail)

    def test_partially_ignoring_columns_still_passes(self):
        """時刻列だけ外すような通常の使い方は今までどおり OK。"""
        result = compare.compare_db_table(
            "T", make_table("T", ["ID", "VAL", "TS"], [["1", "ok", "10:00"]]),
            make_table("T", ["ID", "VAL", "TS"], [["1", "ok", "09:00"]]),
            keys=["ID"], ignore_columns=["TS"])
        self.assertEqual(result.verdict, OK)

    def test_expected_with_only_key_columns_is_a_valid_existence_check(self):
        """期待値がキー列だけなのは「このレコードがあること」の確認。OK でよい。"""
        result = compare.compare_db_table(
            "T", make_table("T", ["ID"], [["1"]]),
            make_table("T", ["ID"], [["1"]]), keys=["ID"])
        self.assertEqual(result.verdict, OK)

    def test_record_mismatch_stays_ng_even_when_all_columns_ignored(self):
        """値を見ていなくても、レコードの過不足は NG のまま。"""
        result = compare.compare_db_table(
            "T", make_table("T", ["ID", "VAL"], [["2", "x"]]),
            make_table("T", ["ID", "VAL"], [["1", "ok"]]),
            keys=["ID"], ignore_columns=["VAL"])
        self.assertEqual(result.verdict, NG)

    def test_truncated_table_is_rejected_for_comparison(self):
        """打ち切り済みの Table を判定に使ったら明示的に NG にすること。"""
        columns = ["ID", "VAL"]
        actual = make_table("T", columns, [["1", "ok"]], truncated_from=999)
        expected = make_table("T", columns, [["1", "ok"]])

        result = compare.compare_db_table("T", actual, expected, keys=["ID"])
        self.assertEqual(result.verdict, NG, "打ち切られたデータで OK 判定してはいけない")
        self.assertIn("打ち切", result.detail)


# =============================================================================
# 6. DB の主キー重複を黙って握りつぶしてはならない
# =============================================================================
class TestDuplicateKeyDetection(unittest.TestCase):
    def test_duplicate_key_in_actual_is_ng(self):
        """実績側に同一キーが 2 件あり片方が異常なら NG。

        dict 索引で後勝ちにすると、正しい方だけが残って OK になる（偽 OK）。
        """
        columns = ["ID", "VAL"]
        actual = make_table("T", columns, [["1", "BROKEN"], ["1", "ok"]])
        expected = make_table("T", columns, [["1", "ok"]])

        result = compare.compare_db_table("T", actual, expected, keys=["ID"])
        self.assertEqual(result.verdict, NG, "キー重複が見逃されている（偽 OK）")
        self.assertIn("重複", result.detail)

    def test_duplicate_key_in_expected_is_ng(self):
        columns = ["ID", "VAL"]
        actual = make_table("T", columns, [["1", "ok"]])
        expected = make_table("T", columns, [["1", "ok"], ["1", "ok"]])

        result = compare.compare_db_table("T", actual, expected, keys=["ID"])
        self.assertEqual(result.verdict, NG)


# =============================================================================
# 2. ログの表示打ち切りがキーワード判定に影響してはならない
# =============================================================================
class TestLogTruncationNotUsedForAssertion(TmpDirCase):
    def _slice_with(self, lines):
        return logs.LogSlice(path=self.tmp / "a.log", lines=lines, method="offset",
                             total_lines_before_trim=len(lines))

    def test_error_in_middle_of_long_log_is_detected(self):
        """中略された範囲にある [ERROR] も禁止キーワードとして検出されること。"""
        log_dir = self.tmp / "log"
        log_dir.mkdir()
        lines = ["line %d" % i for i in range(1000)]
        lines[500] = "2026-08-04 10:00:00 [ERROR] 想定外エラー"
        (log_dir / "a.log").write_text("\n".join(lines), encoding="utf-8")

        settings = self.write_settings({"log_dir": str(log_dir)},
                                       extra={"log": {"patterns": ["*.log"], "encoding": "utf-8",
                                                      "max_lines_in_excel": 10}})
        offsets = {}  # 実行前にファイルが無かった＝全文読み
        slices = logs.collect(settings, offsets, datetime.now(), datetime.now(), log_dir=log_dir)

        _missing, forbidden = logs.check_keywords(slices, [], ["[ERROR]"])
        self.assertEqual(forbidden, ["[ERROR]"], "中略部分の ERROR が見逃されている（偽 OK）")

    def test_rotated_file_is_read_from_head(self):
        """ローテートで中身が入れ替わったら、旧オフセットではなく先頭から読むこと。"""
        log_dir = self.tmp / "log"
        log_dir.mkdir()
        path = log_dir / "a.log"

        path.write_text("OLD-A\nOLD-B\nOLD-C\n", encoding="utf-8")
        settings = self.write_settings({"log_dir": str(log_dir)})
        offsets = logs.snapshot_offsets(settings, log_dir)

        # 別内容・かつ旧サイズより大きいファイルに差し替える
        path.write_text("NEW-1 [ERROR] 先頭行\nNEW-2\nNEW-3\nNEW-4\nNEW-5\n", encoding="utf-8")

        slices = logs.collect(settings, offsets, datetime.now(), datetime.now(), log_dir=log_dir)
        text = "\n".join(s.text for s in slices)
        self.assertIn("NEW-1", text, "差し替え後の先頭行が読めていない")


class TestFirstLogLineCollection(TmpDirCase):
    """初回起動で新しく作られたログの 1 行目を必ず採る。"""

    def _settings(self, log_dir):
        return self.write_settings(
            {"log_dir": str(log_dir)},
            extra={"log": {
                "patterns": ["*.log"],
                "encoding": "utf-8",
                "encoding_fallbacks": ["cp932", "utf-8-sig"],
                "timestamp_regex": r"^(?P<ts>\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})",
                "timestamp_format": "%Y-%m-%d %H:%M:%S",
                "slice_margin_sec": 3,
            }})

    def test_new_utf8_bom_log_keeps_first_line(self):
        """C# StreamWriter が初回作成時に付ける BOM で時刻 regex を失敗させない。"""
        log_dir = self.tmp / "log"
        log_dir.mkdir()
        path = log_dir / "batch.log"
        path.write_bytes(
            b"\xef\xbb\xbf2026-08-15 10:00:00 [INFO] first\n"
            b"2026-08-15 10:00:01 [INFO] second\n")

        slices = logs.collect(
            self._settings(log_dir), {},
            datetime(2026, 8, 15, 10, 0, 0), datetime(2026, 8, 15, 10, 0, 2),
            log_dir=log_dir)

        self.assertEqual(slices[0].method, "new")
        self.assertEqual(slices[0].lines[0], "2026-08-15 10:00:00 [INFO] first")
        self.assertEqual(len(slices[0].lines), 2)

    def test_new_log_keeps_leading_header_without_timestamp(self):
        """新規ファイルは全行が今回分なので、時刻の無い先頭ヘッダも捨てない。"""
        log_dir = self.tmp / "log"
        log_dir.mkdir()
        (log_dir / "batch.log").write_text(
            "Batch version 1.0\n2026-08-15 10:00:00 [INFO] start\n", encoding="utf-8")

        slices = logs.collect(
            self._settings(log_dir), {},
            datetime(2026, 8, 15, 10, 0, 0), datetime(2026, 8, 15, 10, 0, 1),
            log_dir=log_dir)
        self.assertEqual(slices[0].lines[0], "Batch version 1.0")

    def test_rotated_bom_log_keeps_first_timestamped_line(self):
        """同名ローテートは時刻抽出するが、BOM 除去後に 1 行目も対象にする。"""
        log_dir = self.tmp / "log"
        log_dir.mkdir()
        path = log_dir / "batch.log"
        path.write_text("OLD-A\nOLD-B\nOLD-C\n", encoding="utf-8")
        settings = self._settings(log_dir)
        offsets = logs.snapshot_offsets(settings, log_dir)
        path.write_bytes(
            b"\xef\xbb\xbf2026-08-15 10:00:00 [INFO] first-after-rotate\n"
            b"2026-08-15 10:00:01 [INFO] second-after-rotate\n")

        slices = logs.collect(
            settings, offsets,
            datetime(2026, 8, 15, 10, 0, 0), datetime(2026, 8, 15, 10, 0, 2),
            log_dir=log_dir)
        self.assertEqual(slices[0].method, "timestamp")
        self.assertIn("first-after-rotate", slices[0].lines[0])


# =============================================================================
# 5. クリア対象の実パスに危険なものを許してはならない
# =============================================================================
class TestClearDirSafety(TmpDirCase):
    """危険パスの拒否テスト。

    ！重要！ここでは実在の危険パス（/ や ~）を clear_dir に渡してはいけない。
    ガードが未実装・退行した瞬間に本物のファイルが消える。
    そのため「判定関数 assert_safe_to_clear() だけを呼ぶ」設計にし、
    削除を伴う経路は一時ディレクトリ配下でしか実行しない。
    """

    def test_rejects_dangerous_paths(self):
        dangerous = [
            Path(os.path.abspath(os.sep)),   # ファイルシステムルート
            Path.home(),                     # ユーザーホーム
            Path.home().parent,              # /Users
            self.tmp,                        # プロジェクトルートそのもの
            self.tmp.parent,                 # プロジェクトルートの上位
        ]
        for target in dangerous:
            with self.subTest(target=str(target)):
                with self.assertRaises(ConfigError):
                    # 削除は一切行わない。判定関数のみを検証する
                    fsops.assert_safe_to_clear(target, project_root=self.tmp)

    def test_allows_normal_subdirectory(self):
        target = self.tmp / "work" / "in"
        target.mkdir(parents=True)
        fsops.assert_safe_to_clear(target, project_root=self.tmp)  # 例外が出なければ OK

    def test_clear_dir_rejects_dangerous_alias_without_deleting(self):
        """clear_dir 経由でも、危険パスなら 1 件も消さずに例外を投げること。"""
        canary = self.tmp / "canary.txt"
        canary.write_text("must survive", encoding="utf-8")

        settings = self.write_settings({"input_dir": str(self.tmp)})  # = プロジェクトルート
        with self.assertRaises(ConfigError):
            fsops.clear_dir(settings, "input_dir")
        self.assertTrue(canary.exists(), "危険判定なのにファイルが削除された")

    def test_clears_normal_subdirectory(self):
        target = self.tmp / "work" / "in"
        target.mkdir(parents=True)
        (target / "x.txt").write_text("x", encoding="utf-8")
        (target / "sub").mkdir()
        (target / "sub" / "y.txt").write_text("y", encoding="utf-8")

        settings = self.write_settings({"input_dir": str(target)})
        self.assertEqual(fsops.clear_dir(settings, "input_dir"), 2)
        self.assertEqual(list(target.iterdir()), [])


# =============================================================================
# 7. 実行ケースがゼロなら成功にしてはならない
# =============================================================================
class TestZeroCases(TmpDirCase):
    def test_all_disabled_raises(self):
        cases_dir = self.tmp / "cases"
        cases_dir.mkdir()
        (cases_dir / "a.yaml").write_text(
            "id: A\nname: A\nenabled: false\n", encoding="utf-8")
        with self.assertRaises(ConfigError):
            load_cases(cases_dir)

    def test_run_result_with_no_cases_is_not_ok(self):
        run = RunResult(run_id="x", started_at=datetime.now())
        self.assertNotEqual(run.verdict, OK, "0 ケースで OK は偽グリーン")

    def test_run_result_all_skip_is_not_ok(self):
        run = RunResult(run_id="x", started_at=datetime.now())
        case = CaseResult(case_id="A", name="A")
        run.cases.append(case)  # checks 空 = SKIP
        self.assertNotEqual(run.verdict, OK)
        self.assertEqual(case.verdict, SKIP)


# =============================================================================
# 10. ケース ID 重複は設定エラー
# =============================================================================
class TestDuplicateCaseId(TmpDirCase):
    def test_duplicate_id_rejected(self):
        cases_dir = self.tmp / "cases"
        cases_dir.mkdir()
        for name in ("a.yaml", "b.yaml"):
            (cases_dir / name).write_text("id: DUP\nname: %s\n" % name, encoding="utf-8")
        with self.assertRaises(ConfigError) as ctx:
            load_cases(cases_dir)
        self.assertIn("DUP", str(ctx.exception))


# =============================================================================
# 8. Excel 生成が内容で壊れたり、数式として解釈されたりしてはならない
# =============================================================================
class TestExcelContentSafety(TmpDirCase):
    def _build(self, rows):
        from autotest.excel import build_workbook

        run = RunResult(run_id="ut", started_at=datetime.now(), finished_at=datetime.now())
        case = CaseResult(case_id="TC", name="TC")
        case.db_after["T"] = make_table("T", ["ID", "VAL"], rows)
        case.checks.append(compare.CheckResult("dummy", "db", OK, "ok"))
        run.cases.append(case)
        out = self.tmp / "e.xlsx"
        build_workbook(run, out, {"max_db_rows": 500})
        return out

    def test_control_character_does_not_break_workbook(self):
        """NUL 等の制御文字が混ざっても証跡簿の生成が失敗しないこと。"""
        out = self._build([["1", "bad\x00value\x07"]])
        self.assertTrue(out.exists())

    def test_formula_like_value_is_not_written_as_formula(self):
        """`=1+1` のような値が Excel 数式として解釈されないこと。"""
        from openpyxl import load_workbook

        out = self._build([["1", "=1+1"], ["2", "+SUM(A1)"], ["3", "-2"], ["4", "@x"]])
        ws = load_workbook(out)["TC"]
        formula_cells = [c.coordinate for row in ws.iter_rows() for c in row if c.data_type == "f"]
        self.assertEqual(formula_cells, [], "数式セルとして書かれている（インジェクション）")
        texts = [c.value for row in ws.iter_rows() for c in row
                 if isinstance(c.value, str) and "1+1" in c.value]
        self.assertTrue(texts, "値そのものが失われている")


# =============================================================================
# 9. 出力ファイルの glob が複数ヒットしたら曖昧として扱う
# =============================================================================
class TestAmbiguousFileMatch(TmpDirCase):
    def test_multiple_matches_reported(self):
        d = self.tmp / "out"
        d.mkdir()
        (d / "RESULT_1.csv").write_text("a", encoding="utf-8")
        (d / "RESULT_2.csv").write_text("b", encoding="utf-8")
        self.assertEqual(len(fsops.find_files(d, "RESULT_*.csv")), 2)


# =============================================================================
# 11. compare_text_file は actual_path だけでも比較できること
# =============================================================================
class TestCompareTextFileFromPath(TmpDirCase):
    def test_identical_file_is_ok_without_actual_text(self):
        expected = self.tmp / "exp.csv"
        actual = self.tmp / "act.csv"
        expected.write_text("a,b\n1,2\n", encoding="utf-8")
        actual.write_text("a,b\n1,2\n", encoding="utf-8")

        result = compare.compare_text_file("f", actual, expected)
        self.assertEqual(result.verdict, OK, "同一内容なのに NG になっている")


if __name__ == "__main__":
    unittest.main(verbosity=2)


# =============================================================================
# 接続エラーの切り分け（実 DB が無い環境でも検証できるようにする）
# =============================================================================
class TestConnectionDiagnosis(unittest.TestCase):
    """pyodbc が実際に返すエラー文字列から、正しい助言が出せること。

    実 DB や ODBC ドライバが無い開発環境でも、切り分けロジックだけは
    検証できるようにしておく（ここが外れると現場で原因に辿り着けない）。
    """

    def _hints(self, message):
        from autotest import db as db_mod
        return " ".join(db_mod.diagnose_connection_error(Exception(message)))

    def test_driver_not_found(self):
        msg = ("('IM002', \"[IM002] [unixODBC][Driver Manager]Data source name not found "
               "and no default driver specified (0) (SQLDriverConnect)\")")
        self.assertIn("driver", self._hints(msg).lower() + "driver")  # 助言が空でないこと
        self.assertIn("一致していません", self._hints(msg))

    def test_server_unreachable(self):
        msg = ("('08001', '[08001] [Microsoft][ODBC Driver 18 for SQL Server]"
               "TCP Provider: Error code 0x2749 (10061) (SQLDriverConnect)')")
        h = self._hints(msg)
        self.assertIn("到達できていません", h)
        self.assertIn("1433", h)

    def test_login_failed(self):
        msg = ("('28000', \"[28000] [Microsoft][ODBC Driver 18 for SQL Server][SQL Server]"
               "Login failed for user 'sa'. (18456)\")")
        h = self._hints(msg)
        self.assertIn("認証で拒否", h)
        self.assertIn("AUTOTEST_DB_PASSWORD", h)

    def test_timeout(self):
        msg = ("('HYT00', '[HYT00] [Microsoft][ODBC Driver 18 for SQL Server]"
               "Login timeout expired (0)')")
        self.assertIn("タイムアウト", self._hints(msg))

    def test_password_is_masked_in_connection_string(self):
        """接続文字列を表示するとき、パスワードが露出しないこと。"""
        import os
        from autotest import db as db_mod
        os.environ["UT_DB_PW"] = "s3cr3t-value"
        settings = Settings(
            raw={"batch": {"exe_path": "x"},
                 "database": {"server": "s", "database": "d", "user": "u",
                              "password_env": "UT_DB_PW", "driver": "X"},
                 "paths": {"a": "/tmp"}},
            source=Path("s.yaml"), project_root=Path("/tmp"))
        conn_str, shown = db_mod.build_connection_string(settings)
        self.assertIn("s3cr3t-value", conn_str)      # 実際の接続には使う
        self.assertNotIn("s3cr3t-value", shown)      # 表示には出さない
        self.assertIn("PWD=********", shown)


# =============================================================================
# 接続先文字列の解釈（SQL Server 特有の記法）
# =============================================================================
class TestServerStringParsing(unittest.TestCase):
    """SQL Server の接続先はポート区切りがコロンではなくカンマ。

    ここを取り違えると、到達確認が誤ったホスト/ポートを見て
    「繋がらない理由」を誤診断してしまう。
    """

    def _parse(self, text):
        from autotest import db as db_mod
        return db_mod.parse_server(text)

    def test_bare_host_uses_default_port(self):
        self.assertEqual(self._parse("SQLSRV01"), ("SQLSRV01", 1433, None))

    def test_explicit_port(self):
        self.assertEqual(self._parse("SQLSRV01,14330"), ("SQLSRV01", 14330, None))

    def test_named_instance_has_dynamic_port(self):
        host, port, inst = self._parse("SQLSRV01\\SQLEXPRESS")
        self.assertEqual((host, inst), ("SQLSRV01", "SQLEXPRESS"))
        self.assertIsNone(port, "名前付きインスタンスのポートを 1433 と決めつけてはいけない")

    def test_named_instance_with_explicit_port(self):
        self.assertEqual(self._parse("SQLSRV01\\SQLEXPRESS,49172"),
                         ("SQLSRV01", 49172, "SQLEXPRESS"))

    def test_rds_endpoint(self):
        host, port, inst = self._parse("db1.abc123.ap-northeast-1.rds.amazonaws.com,1433")
        self.assertEqual(port, 1433)
        self.assertTrue(host.endswith("rds.amazonaws.com"))
        self.assertIsNone(inst)

    def test_colon_is_not_a_port_separator(self):
        """コロン区切りは SQL Server の記法ではない。ホスト名の一部として扱われる。"""
        host, port, _inst = self._parse("SQLSRV01:1433")
        self.assertEqual(host, "SQLSRV01:1433")
        self.assertEqual(port, 1433, "既定ポートが補われること（実際の接続は失敗する想定）")


# =============================================================================
# パスワード環境変数の診断（setx 由来の事故を見つけられること）
# =============================================================================
class TestPasswordEnvDiagnosis(unittest.TestCase):
    def _lines(self, value):
        from autotest import db as db_mod
        return db_mod.diagnose_password_env("AUTOTEST_DB_PASSWORD", value)

    def test_not_set_explains_setx_behaviour(self):
        text = " ".join(self._lines(None))
        self.assertIn("NG", text)
        self.assertIn("setx", text, "setx が新プロセスにしか効かない点を案内すること")

    def test_empty_string_is_ng(self):
        self.assertIn("NG", " ".join(self._lines("")))

    def test_normal_value_reports_length_only(self):
        lines = self._lines("GoodPass123")
        text = " ".join(lines)
        self.assertIn("OK", text)
        self.assertIn("11 文字", text)
        self.assertNotIn("GoodPass123", text, "パスワードそのものを出力してはいけない")

    def test_special_characters_warned(self):
        text = " ".join(self._lines("p@ss%wo&rd"))
        self.assertIn("警告", text)
        self.assertNotIn("p@ss%wo&rd", text, "パスワードそのものを出力してはいけない")

    def test_surrounding_whitespace_warned(self):
        self.assertIn("空白", " ".join(self._lines(" secret ")))

    def test_quoted_value_warned(self):
        self.assertIn("引用符", " ".join(self._lines('"secret"')))


# =============================================================================
# pyodbc の API 前提（Cursor に timeout 属性は無い）
# =============================================================================
class TestPyodbcCursorApi(unittest.TestCase):
    """クエリのタイムアウトは Connection の属性であり Cursor には無い。

    cur.timeout = ... と書くと実行時に AttributeError になり、
    DB 接続自体は成功しているのに最初のスナップショットで落ちる。
    実際に本番環境で発生したため、同じ書き方に戻らないよう固定する。
    """

    def test_query_does_not_touch_cursor_timeout(self):
        import sys as _sys
        from types import ModuleType

        calls = {"conn_timeout": None}

        class FakeCursor:
            description = [("ID",), ("VAL",)]

            def __setattr__(self, name, value):
                if name == "timeout":
                    raise AttributeError("'pyodbc.Cursor' object has no attribute 'timeout'")
                object.__setattr__(self, name, value)

            def execute(self, sql, params=None):
                return self

            def fetchall(self):
                return [("1", "a")]

            def close(self):
                pass

        class FakeConn:
            def __setattr__(self, name, value):
                if name == "timeout":
                    calls["conn_timeout"] = value
                object.__setattr__(self, name, value)

            def cursor(self):
                return FakeCursor()

            def commit(self):
                pass

            def rollback(self):
                pass

            def close(self):
                pass

        fake_pyodbc = ModuleType("pyodbc")
        fake_pyodbc.connect = lambda *a, **k: FakeConn()
        fake_pyodbc.drivers = lambda: ["ODBC Driver 18 for SQL Server"]
        saved = _sys.modules.get("pyodbc")
        _sys.modules["pyodbc"] = fake_pyodbc
        try:
            from autotest.db import PyodbcClient

            os.environ["UT_PW"] = "pw"
            settings = Settings(
                raw={"batch": {"exe_path": "x"},
                     "database": {"server": "s,1433", "database": "d", "user": "u",
                                  "password_env": "UT_PW", "query_timeout_sec": 42},
                     "paths": {"a": "/tmp"}},
                source=Path("s.yaml"), project_root=Path("/tmp"))

            client = PyodbcClient(settings)
            self.assertEqual(calls["conn_timeout"], 42,
                             "クエリタイムアウトは Connection 側に設定すること")

            # ここで AttributeError が出るなら cur.timeout を触っている
            columns, rows = client.query("SELECT ID, VAL FROM T")
            self.assertEqual(columns, ["ID", "VAL"])
            self.assertEqual(rows, [["1", "a"]])

            client.execute_script("DELETE FROM T\nGO\n")
        finally:
            if saved is not None:
                _sys.modules["pyodbc"] = saved
            else:
                _sys.modules.pop("pyodbc", None)


# =============================================================================
# 日付プレースホルダ（batch が日付別フォルダを使う構成への対応）
# =============================================================================
class TestDatePlaceholders(unittest.TestCase):
    """Backup/20260803 のような日付別フォルダを設定で表現できること。"""

    def _settings(self, paths, base=None):
        s = Settings(raw={"batch": {"exe_path": "x"},
                          "database": {"server": "s", "database": "d"},
                          "paths": paths},
                     source=Path("s.yaml"), project_root=Path("/tmp/proj"))
        s.set_base_date(base)
        return s

    def test_date_expands_to_yyyymmdd(self):
        s = self._settings({"backup_dir": "/data/Backup/{date}"}, "20260803")
        self.assertEqual(str(s.resolve_dir("backup_dir")), "/data/Backup/20260803")

    def test_custom_format(self):
        s = self._settings({"d": "/data/{date:%Y}/{date:%m}/{date:%d}"}, "20260803")
        self.assertEqual(str(s.resolve_dir("d")), "/data/2026/08/03")

    def test_negative_offset_for_previous_day(self):
        """前日データを参照するケースがあるためオフセットを使えること。"""
        s = self._settings({"d": "/data/{date-1}"}, "20260803")
        self.assertEqual(str(s.resolve_dir("d")), "/data/20260802")

    def test_offset_crosses_month_boundary(self):
        s = self._settings({"d": "/data/{date-1}"}, "20260801")
        self.assertEqual(str(s.resolve_dir("d")), "/data/20260731")

    def test_positive_offset(self):
        s = self._settings({"d": "/data/{date+1}"}, "20261231")
        self.assertEqual(str(s.resolve_dir("d")), "/data/20270101")

    def test_defaults_to_today(self):
        from datetime import date as _date
        s = self._settings({"d": "/data/{date}"})
        self.assertEqual(str(s.resolve_dir("d")), "/data/" + _date.today().strftime("%Y%m%d"))

    def test_accepts_hyphenated_input(self):
        s = self._settings({"d": "/data/{date}"}, "2026-08-03")
        self.assertEqual(str(s.resolve_dir("d")), "/data/20260803")

    def test_invalid_date_is_config_error(self):
        s = self._settings({"d": "/data/{date}"})
        with self.assertRaises(ConfigError):
            s.set_base_date("2026年8月3日")

    def test_path_without_placeholder_is_untouched(self):
        s = self._settings({"d": "/data/fixed"}, "20260803")
        self.assertEqual(str(s.resolve_dir("d")), "/data/fixed")

    def test_expand_applies_to_patterns_and_args(self):
        s = self._settings({"d": "/tmp"}, "20260803")
        self.assertEqual(s.expand("RESULT_{date}.csv"), "RESULT_20260803.csv")
        self.assertEqual(s.expand("--date"), "--date")


# =============================================================================
# 入れ子フォルダの保護（Receive の下に Backup がある構成）
# =============================================================================
class TestNestedFolderProtection(TmpDirCase):
    """クリア対象の配下に既存データのフォルダがある構成を壊さないこと。

    実案件の構成: Receive/ にエラーファイルが置かれ、同じ Receive/ の下に
    Backup/20260802/ のような履歴フォルダがある。Receive をクリアすると
    履歴ごと消えてしまう（rmtree は復元不能）。
    """

    def _layout(self):
        recv = self.tmp / "work" / "TAS" / "Receive"
        (recv / "Backup" / "20260802").mkdir(parents=True)
        (recv / "Backup" / "20260802" / "old.csv").write_text("履歴", encoding="utf-8")
        (recv / "ERR_001.csv").write_text("err", encoding="utf-8")
        return recv

    def test_nested_alias_is_protected_automatically(self):
        """配下のフォルダが論理名として定義済みなら、明示しなくても残ること。"""
        recv = self._layout()
        settings = self.write_settings({"error_dir": str(recv), "backup_dir": str(recv / "Backup")})

        removed = fsops.clear_dir(settings, "error_dir")
        self.assertEqual(removed, 1, "エラーファイルだけが消えること")
        self.assertTrue((recv / "Backup" / "20260802" / "old.csv").exists(),
                        "履歴データが消えている（復元不能）")
        self.assertFalse((recv / "ERR_001.csv").exists())

    def test_explicit_exclude_protects_undeclared_folder(self):
        """論理名でない補助フォルダも exclude で残せること。"""
        recv = self._layout()
        (recv / "Keep").mkdir()
        (recv / "Keep" / "a.txt").write_text("z", encoding="utf-8")
        settings = self.write_settings({"error_dir": str(recv)})

        fsops.clear_dir(settings, "error_dir", exclude=["Backup", "Keep"])
        self.assertTrue((recv / "Backup").exists())
        self.assertTrue((recv / "Keep" / "a.txt").exists())
        self.assertFalse((recv / "ERR_001.csv").exists())

    def test_file_count_excludes_directories(self):
        """assert の件数はフォルダを数えないこと（Backup が 1 件と数えられない）。"""
        recv = self._layout()
        self.assertEqual(len(fsops.find_files(recv, "*")), 1,
                         "サブフォルダがファイルとして数えられている")

    def test_listing_separates_files_and_folders(self):
        """証跡一覧はファイルとフォルダを区別できる形で返すこと。"""
        recv = self._layout()
        entries = fsops.list_dir(recv)
        dirs = [e for e in entries if e.is_dir]
        files = [e for e in entries if not e.is_dir]
        self.assertEqual([d.name for d in dirs], ["Backup"])
        self.assertEqual([f.name for f in files], ["ERR_001.csv"])


# =============================================================================
# {batch_start} プレースホルダ（batch が触った行だけを抽出する）
# =============================================================================
class TestBatchStartPlaceholder(TmpDirCase):
    """スナップショット SQL に batch 起動時刻を埋め込めること。

    基準時刻は DB サーバから取る。テスト機の時計を使うと、両者にずれが
    あるとき対象行を取りこぼす／無関係な行まで拾うため。
    """

    def _runner(self):
        from autotest.orchestrator import CaseRunner
        settings = self.write_settings({"log_dir": str(self.tmp)})
        return CaseRunner(settings, self.tmp)

    def test_placeholder_replaced_with_db_time(self):
        import datetime as _dt
        runner = self._runner()
        runner._batch_start_mark = _dt.datetime(2026, 8, 5, 9, 15, 22, 123456)

        sql = runner._expand_sql(
            "SELECT * FROM T_ORDER WHERE LAST_UPDATE >= '{batch_start}'")
        # 言語設定に依存しない ISO 8601（T 区切り）で埋め込む
        self.assertIn("'2026-08-05T09:15:22.123'", sql)
        self.assertNotIn("{batch_start}", sql)

    def test_falls_back_to_local_time_when_db_time_unavailable(self):
        runner = self._runner()
        runner._batch_start_mark = None
        sql = runner._expand_sql("WHERE LAST_UPDATE >= '{batch_start}'")
        self.assertNotIn("{batch_start}", sql, "展開されずに SQL へ渡ってはいけない")

    def test_date_placeholder_also_works_in_sql(self):
        runner = self._runner()
        runner.settings.set_base_date("20260805")
        sql = runner._expand_sql("WHERE ORDER_DATE = '{date}'")
        self.assertIn("'20260805'", sql)

    def test_sql_without_placeholder_is_untouched(self):
        runner = self._runner()
        original = "SELECT ORDER_ID FROM T_ORDER ORDER BY ORDER_ID"
        self.assertEqual(runner._expand_sql(original), original)


class TestSqlDatetimeLiteral(unittest.TestCase):
    """datetime リテラルは言語設定に依存しない書式で埋め込むこと。

    'yyyy-mm-dd hh:mi:ss'（空白区切り）は datetime 型では SET LANGUAGE /
    DATEFORMAT の影響を受け、環境によっては変換エラー（241）になる。
    ISO 8601 の T 区切りは言語設定に依存しない。
    """

    def test_uses_iso8601_t_separator(self):
        import datetime as _dt
        from autotest import db as db_mod
        got = db_mod.format_sql_datetime(_dt.datetime(2026, 8, 5, 0, 40, 47, 872000))
        self.assertEqual(got, "2026-08-05T00:40:47.872")
        self.assertNotIn(" ", got, "空白区切りは言語設定に依存するため使わない")

    def test_conversion_error_is_diagnosed(self):
        from autotest import db as db_mod
        msg = ("('22007', '[22007] [Microsoft][ODBC Driver 17 for SQL Server]"
               "文字列から日付と時刻、またはそのいずれかへの変換中に、変換が失敗しました。"
               "(241) (SQLExecDirectW)')")
        hints = " ".join(db_mod.diagnose_connection_error(Exception(msg)))
        self.assertIn("日付時刻", hints)
        self.assertIn("batch_start:%Y%m%d", hints, "char 列向けの書式指定を案内すること")


class TestBatchStartFormatting(TmpDirCase):
    def _runner(self):
        from autotest.orchestrator import CaseRunner
        return CaseRunner(self.write_settings({"log_dir": str(self.tmp)}), self.tmp)

    def test_default_is_iso8601(self):
        import datetime as _dt
        r = self._runner()
        r._batch_start_mark = _dt.datetime(2026, 8, 5, 0, 40, 47, 872000)
        self.assertIn("'2026-08-05T00:40:47.872'",
                      r._expand_sql("WHERE UPDATED_AT >= '{batch_start}'"))

    def test_custom_format_for_char_columns(self):
        """日付が char(8) に 'yyyyMMdd' で入っている列にも合わせられること。"""
        import datetime as _dt
        r = self._runner()
        r._batch_start_mark = _dt.datetime(2026, 8, 5, 0, 40, 47)
        self.assertIn("'20260805'",
                      r._expand_sql("WHERE UPDATE_YMD >= '{batch_start:%Y%m%d}'"))

    def test_custom_format_with_slashes(self):
        import datetime as _dt
        r = self._runner()
        r._batch_start_mark = _dt.datetime(2026, 8, 5, 0, 40, 47)
        self.assertIn("'2026/08/05 00:40:47'",
                      r._expand_sql("WHERE UPD >= '{batch_start:%Y/%m/%d %H:%M:%S}'"))


class TestQueryErrorIncludesSql(unittest.TestCase):
    """クエリ失敗時、実際に送った SQL がエラーに含まれること。

    含まれないと、プレースホルダがどう展開されたか分からず原因に辿り着けない。
    """

    def test_sql_is_in_error_message(self):
        import sys as _sys
        from types import ModuleType

        class FailingCursor:
            description = None

            def execute(self, sql, params=None):
                raise RuntimeError("22007 conversion failed")

            def close(self):
                pass

        class FakeConn:
            def cursor(self):
                return FailingCursor()

            def rollback(self):
                pass

            def close(self):
                pass

        fake = ModuleType("pyodbc")
        fake.connect = lambda *a, **k: FakeConn()
        fake.drivers = lambda: []
        saved = _sys.modules.get("pyodbc")
        _sys.modules["pyodbc"] = fake
        try:
            from autotest.db import DbError, PyodbcClient

            os.environ["UT_PW2"] = "pw"
            settings = Settings(
                raw={"batch": {"exe_path": "x"},
                     "database": {"server": "s", "database": "d", "user": "u",
                                  "password_env": "UT_PW2"},
                     "paths": {"a": "/tmp"}},
                source=Path("s.yaml"), project_root=Path("/tmp"))
            client = PyodbcClient(settings)
            with self.assertRaises(DbError) as ctx:
                client.query("SELECT * FROM T WHERE UPD >= '2026-08-05T00:40:47.872'")
            self.assertIn("2026-08-05T00:40:47.872", str(ctx.exception),
                          "展開後の SQL がエラーに含まれていない")
        finally:
            if saved is not None:
                _sys.modules["pyodbc"] = saved
            else:
                _sys.modules.pop("pyodbc", None)


# =============================================================================
# {batch_start} を使わないケースでは DB 時刻を問い合わせないこと
# =============================================================================
class TestBatchStartSource(TmpDirCase):
    """{batch_start} の基準時刻は DB サーバの時計から取ること。

    テスト機と DB でずれていると batch が更新した行を取りこぼす
    （＝異常の見逃し）。DB から取れないときだけテスト機の時計で代用し、
    その場合は取りこぼさないようマージン分さかのぼる。
    """

    class _Client:
        def __init__(self, value="2026-08-05 09:15:22"):
            self.value = value
            self.queries = []

        def query(self, sql, params=None):
            self.queries.append(sql)
            if self.value is None:
                raise RuntimeError("not available")
            return ["now"], [[self.value]]

        def execute_script(self, sql):
            pass

    def _case(self, sql):
        from autotest.config import TestCase
        return TestCase(case_id="T", name="T", source=self.tmp / "t.yaml",
                        snapshot={"tables": [{"name": "T_ORDER", "sql": sql}]})

    def _runner(self, margin=None):
        from autotest.orchestrator import CaseRunner
        extra = {"database": {"server": "s", "database": "d"}}
        if margin is not None:
            extra["database"]["batch_start_margin_sec"] = margin
        return CaseRunner(self.write_settings({"log_dir": str(self.tmp)}, extra=extra), self.tmp)

    def test_uses_db_time(self):
        client = self._Client("2026-08-05 09:15:22")
        mark = self._runner()._resolve_batch_start(
            self._case("WHERE UPD >= '{batch_start}'"), client)
        self.assertEqual(mark.strftime("%Y-%m-%d %H:%M:%S"), "2026-08-05 09:15:22")
        self.assertTrue(client.queries, "DB へ問い合わせていない")

    def test_not_queried_when_placeholder_unused(self):
        """使っていないケースで余計な問い合わせをしないこと。"""
        client = self._Client()
        mark = self._runner()._resolve_batch_start(
            self._case("SELECT * FROM T_ORDER ORDER BY ORDER_ID"), client)
        self.assertIsNone(mark)
        self.assertEqual(client.queries, [])

    def test_falls_back_to_local_clock_with_margin(self):
        """DB から取れないときはテスト機の時計。取りこぼし防止に数秒さかのぼる。"""
        from datetime import datetime as _dtm
        client = self._Client(None)          # 常に失敗する
        before = _dtm.now()
        mark = self._runner(margin=30)._resolve_batch_start(
            self._case("WHERE UPD >= '{batch_start}'"), client)
        self.assertIsNotNone(mark)
        self.assertGreaterEqual((before - mark).total_seconds(), 29, "マージンが引かれていない")

    def test_falls_back_to_getdate(self):
        """SYSDATETIME が使えない環境向けに GETDATE を試すこと。"""
        class OnlyGetdate(self._Client):
            def query(self, sql, params=None):
                self.queries.append(sql)
                if "SYSDATETIME" in sql:
                    raise RuntimeError("not supported")
                return ["now"], [["2026-08-05 09:15:22"]]

        client = OnlyGetdate()
        mark = self._runner()._resolve_batch_start(
            self._case("WHERE UPD >= '{batch_start}'"), client)
        self.assertIsNotNone(mark)
        self.assertTrue(any("GETDATE" in q for q in client.queries))


class TestLocalConfigPreference(TmpDirCase):
    """config/settings.local.yaml があればそちらを既定にすること。

    settings.yaml はリポジトリ側が更新し続けるため、各自の環境値を
    そこに書くと毎回 git pull で衝突する。local 側は .gitignore 済み。
    """

    def test_local_config_is_preferred(self):
        import autotest.cli as cli
        saved = cli.PROJECT_ROOT
        try:
            cli.PROJECT_ROOT = self.tmp
            (self.tmp / "config").mkdir()
            (self.tmp / "config" / "settings.yaml").write_text("a: 1", encoding="utf-8")
            self.assertEqual(cli._default_config().name, "settings.yaml")

            (self.tmp / "config" / "settings.local.yaml").write_text("a: 2", encoding="utf-8")
            self.assertEqual(cli._default_config().name, "settings.local.yaml")
        finally:
            cli.PROJECT_ROOT = saved

    def test_local_config_is_gitignored(self):
        """*.local.yaml が .gitignore に入っていること。"""
        gitignore = (Path(__file__).resolve().parent.parent / ".gitignore").read_text(encoding="utf-8")
        self.assertIn("*.local.yaml", gitignore)


# =============================================================================
# batch 設定ファイルの差し替えと復元
# =============================================================================
class TestReplaceFiles(TmpDirCase):
    """ケースごとに batch の設定ファイルを差し替え、実行後に必ず戻すこと。

    戻し損ねると環境が変更されたまま残り、次のケースや手動実行が
    壊れた設定で動いてしまう。
    """

    def _setup_layout(self):
        batch_dir = self.tmp / "app" / "batch"
        batch_dir.mkdir(parents=True)
        (batch_dir / "OrderBatch.exe.config").write_text("ORIGINAL", encoding="utf-8")
        case_dir = self.tmp / "cases" / "TC"
        (case_dir / "config").mkdir(parents=True)
        (case_dir / "config" / "case01.config").write_text("CASE01", encoding="utf-8")
        return batch_dir, case_dir

    def _settings(self, batch_dir):
        return self.write_settings({"batch_dir": str(batch_dir)})

    def test_replaced_then_restored(self):
        batch_dir, case_dir = self._setup_layout()
        settings = self._settings(batch_dir)
        target = batch_dir / "OrderBatch.exe.config"

        replaced = fsops.replace_files(
            settings, case_dir,
            [{"src": "config/case01.config", "dest_dir": "batch_dir",
              "name": "OrderBatch.exe.config"}],
            backup_root=self.tmp / "backup")
        self.assertEqual(target.read_text(encoding="utf-8"), "CASE01", "差し替わっていない")

        self.assertEqual(fsops.restore_files(replaced), [])
        self.assertEqual(target.read_text(encoding="utf-8"), "ORIGINAL", "元に戻っていない")

    def test_file_created_when_absent_is_removed_on_restore(self):
        """元々無かったファイルは、復元時に削除されること（ゴミを残さない）。"""
        batch_dir, case_dir = self._setup_layout()
        settings = self._settings(batch_dir)
        target = batch_dir / "extra.config"

        replaced = fsops.replace_files(
            settings, case_dir,
            [{"src": "config/case01.config", "dest_dir": "batch_dir", "name": "extra.config"}],
            backup_root=self.tmp / "backup")
        self.assertTrue(target.exists())

        fsops.restore_files(replaced)
        self.assertFalse(target.exists(), "差し替えで作ったファイルが残っている")

    def test_missing_source_is_config_error(self):
        batch_dir, case_dir = self._setup_layout()
        with self.assertRaises(ConfigError):
            fsops.replace_files(
                self._settings(batch_dir), case_dir,
                [{"src": "config/nope.config", "dest_dir": "batch_dir"}],
                backup_root=self.tmp / "backup")

    def test_restore_continues_after_one_failure(self):
        """1 つ戻せなくても、残りは戻すこと（中途半端な状態で放置しない）。"""
        batch_dir, case_dir = self._setup_layout()
        settings = self._settings(batch_dir)
        (batch_dir / "second.config").write_text("ORIG2", encoding="utf-8")

        replaced = fsops.replace_files(
            settings, case_dir,
            [{"src": "config/case01.config", "dest_dir": "batch_dir", "name": "OrderBatch.exe.config"},
             {"src": "config/case01.config", "dest_dir": "batch_dir", "name": "second.config"}],
            backup_root=self.tmp / "backup")

        # 1 件目の退避ファイルを壊して復元を失敗させる
        replaced[0].backup.unlink()
        replaced[0].existed = True
        replaced[0].backup = self.tmp / "backup" / "gone.orig"

        problems = fsops.restore_files(replaced)
        self.assertEqual((batch_dir / "second.config").read_text(encoding="utf-8"), "ORIG2",
                         "2 件目が戻っていない")
        # 戻せなかった事実を黙って握りつぶさないこと（環境が変更されたまま残る）
        self.assertTrue(problems, "復元失敗が報告されていない")
        self.assertIn("OrderBatch.exe.config", " ".join(problems))

    def test_preflight_rejects_cleaning_the_replace_target(self):
        """差し替え先を clean_dirs に入れる矛盾を検出すること。"""
        from autotest.config import TestCase
        from autotest.orchestrator import preflight_case

        batch_dir, case_dir = self._setup_layout()
        settings = self._settings(batch_dir)
        settings.raw["batch"]["exe_path"] = str(batch_dir / "OrderBatch.exe.config")

        case = TestCase(case_id="TC", name="TC", source=self.tmp / "cases" / "TC.yaml",
                        setup={"clean_dirs": ["batch_dir"],
                               "replace_files": [{"src": "config/case01.config",
                                                  "dest_dir": "batch_dir"}]})
        problems = " ".join(preflight_case(settings, case))
        self.assertIn("clean_dirs", problems)


# =============================================================================
# ケース絞り込み（--case / --tag）
# =============================================================================
class TestCaseFiltering(TmpDirCase):
    def _make_cases(self):
        d = self.tmp / "cases"
        d.mkdir()
        defs = [
            ("A", "TC_A", ["正常系", "単体"]),
            ("B", "TC_B", ["異常系", "単体"]),
            ("C", "TC_C", ["異常系", "環境不備"]),
        ]
        for fname, cid, tags in defs:
            (d / (fname + ".yaml")).write_text(
                "id: %s\nname: %s\ntags: [%s]\n" % (cid, cid, ", ".join(tags)),
                encoding="utf-8")
        return d

    def test_multiple_cases_are_all_selected(self):
        got = load_cases(self._make_cases(), only=["TC_A", "TC_C"])
        self.assertEqual(sorted(c.case_id for c in got), ["TC_A", "TC_C"])

    def test_tag_selects_all_matching(self):
        got = load_cases(self._make_cases(), tags=["異常系"])
        self.assertEqual(sorted(c.case_id for c in got), ["TC_B", "TC_C"])

    def test_multiple_tags_are_or(self):
        """複数タグはいずれかに一致すれば対象（AND ではない）。"""
        got = load_cases(self._make_cases(), tags=["正常系", "環境不備"])
        self.assertEqual(sorted(c.case_id for c in got), ["TC_A", "TC_C"])

    def test_case_and_tag_combine_as_and(self):
        got = load_cases(self._make_cases(), only=["TC_A", "TC_B"], tags=["異常系"])
        self.assertEqual([c.case_id for c in got], ["TC_B"])

    def test_unknown_case_id_is_error(self):
        """打ち間違いを黙って 0 件実行にしないこと。"""
        with self.assertRaises(ConfigError) as ctx:
            load_cases(self._make_cases(), only=["TC_TYPO"])
        self.assertIn("TC_TYPO", str(ctx.exception))

    def test_unknown_tag_is_error(self):
        with self.assertRaises(ConfigError):
            load_cases(self._make_cases(), tags=["存在しないタグ"])


class TestListAlignment(unittest.TestCase):
    """全角を含む一覧が桁ずれしないこと。"""

    def test_east_asian_width_counted_as_two(self):
        from autotest.cli import _display_width
        self.assertEqual(_display_width("abc"), 3)
        self.assertEqual(_display_width("異常系"), 6)
        self.assertEqual(_display_width("TC_異常"), 7)

    def test_pad_aligns_by_display_width(self):
        from autotest.cli import _display_width, _pad
        for text in ("TC001", "異常系/環境不備", "見本"):
            self.assertEqual(_display_width(_pad(text, 20)), 20)

    def test_pad_keeps_separation_when_overflowing(self):
        from autotest.cli import _pad
        self.assertTrue(_pad("very_long_case_id_here", 5).endswith(" "))


# =============================================================================
# 機能別のフォルダ構成（サブフォルダ = 暗黙のタグ）
# =============================================================================
class TestCaseFolderStructure(TmpDirCase):
    """機能ごとにサブフォルダで整理し、--tag <機能名> でまとめて実行できること。"""

    def _build(self):
        root = self.tmp / "cases"
        (root / "受注" / "TC_ORDER01" / "input").mkdir(parents=True)
        (root / "請求").mkdir(parents=True)
        (root / "受注" / "TC_ORDER01.yaml").write_text(
            "id: TC_ORDER01\nname: 受注取込\ntags: [正常系]\nexecute: {batch: order}\n",
            encoding="utf-8")
        (root / "請求" / "TC_INV01.yaml").write_text(
            "id: TC_INV01\nname: 請求処理\ntags: [正常系]\nexecute: {batch: invoice}\n",
            encoding="utf-8")
        (root / "TC_TOP.yaml").write_text(
            "id: TC_TOP\nname: トップ直下\n", encoding="utf-8")
        return root

    def test_subdirectories_are_loaded(self):
        got = {c.case_id for c in load_cases(self._build())}
        self.assertEqual(got, {"TC_ORDER01", "TC_INV01", "TC_TOP"})

    def test_folder_name_becomes_tag(self):
        by_id = {c.case_id: c for c in load_cases(self._build())}
        self.assertIn("受注", by_id["TC_ORDER01"].tags)
        self.assertIn("請求", by_id["TC_INV01"].tags)
        self.assertEqual(by_id["TC_TOP"].tags, [], "トップ直下は暗黙タグ無し")

    def test_declared_tags_are_kept(self):
        by_id = {c.case_id: c for c in load_cases(self._build())}
        self.assertIn("正常系", by_id["TC_ORDER01"].tags)

    def test_filter_by_folder_tag(self):
        got = load_cases(self._build(), tags=["受注"])
        self.assertEqual([c.case_id for c in got], ["TC_ORDER01"])

    def test_nested_folders_become_multiple_tags(self):
        root = self.tmp / "cases"
        (root / "受注" / "取込").mkdir(parents=True)
        (root / "受注" / "取込" / "TC_X.yaml").write_text(
            "id: TC_X\nname: X\n", encoding="utf-8")
        by_id = {c.case_id: c for c in load_cases(root)}
        self.assertEqual(by_id["TC_X"].tags, ["受注", "取込"])

    def test_yaml_inside_case_material_folder_is_ignored(self):
        """資材フォルダ内の YAML をケース定義と誤認しないこと。"""
        root = self._build()
        (root / "受注" / "TC_ORDER01" / "input" / "notacase.yaml").write_text(
            "id: SHOULD_NOT_LOAD\nname: 資材\n", encoding="utf-8")
        got = {c.case_id for c in load_cases(root)}
        self.assertNotIn("SHOULD_NOT_LOAD", got)

    def test_material_dir_resolves_under_subfolder(self):
        by_id = {c.case_id: c for c in load_cases(self._build())}
        self.assertTrue(str(by_id["TC_ORDER01"].dir).endswith("受注/TC_ORDER01"))


# =============================================================================
# 1 ファイル 1 ケース書式のベースライン
#   これがこのツール唯一のケース定義書式。読み取りの解釈が変わっていないことを
#   固定しておく（低コストの安全策）。
# =============================================================================
class TestSingleCaseFileFormat(TmpDirCase):
    def _write(self, text, name="TC_SINGLE.yaml"):
        root = self.tmp / "cases"
        root.mkdir(parents=True, exist_ok=True)
        (root / name).write_text(text, encoding="utf-8")
        return root

    def test_all_sections_are_read_as_written(self):
        root = self._write(
            "id: TC_SINGLE\n"
            "name: 単一\n"
            "tags: [正常系]\n"
            "setup: {clean_dirs: [input_dir]}\n"
            "assert: {exit_code: 0}\n")
        case = load_cases(root)[0]
        self.assertEqual(case.case_id, "TC_SINGLE")
        self.assertEqual(case.name, "単一")
        self.assertEqual(case.tags, ["正常系"])
        self.assertEqual(case.setup["clean_dirs"], ["input_dir"])
        self.assertEqual(case.assertions["exit_code"], 0)

    def test_id_defaults_to_file_name(self):
        """id: を書かない場合はファイル名がケース ID になること。"""
        root = self._write("name: ID なし\n", name="TC_FROM_NAME.yaml")
        self.assertEqual(load_cases(root)[0].case_id, "TC_FROM_NAME")

    def test_unknown_top_level_key_is_rejected(self):
        """綴り間違い（および廃止された cases: 形式）を黙って無視しないこと。"""
        root = self._write("id: TC_X\nname: X\ncases: []\n")
        with self.assertRaises(ConfigError) as ctx:
            load_cases(root)
        self.assertIn("cases", str(ctx.exception))


class TestDbLockKey(TmpDirCase):
    """orchestrator が読む項目は、ケース定義の許可キーにも入っていること。

    許可キーから漏れていると、ドキュメントどおりに書いたケースが
    「未知の項目」として読み込み時に弾かれる。
    """

    def _write(self, body):
        root = self.tmp / "cases"
        root.mkdir(parents=True, exist_ok=True)
        (root / "TC_LOCK.yaml").write_text(body, encoding="utf-8")
        return root

    def test_db_lock_is_accepted(self):
        root = self._write(
            "id: TC_LOCK\nname: ロック\n"
            "setup:\n  db_lock: \"BEGIN TRAN; UPDATE T WITH (TABLOCKX) SET A = A\"\n")
        self.assertIn("BEGIN TRAN", load_cases(root)[0].setup["db_lock"])

    def test_non_string_db_lock_is_rejected(self):
        root = self._write(
            "id: TC_LOCK\nname: ロック\n"
            "setup:\n  db_lock: [\"BEGIN TRAN\", \"UPDATE T SET A = A\"]\n")
        with self.assertRaises(ConfigError) as ctx:
            load_cases(root)
        self.assertIn("db_lock", str(ctx.exception))


# =============================================================================
# 絞り込み実行の証跡（何を実行しなかったかも残す）
# =============================================================================
class TestFilteredRunEvidence(TmpDirCase):
    """一部だけ実行した結果を「全体が合格した」と誤読させないこと。"""

    def _run_result(self, executed, total, description):
        from autotest.models import RunResult
        run = RunResult(run_id="ut", started_at=datetime.now(), finished_at=datetime.now(),
                        filter_description=description, total_available=total)
        for i in range(executed):
            case = CaseResult(case_id="TC%d" % i, name="case%d" % i)
            case.checks.append(compare.CheckResult("c", "db", OK, "ok"))
            run.cases.append(case)
        return run

    def _summary_values(self, run):
        from openpyxl import load_workbook
        from autotest.excel import build_workbook
        out = self.tmp / "e.xlsx"
        build_workbook(run, out, {})
        ws = load_workbook(out)["サマリ"]
        return {r[0].value: r[1].value for r in ws.iter_rows(max_row=16) if r[0].value}

    def test_filter_condition_is_recorded(self):
        values = self._summary_values(self._run_result(3, 6, "タグ指定: 異常系"))
        self.assertEqual(values["実行対象"], "タグ指定: 異常系")

    def test_excluded_count_is_shown(self):
        values = self._summary_values(self._run_result(3, 6, "タグ指定: 異常系"))
        self.assertIn("全 6 件中", values["実行ケース数"])
        self.assertIn("3 件は今回実行していません", values["実行ケース数"])

    def test_full_run_says_no_filter(self):
        values = self._summary_values(self._run_result(6, 6, ""))
        self.assertIn("絞り込みなし", values["実行対象"])
        self.assertEqual(str(values["実行ケース数"]), "6")

    def test_filter_description_from_args(self):
        from autotest.cli import _filter_description

        class Args:
            cases = None
            tags = ["異常系", "環境不備"]
        self.assertIn("異常系", _filter_description(Args()))

        class Args2:
            cases = ["TC001"]
            tags = None
        self.assertIn("TC001", _filter_description(Args2()))

        class Args3:
            cases = None
            tags = None
        self.assertEqual(_filter_description(Args3()), "")

    def test_filter_slug_is_filename_safe(self):
        from autotest.cli import _filter_slug

        class Args:
            cases = None
            tags = ["異常系"]
        slug = _filter_slug(Args())
        self.assertNotIn("/", slug)
        self.assertNotIn("\\\\", slug)
        self.assertIn("tag-", slug)


# =============================================================================
# 人による最終確認（manual: true）
# =============================================================================
class TestManualReview(unittest.TestCase):
    """自動判定だけで確定させず、人が証跡を見て判定する項目を扱えること。"""

    def _check(self, verdict, detail="d"):
        return compare.CheckResult("c", "db", verdict, detail)

    def test_ok_becomes_review_when_manual(self):
        from autotest.models import REVIEW
        from autotest.orchestrator import _apply_manual
        got = _apply_manual(self._check(OK, "2 件一致"), manual=True)
        self.assertEqual(got.verdict, REVIEW)
        self.assertIn("2 件一致", got.detail, "自動比較の結果は残すこと")
        self.assertIn("目視", got.detail)

    def test_ng_stays_ng_even_when_manual(self):
        """NG は人の確認を待たずに問題として扱うこと。"""
        from autotest.orchestrator import _apply_manual
        self.assertEqual(_apply_manual(self._check(NG), manual=True).verdict, NG)

    def test_untouched_when_not_manual(self):
        from autotest.orchestrator import _apply_manual
        self.assertEqual(_apply_manual(self._check(OK), manual=False).verdict, OK)

    def test_case_verdict_is_review_when_any_check_pending(self):
        from autotest.models import REVIEW
        case = CaseResult(case_id="T", name="T")
        case.checks.append(self._check(OK))
        case.checks.append(self._check(REVIEW))
        self.assertEqual(case.verdict, REVIEW)
        self.assertEqual(case.review_count, 1)

    def test_ng_takes_precedence_over_review(self):
        """NG と要確認が混在したら NG。合格でないことを優先して示す。"""
        from autotest.models import REVIEW
        case = CaseResult(case_id="T", name="T")
        case.checks.append(self._check(REVIEW))
        case.checks.append(self._check(NG))
        self.assertEqual(case.verdict, NG)

    def test_run_verdict_is_review(self):
        from autotest.models import REVIEW, RunResult
        run = RunResult(run_id="r", started_at=datetime.now())
        ok_case = CaseResult(case_id="A", name="A")
        ok_case.checks.append(self._check(OK))
        pending = CaseResult(case_id="B", name="B")
        pending.checks.append(self._check(REVIEW))
        run.cases.extend([ok_case, pending])
        self.assertEqual(run.verdict, REVIEW, "確認待ちが残る間は合格にしない")
        self.assertEqual(run.review_count, 1)

    def test_review_run_is_not_ok(self):
        from autotest.models import REVIEW
        self.assertNotEqual(REVIEW, OK)


class TestYamlReviewPoints(TmpDirCase):
    """assert の種類に縛られない人工確認観点を YAML から作れること。"""

    def test_review_points_become_review_checks_without_skip_noise(self):
        from autotest.models import REVIEW
        from autotest.orchestrator import CaseRunner

        cases_dir = self.tmp / "cases"
        cases_dir.mkdir()
        (cases_dir / "R.yaml").write_text(
            "id: R\nname: R\n"
            "review:\n"
            "  - DB 実行前後を比較し、更新状態が仕様どおりであること\n"
            "  - |\n"
            "      実行ログを確認すること。\n"
            "      想定外のエラーが無いこと。\n",
            encoding="utf-8")
        case = load_cases(cases_dir)[0]
        runner = CaseRunner(
            self.write_settings({"output_dir": str(self.tmp / "out")}), self.tmp / "run")

        checks = runner._assert_all(case, None, CaseResult("R", "R"), [])

        self.assertEqual(
            [c.name for c in checks],
            ["DB 実行前後を比較し、更新状態が仕様どおりであること",
             "実行ログを確認すること。\n想定外のエラーが無いこと。"])
        self.assertTrue(all(c.category == "review" for c in checks))
        self.assertTrue(all(c.verdict == REVIEW for c in checks))
        self.assertEqual(checks[0].detail, "")

    def test_review_point_requires_non_empty_string(self):
        cases_dir = self.tmp / "invalid_cases"
        cases_dir.mkdir()
        (cases_dir / "R.yaml").write_text(
            "id: R\nname: R\nreview:\n  - ''\n  - {content: mapping は不可}\n",
            encoding="utf-8")
        with self.assertRaises(ConfigError) as ctx:
            load_cases(cases_dir)
        self.assertIn("review[0]", str(ctx.exception))
        self.assertIn("review[1]", str(ctx.exception))

    def test_review_point_rejects_duplicate_content(self):
        cases_dir = self.tmp / "duplicate_cases"
        cases_dir.mkdir()
        (cases_dir / "R.yaml").write_text(
            "id: R\nname: R\nreview:\n"
            "  - 同じ確認内容\n"
            "  - 同じ確認内容\n",
            encoding="utf-8")
        with self.assertRaises(ConfigError) as ctx:
            load_cases(cases_dir)
        self.assertIn("重複", str(ctx.exception))


class TestManualReviewAssertionTypes(TmpDirCase):
    """DB/内容比較以外に manual を書いても黙って無視されないこと。"""

    def _runner(self):
        from autotest.orchestrator import CaseRunner
        output = self.tmp / "out"
        output.mkdir()
        settings = self.write_settings({"output_dir": str(output)})
        return CaseRunner(settings, self.tmp), output

    def test_exists_ok_becomes_review(self):
        runner, output = self._runner()
        (output / "a.csv").write_text("x", encoding="utf-8")
        checks = runner._assert_files({
            "files": [{"name": "存在", "exists": {"dir": "output_dir", "pattern": "*.csv", "count": 1},
                       "manual": True}]})
        self.assertEqual(checks[0].verdict, REVIEW)

    def test_log_ok_becomes_review(self):
        runner, _ = self._runner()
        check = runner._assert_log(
            {"log": {"must_not_contain": ["ERROR"], "manual": True}}, [])
        self.assertEqual(check.verdict, REVIEW)

    def test_exit_code_ok_becomes_review(self):
        from autotest.models import ExecutionInfo
        runner, _ = self._runner()
        result = CaseResult("T", "T")
        result.execution = ExecutionInfo(exit_code=0)
        check = runner._assert_exit_code(
            {"exit_code": {"expected": 0, "manual": True}}, result)
        self.assertEqual(check.verdict, REVIEW)


class TestManualReviewExcel(TmpDirCase):
    def _build(self, verdicts):
        from openpyxl import load_workbook
        from autotest.excel import build_workbook
        from autotest.models import RunResult

        run = RunResult(run_id="ut", started_at=datetime.now(), finished_at=datetime.now())
        case = CaseResult(case_id="TC", name="TC")
        for v in verdicts:
            case.checks.append(compare.CheckResult("chk-" + v, "db", v, "detail"))
        run.cases.append(case)
        out = self.tmp / "e.xlsx"
        build_workbook(run, out, {})
        return load_workbook(out)["TC"]

    def test_input_columns_added_when_review_exists(self):
        from autotest.models import REVIEW
        ws = self._build([OK, REVIEW])
        headers = {c.value for row in ws.iter_rows() for c in row
                   if c.value in ("判定結果", "確認者", "確認日")}
        self.assertEqual(headers, {"判定結果", "確認者", "確認日"})
        values = [c.value for row in ws.iter_rows() for c in row]
        self.assertIn("未確認", values)

    def test_no_input_columns_without_review(self):
        ws = self._build([OK, OK])
        headers = [c.value for row in ws.iter_rows() for c in row if c.value == "確認者"]
        self.assertEqual(headers, [], "確認待ちが無いのに記入欄を出さないこと")

    def test_guidance_note_present(self):
        from autotest.models import REVIEW
        ws = self._build([REVIEW])
        texts = [str(c.value) for row in ws.iter_rows() for c in row
                 if isinstance(c.value, str) and "未確認" in c.value and "記入" in c.value]
        self.assertTrue(texts, "確認手順の案内が出ていない")


# =============================================================================
# DB 値は整形せずそのまま（表示のための加工が判定を汚さないこと）
# =============================================================================
class TestDbValueLossless(unittest.TestCase):
    """DB にある値をそのまま扱う。丸め・切り捨て・マスクは判定より後。

    表示のための加工を判定より前に適用すると、値が違うのに一致と
    みなされる（偽 OK）。実際に mask / binary / 秒精度 / Decimal 丸めの
    4 パターンで発生していた。
    """

    def _compare(self, actual, expected, mask=None):
        from autotest.db import to_text
        a = Table("T", ["ID", "V"], [["1", to_text(actual)]], mask_columns=mask or [])
        e = Table("T", ["ID", "V"], [["1", to_text(expected)]], mask_columns=mask or [])
        return compare.compare_db_table("T", a, e, keys=["ID"])

    def test_masked_column_still_detects_difference(self):
        """マスク対象でも値が違えば NG。別人が同一視されてはいけない。"""
        result = self._compare("bob", "alice", mask=["V"])
        self.assertEqual(result.verdict, NG)

    def test_masked_values_are_hidden_in_diff(self):
        """差異は検出しつつ、差分表には生値を出さないこと。"""
        result = self._compare("bob", "alice", mask=["V"])
        shown = " ".join(str(v) for v in result.diff_table.rows[0])
        self.assertNotIn("alice", shown)
        self.assertNotIn("bob", shown)
        self.assertIn("MASKED", shown)

    def test_binary_tail_difference_detected(self):
        """先頭が同じで後半だけ異なる BLOB を同一視しないこと。"""
        head = bytes(range(16))
        self.assertEqual(self._compare(head + b"AAAA", head + b"BBBB").verdict, NG)

    def test_datetime_microseconds_preserved(self):
        import datetime as _dt
        a = _dt.datetime(2026, 8, 5, 9, 0, 0, 111111)
        b = _dt.datetime(2026, 8, 5, 9, 0, 0, 999999)
        self.assertEqual(self._compare(a, b).verdict, NG)

    def test_decimal_not_rounded(self):
        from decimal import Decimal as _Dec
        self.assertEqual(
            self._compare(_Dec("100.004"), _Dec("100.001")).verdict, NG)

    def test_identical_values_are_ok(self):
        import datetime as _dt
        self.assertEqual(
            self._compare(_dt.datetime(2026, 8, 5, 9, 0, 0, 1),
                          _dt.datetime(2026, 8, 5, 9, 0, 0, 1)).verdict, OK)
        self.assertEqual(self._compare("alice", "alice", mask=["V"]).verdict, OK)

    def test_to_text_is_lossless(self):
        import datetime as _dt
        from decimal import Decimal as _Dec
        from autotest.db import to_text
        self.assertEqual(to_text(_dt.datetime(2026, 8, 5, 9, 0, 0, 123456)),
                         "2026-08-05 09:00:00.123456")
        self.assertEqual(to_text(_Dec("100.0040")), "100.0040", "末尾 0 も落とさない")
        self.assertEqual(to_text(b"\x00\x01\xff"), "0x0001FF")
        self.assertEqual(to_text(None), "(NULL)")
        self.assertEqual(to_text("  空白  "), "  空白  ", "前後空白を残すこと")

    def test_snapshot_does_not_mask_or_format(self):
        """snapshot が返す Table は生値で、マスク列名だけを持つこと。"""
        from autotest.db import DbClient

        class FakeClient(DbClient):
            def query(self, sql, params=None):
                return ["ID", "NAME"], [["1", "alice"]]

            def execute_script(self, sql):
                pass

        table = FakeClient().snapshot({"name": "T", "sql": "SELECT * FROM T", "mask": ["NAME"]}, {})
        self.assertEqual(table.rows, [["1", "alice"]], "比較用データはマスクしない")
        self.assertEqual(table.mask_columns, ["NAME"])


class TestMaskAppliedOnlyInExcel(TmpDirCase):
    def test_excel_output_is_masked(self):
        from openpyxl import load_workbook
        from autotest.excel import build_workbook
        from autotest.models import RunResult

        run = RunResult(run_id="ut", started_at=datetime.now(), finished_at=datetime.now())
        case = CaseResult(case_id="TC", name="TC")
        case.db_after["T"] = Table("T", ["ID", "NAME"], [["1", "alice"]], mask_columns=["NAME"])
        case.checks.append(compare.CheckResult("c", "db", OK, "ok"))
        run.cases.append(case)

        out = self.tmp / "e.xlsx"
        build_workbook(run, out, {})
        values = [str(c.value) for row in load_workbook(out)["TC"].iter_rows() for c in row
                  if c.value is not None]
        self.assertNotIn("alice", values, "Excel に生値が出ている")
        self.assertIn("***MASKED***", values)


# =============================================================================
# manual: true は expected 無しでも成立すること（preflight で弾かない）
# =============================================================================
class TestManualWithoutExpected(TmpDirCase):
    def _case(self, assertions, snapshot=None, collect=None):
        from autotest.config import TestCase
        return TestCase(case_id="T", name="T", source=self.tmp / "T.yaml",
                        assertions=assertions, snapshot=snapshot, collect=collect)

    def _settings(self):
        exe = self.tmp / "b.exe"
        exe.write_text("x", encoding="utf-8")
        return self.write_settings({"work_dir": str(self.tmp / "w")},
                                   extra={"batch": {"exe_path": str(exe)}})

    def test_manual_db_without_expected_passes_preflight(self):
        from autotest.orchestrator import preflight_case
        problems = preflight_case(self._settings(),
                                  self._case({"db": [{"table": "T_ORDER", "manual": True}]},
                                             snapshot={"tables": [{"name": "T_ORDER", "sql": "SELECT 1"}]}))
        self.assertEqual(problems, [], "manual 運用が preflight で成立しない")

    def test_non_manual_db_still_requires_expected(self):
        from autotest.orchestrator import preflight_case
        problems = preflight_case(self._settings(),
                                  self._case({"db": [{"table": "T_ORDER"}]}))
        self.assertTrue(problems)
        self.assertIn("manual", " ".join(problems), "対処方法を案内すること")

    def test_manual_file_without_preview_is_rejected(self):
        from autotest.orchestrator import preflight_case
        assertions = {"files": [{"name": "result", "actual": {"dir": "work_dir", "pattern": "*.csv"},
                                  "manual": True}]}
        problems = preflight_case(self._settings(), self._case(assertions))
        self.assertIn("preview: true", " ".join(problems))

    def test_manual_file_with_expected_still_requires_preview(self):
        from autotest.orchestrator import preflight_case
        expected = self.tmp / "expected.csv"
        expected.write_text("x", encoding="utf-8")
        assertions = {"files": [{"name": "result", "actual": {"dir": "work_dir", "pattern": "*.csv"},
                                  "expected": "expected.csv", "manual": True}]}
        problems = preflight_case(self._settings(), self._case(assertions))
        self.assertIn("preview: true", " ".join(problems))

    def test_manual_file_with_matching_preview_passes(self):
        from autotest.orchestrator import preflight_case
        assertions = {"files": [{"name": "result", "actual": {"dir": "work_dir", "pattern": "*.csv"},
                                  "manual": True}]}
        collect = {"files": [{"dir": "work_dir", "pattern": "*.csv", "preview": True}]}
        problems = preflight_case(self._settings(), self._case(assertions, collect=collect))
        self.assertEqual(problems, [])


# =============================================================================
# teardown は異常経路でも実行されること
# =============================================================================
class TestTeardownOnFailure(TmpDirCase):
    def test_teardown_runs_when_case_fails_midway(self):
        """証跡採取で落ちても、setup が入れたデータを DB に残さないこと。"""
        from autotest import db as dbm, render
        from autotest.config import TestCase
        from autotest.orchestrator import CaseRunner

        executed = []

        class FakeClient:
            def query(self, sql, params=None):
                return [], []

            def execute_script(self, sql):
                executed.append(sql.strip())

            def close(self):
                pass

        exe = self.tmp / "b.exe"
        exe.write_text("x", encoding="utf-8")
        (self.tmp / "w").mkdir()
        settings = self.write_settings({"work_dir": str(self.tmp / "w"), "log_dir": str(self.tmp / "w")},
                                       extra={"batch": {"exe_path": str(exe)}})
        case = TestCase(case_id="T", name="T", source=self.tmp / "T.yaml",
                        setup={"sql": ["SETUP"]}, teardown={"sql": ["TEARDOWN"]},
                        collect={"folder_evidence": ["work_dir"]})

        saved_create, saved_listing = dbm.create_client, render.Renderer.folder_listing
        try:
            dbm.create_client = lambda *a, **k: FakeClient()
            render.Renderer.folder_listing = lambda *a, **k: (_ for _ in ()).throw(RuntimeError("撮影失敗"))
            result = CaseRunner(settings, self.tmp / "out").run(case)
        finally:
            dbm.create_client, render.Renderer.folder_listing = saved_create, saved_listing

        self.assertIn("SETUP", executed)
        self.assertIn("TEARDOWN", executed, "異常終了時に teardown が実行されていない")
        self.assertEqual(result.verdict, "ERROR")


# =============================================================================
# dry-run はフォルダを作らない / 通常実行も無関係な論理名は作らない
# =============================================================================
class TestDirectoryCreationScope(TmpDirCase):
    def _settings(self):
        exe = self.tmp / "b.exe"
        exe.write_text("x", encoding="utf-8")
        return self.write_settings(
            {"input_dir": str(self.tmp / "sb" / "in"),
             "log_dir": str(self.tmp / "sb" / "log"),
             "other_batch_dir": str(self.tmp / "OTHER" / "data")},
            extra={"batch": {"exe_path": str(exe)},
                   "folder_evidence": {"targets": ["input_dir"]}})

    def _case(self):
        from autotest.config import TestCase
        return TestCase(case_id="T", name="T", source=self.tmp / "T.yaml",
                        setup={"clean_dirs": ["input_dir"]})

    def test_unrelated_alias_not_created(self):
        """他機能のフォルダまで作らないこと（到達不能な共有で巻き添えになる）。"""
        from autotest.orchestrator import CaseRunner
        runner = CaseRunner(self._settings(), self.tmp / "out")
        used = runner._aliases_used_by(self._case())
        self.assertIn("input_dir", used)
        self.assertIn("log_dir", used)
        self.assertNotIn("other_batch_dir", used)

    def test_dry_run_creates_nothing(self):
        from autotest import db as dbm, render
        from autotest.orchestrator import CaseRunner

        saved_create, saved_listing = dbm.create_client, render.Renderer.folder_listing
        try:
            render.Renderer.folder_listing = lambda self_, d, e, a, p, m: self_.out_dir / "x.png"
            CaseRunner(self._settings(), self.tmp / "out", dry_run=True).run(self._case())
        finally:
            dbm.create_client, render.Renderer.folder_listing = saved_create, saved_listing

        self.assertFalse((self.tmp / "sb").exists(), "dry-run がフォルダを作っている")
        self.assertFalse((self.tmp / "OTHER").exists())


# =============================================================================
# setup.batches（投入後・主 batch 前に実行する前置 batch）
# =============================================================================
class TestSetupBatches(TmpDirCase):
    def _settings(self):
        exe = self.tmp / "batch.exe"
        exe.write_text("x", encoding="utf-8")
        return self.write_settings(
            {"input_dir": str(self.tmp / "in"), "log_dir": str(self.tmp / "log")},
            extra={
                "batch": {"exe_path": str(exe)},
                "batches": {"prepare": {"exe_path": str(exe)}},
            })

    def test_runs_after_input_and_replacement_in_list_order(self):
        """YAML の mapping 順ではなく、投入完了後という lifecycle を保証する。"""
        from autotest import orchestrator as orch
        from autotest.config import TestCase
        from autotest.models import ExecutionInfo

        events = []

        class FakeClient:
            def execute_script(self, sql):
                events.append("sql")

        def fake_put(*args, **kwargs):
            events.append("input_files")

        def fake_replace(*args, **kwargs):
            events.append("replace_files")
            return []

        def fake_run(settings, args, dry_run=False, batch_name=None, on_progress=None):
            events.append("batch:%s:%s" % (batch_name, args[0]))
            info = ExecutionInfo()
            info.exit_code = 0
            return info

        case = TestCase(
            case_id="T", name="T", source=self.tmp / "T.yaml",
            setup={
                "sql": ["SELECT 1"],
                "input_files": [{"src": "input.csv", "dest_dir": "input_dir"}],
                "replace_files": [{"src": "case.config", "dest_dir": "input_dir"}],
                # 同じ batch を複数回実行でき、記載順を保持する。
                "batches": [
                    {"batch": "prepare", "args": ["first"]},
                    {"batch": "prepare", "args": ["second"]},
                ],
            })

        saved_put = fsops.put_input_files
        saved_replace = fsops.replace_files
        saved_run = orch.run_batch
        try:
            fsops.put_input_files = fake_put
            fsops.replace_files = fake_replace
            orch.run_batch = fake_run
            orch.CaseRunner(self._settings(), self.tmp / "out")._setup(case, FakeClient())
        finally:
            fsops.put_input_files = saved_put
            fsops.replace_files = saved_replace
            orch.run_batch = saved_run

        self.assertEqual(events, [
            "sql", "input_files", "replace_files",
            "batch:prepare:first", "batch:prepare:second",
        ])

    def test_unexpected_setup_batch_exit_code_stops_setup(self):
        from autotest import orchestrator as orch
        from autotest.config import TestCase
        from autotest.models import ExecutionInfo

        calls = []

        def fake_run(settings, args, dry_run=False, batch_name=None, on_progress=None):
            calls.append(batch_name)
            info = ExecutionInfo()
            info.exit_code = 9
            info.stderr = "prepare failed"
            return info

        case = TestCase(
            case_id="T", name="T", source=self.tmp / "T.yaml",
            setup={"batches": ["prepare", "prepare"]})

        saved_run = orch.run_batch
        try:
            orch.run_batch = fake_run
            with self.assertRaises(ConfigError) as ctx:
                orch.CaseRunner(self._settings(), self.tmp / "out")._setup(case, object())
        finally:
            orch.run_batch = saved_run

        self.assertEqual(calls, ["prepare"], "失敗後に次の前置 batch を実行している")
        self.assertIn("prepare failed", str(ctx.exception))

    def test_preflight_checks_every_setup_batch_before_changes(self):
        from autotest.config import TestCase
        from autotest.orchestrator import preflight_case

        settings = self._settings()
        case = TestCase(
            case_id="T", name="T", source=self.tmp / "T.yaml",
            setup={"batches": [{"batch": "missing"}]})
        problems = "\n".join(preflight_case(settings, case))
        self.assertIn("setup.batches[0]", problems)
        self.assertIn("missing", problems)


# =============================================================================
# ケース定義のスキーマ検証
# =============================================================================
class TestCaseSchemaValidation(TmpDirCase):
    def _load(self, body):
        d = self.tmp / "cases"
        d.mkdir(exist_ok=True)
        (d / "x.yaml").write_text(body, encoding="utf-8")
        return load_cases(d)

    def _expect_error(self, body, fragment):
        with self.assertRaises(ConfigError) as ctx:
            self._load(body)
        self.assertIn(fragment, str(ctx.exception))

    def test_dict_without_alias_is_rejected(self):
        """実行時に KeyError で落ちるのではなく、読込時に説明すること。"""
        self._expect_error("id: A\nname: A\nsetup:\n  clean_dirs:\n    - {}\n", "alias")

    def test_args_as_string_is_rejected(self):
        """文字列を渡すと 1 文字ずつ .exe に渡ってしまう。"""
        self._expect_error('id: B\nname: B\nexecute:\n  args: "--mode daily"\n', "リストで指定")

    def test_setup_batches_after_input_files_is_valid(self):
        got = self._load(
            "id: B2\nname: B2\nsetup:\n"
            "  input_files:\n    - {src: input.csv, dest_dir: input_dir}\n"
            "  batches:\n"
            "    - {batch: prepare, args: ['--case', B2]}\n"
            "    - prepare\n")
        self.assertEqual(len(got[0].setup["batches"]), 2)

    def test_setup_batches_args_as_string_is_rejected(self):
        self._expect_error(
            'id: B3\nname: B3\nsetup:\n  batches:\n    - batch: prepare\n      args: "--case B3"\n',
            "setup.batches[0].args")

    def test_setup_batches_unknown_key_is_rejected(self):
        self._expect_error(
            "id: B4\nname: B4\nsetup:\n  batches:\n    - batch: prepare\n      arg: []\n",
            "未知の項目")

    def test_quoted_false_is_treated_as_false(self):
        """enabled: "false" を True と誤解しないこと。"""
        with self.assertRaises(ConfigError) as ctx:
            self._load('id: C\nname: C\nenabled: "false"\n')
        self.assertIn("実行可能なケースが 0 件", str(ctx.exception))

    def test_path_traversal_in_id_rejected(self):
        self._expect_error('id: "../../etc/evil"\nname: D\n', "パス区切り")

    def test_path_traversal_in_src_rejected(self):
        self._expect_error(
            'id: E\nname: E\nsetup:\n  input_files:\n    - src: "../../secret.csv"\n',
            "上位フォルダ")

    def test_unknown_key_is_rejected(self):
        """綴り間違いを黙って無視しないこと。"""
        self._expect_error("id: F\nname: F\nsetupp:\n  clean_dirs: [a]\n", "未知の項目")

    def test_misspelled_manual_inside_db_is_rejected(self):
        self._expect_error(
            "id: F2\nname: F2\nassert:\n  db:\n    - table: T\n      expected: e.csv\n      manul: true\n",
            "未知の項目")

    def test_quoted_manual_false_is_rejected(self):
        self._expect_error(
            'id: F3\nname: F3\nassert:\n  db:\n    - table: T\n      expected: e.csv\n      manual: "false"\n',
            "引用符")

    def test_file_content_without_expected_or_manual_is_rejected(self):
        self._expect_error(
            "id: F4\nname: F4\nassert:\n  files:\n    - actual: {dir: output_dir, pattern: '*.csv'}\n",
            "expected")

    def test_manual_is_allowed_for_exists_and_log(self):
        got = self._load(
            "id: F5\nname: F5\nassert:\n"
            "  files:\n    - exists: {dir: output_dir, pattern: '*.csv', count: 1}\n      manual: true\n"
            "  log:\n    must_not_contain: ['ERROR']\n    manual: true\n")
        self.assertTrue(got[0].assertions["files"][0]["manual"])

    def test_empty_wrong_assert_types_are_not_hidden_by_defaults(self):
        self._expect_error("id: F6\nname: F6\nassert: []\n", "マッピング")
        self._expect_error("id: F7\nname: F7\nassert:\n  db: {}\n", "リスト")

    def test_tags_must_be_list(self):
        self._expect_error("id: G\nname: G\ntags: 正常系\n", "リストで指定")

    def test_valid_case_loads(self):
        got = self._load('id: H\nname: H\ntags: [正常系]\nexecute:\n  args: ["--mode", "daily"]\n')
        self.assertEqual(got[0].case_id, "H")
        self.assertEqual(got[0].execute["args"], ["--mode", "daily"])


# =============================================================================
# .exe 終了後のログ書き込み待ち
# =============================================================================
class TestLogSettleWait(TmpDirCase):
    """.exe 終了直後にログを読むと末尾が欠けることへの対策。

    log4net / NLog の非同期アペンダはプロセス終了後に遅れて flush される。
    すぐ読むと「処理正常終了」が見つからず誤って NG になる。
    """

    def _settings(self, log_dir):
        return self.write_settings({"log_dir": str(log_dir)})

    def _log_dir(self):
        d = self.tmp / "log"
        d.mkdir()
        (d / "batch.log").write_text("start\n", encoding="utf-8")
        return d

    def test_waits_for_delayed_flush(self):
        import threading
        import time as _t

        d = self._log_dir()
        target = d / "batch.log"

        def delayed():
            _t.sleep(0.5)
            with target.open("a", encoding="utf-8") as f:
                f.write("処理正常終了\n")

        offsets = logs.snapshot_offsets(self._settings(d), d)
        threading.Thread(target=delayed).start()

        logs.wait_until_stable(self._settings(d), d, max_wait_sec=6,
                               min_wait_sec=1.0, stable_for_sec=0.6)
        slices = logs.collect(self._settings(d), offsets, datetime.now(), datetime.now(), log_dir=d)
        missing, _ = logs.check_keywords(slices, ["処理正常終了"], [])
        self.assertEqual(missing, [], "遅延書き込みを取りこぼしている")

    def test_returns_quickly_when_already_settled(self):
        """既に書き終わっている場合に上限まで待たないこと。"""
        d = self._log_dir()
        waited = logs.wait_until_stable(self._settings(d), d, max_wait_sec=10,
                                        min_wait_sec=0.3, stable_for_sec=0.3)
        self.assertLess(waited, 3.0, "落ち着いているのに待ちすぎている")

    def test_respects_max_wait(self):
        """書き込みが続いていても上限で打ち切ること。"""
        import threading
        import time as _t

        d = self._log_dir()
        target = d / "batch.log"
        stop = threading.Event()

        def keep_writing():
            while not stop.is_set():
                with target.open("a", encoding="utf-8") as f:
                    f.write("x\n")
                _t.sleep(0.1)

        writer = threading.Thread(target=keep_writing)
        writer.start()
        try:
            waited = logs.wait_until_stable(self._settings(d), d, max_wait_sec=1.0,
                                            min_wait_sec=0.2, stable_for_sec=0.5)
        finally:
            stop.set()
            writer.join()
        self.assertLess(waited, 3.0, "上限を超えて待ち続けている")

    def test_disabled_when_zero(self):
        d = self._log_dir()
        self.assertEqual(logs.wait_until_stable(self._settings(d), d, max_wait_sec=0), 0.0)


# =============================================================================
# 障害注入: 破損ファイルの生成
# =============================================================================
class TestCorruptFileGeneration(TmpDirCase):
    """解凍エラー等の異常系を、正常な資材から決まった手順で再現できること。

    壊れたバイナリを直接リポジトリへ置くと、中身が読めず「何がどう
    壊れているのか」が後から分からなくなる。
    """

    def _zip(self):
        import io
        import zipfile
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            z.writestr("data.csv", "ORDER_NO,AMOUNT\n" +
                       "\n".join("A%04d,%d" % (i, i * 100) for i in range(200)))
        return buf.getvalue()

    def _zip_status(self, data):
        import io
        import zipfile
        try:
            zf = zipfile.ZipFile(io.BytesIO(data))
            return "crc" if zf.testzip() else "ok"
        except Exception:
            return "broken"

    def test_truncate_breaks_zip(self):
        self.assertEqual(self._zip_status(fsops.corrupt_bytes(self._zip(), {"mode": "truncate"})),
                         "broken")

    def test_empty_breaks_zip(self):
        got = fsops.corrupt_bytes(self._zip(), {"mode": "empty"})
        self.assertEqual(got, b"")

    def test_flip_causes_crc_error(self):
        """既定は中央（データ部）。末尾だと zip は壊れたと判定しないことがある。"""
        self.assertEqual(self._zip_status(fsops.corrupt_bytes(self._zip(), {"mode": "flip"})),
                         "crc")

    def test_truncate_with_explicit_size(self):
        got = fsops.corrupt_bytes(self._zip(), {"mode": "truncate", "size": 10})
        self.assertEqual(len(got), 10)

    def test_unknown_mode_is_config_error(self):
        with self.assertRaises(ConfigError):
            fsops.corrupt_bytes(b"x", {"mode": "explode"})

    def test_put_input_files_writes_corrupted_copy(self):
        """投入時に生成され、元の資材は変更されないこと。"""
        case_dir = self.tmp / "case"
        (case_dir / "input").mkdir(parents=True)
        source = case_dir / "input" / "DATA.zip"
        original = self._zip()
        source.write_bytes(original)

        dest_dir = self.tmp / "in"
        settings = self.write_settings({"input_dir": str(dest_dir)})
        fsops.put_input_files(settings, case_dir, [
            {"src": "input/DATA.zip", "dest_dir": "input_dir",
             "corrupt": {"mode": "truncate"}}])

        self.assertEqual(source.read_bytes(), original, "元の資材を壊してはいけない")
        self.assertEqual(self._zip_status((dest_dir / "DATA.zip").read_bytes()), "broken")

    def test_shorthand_string_form(self):
        case_dir = self.tmp / "case"
        (case_dir / "input").mkdir(parents=True)
        (case_dir / "input" / "a.zip").write_bytes(self._zip())
        dest_dir = self.tmp / "in"
        settings = self.write_settings({"input_dir": str(dest_dir)})
        fsops.put_input_files(settings, case_dir, [
            {"src": "input/a.zip", "dest_dir": "input_dir", "corrupt": "empty"}])
        self.assertEqual((dest_dir / "a.zip").read_bytes(), b"")

    def test_put_input_files_copies_directory_tree(self):
        """フォルダを指定すると、フォルダ名と配下の構造を保って投入できること。"""
        case_dir = self.tmp / "case"
        source = case_dir / "input" / "DATA_SET"
        (source / "sub" / "empty").mkdir(parents=True)
        (source / "root.txt").write_text("root", encoding="utf-8")
        (source / "sub" / "child.txt").write_text("child", encoding="utf-8")

        dest_dir = self.tmp / "in"
        settings = self.write_settings({"input_dir": str(dest_dir)})
        placed = fsops.put_input_files(settings, case_dir, [
            {"src": "input/DATA_SET", "dest_dir": "input_dir"}])

        copied = dest_dir / "DATA_SET"
        self.assertEqual(placed, [copied])
        self.assertEqual((copied / "root.txt").read_text(encoding="utf-8"), "root")
        self.assertEqual((copied / "sub" / "child.txt").read_text(encoding="utf-8"), "child")
        self.assertTrue((copied / "sub" / "empty").is_dir())

    def test_put_input_directory_supports_rename_and_merge(self):
        """rename がフォルダにも効き、既存フォルダには上書きマージすること。"""
        case_dir = self.tmp / "case"
        source = case_dir / "input" / "source"
        source.mkdir(parents=True)
        (source / "same.txt").write_text("new", encoding="utf-8")
        (source / "added.txt").write_text("added", encoding="utf-8")

        dest_dir = self.tmp / "in"
        target = dest_dir / "RENAMED"
        target.mkdir(parents=True)
        (target / "same.txt").write_text("old", encoding="utf-8")
        (target / "kept.txt").write_text("kept", encoding="utf-8")
        settings = self.write_settings({"input_dir": str(dest_dir)})
        fsops.put_input_files(settings, case_dir, [
            {"src": "input/source", "dest_dir": "input_dir", "rename": "RENAMED"}])

        self.assertEqual((target / "same.txt").read_text(encoding="utf-8"), "new")
        self.assertEqual((target / "added.txt").read_text(encoding="utf-8"), "added")
        self.assertEqual((target / "kept.txt").read_text(encoding="utf-8"), "kept")

    def test_put_input_directory_rejects_corrupt(self):
        case_dir = self.tmp / "case"
        (case_dir / "input" / "folder").mkdir(parents=True)
        settings = self.write_settings({"input_dir": str(self.tmp / "in")})
        with self.assertRaises(ConfigError) as ctx:
            fsops.put_input_files(settings, case_dir, [
                {"src": "input/folder", "dest_dir": "input_dir", "corrupt": "empty"}])
        self.assertIn("フォルダ", str(ctx.exception))


class TestKeepEnvOption(TmpDirCase):
    """--keep-env は調査のため後始末を行わないこと。"""

    def test_teardown_skipped_and_reported(self):
        from autotest import db as dbm
        from autotest.config import TestCase
        from autotest.orchestrator import CaseRunner

        executed = []

        class FakeClient:
            def query(self, sql, params=None):
                return [], []

            def execute_script(self, sql):
                executed.append(sql.strip())

            def close(self):
                pass

        exe = self.tmp / "b.exe"
        exe.write_text("x", encoding="utf-8")
        settings = self.write_settings({"log_dir": str(self.tmp)},
                                       extra={"batch": {"exe_path": str(exe)}})
        case = TestCase(case_id="T", name="T", source=self.tmp / "T.yaml",
                        setup={"sql": ["SETUP"]}, teardown={"sql": ["TEARDOWN"]})

        saved = dbm.create_client
        try:
            dbm.create_client = lambda *a, **k: FakeClient()
            CaseRunner(settings, self.tmp / "out", keep_env=True).run(case)
        finally:
            dbm.create_client = saved

        self.assertIn("SETUP", executed)
        self.assertNotIn("TEARDOWN", executed, "--keep-env なのに teardown が走っている")
