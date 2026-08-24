# -*- coding: utf-8 -*-
"""人工確認の入力検証と finalize 用データの回帰テスト。"""

import shutil
import sys
import tempfile
import unittest
from datetime import datetime
from pathlib import Path
from types import SimpleNamespace

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from openpyxl import load_workbook  # noqa: E402

from autotest import review  # noqa: E402
from autotest import result_store  # noqa: E402
from autotest.config import ConfigError  # noqa: E402
from autotest.config import Settings  # noqa: E402
from autotest.excel import build_workbook  # noqa: E402
from autotest.models import NG, OK, REVIEW, CaseResult, CheckResult, RunResult  # noqa: E402


class TestFinalizeReview(unittest.TestCase):
    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp(prefix="autotest_review_"))

    def tearDown(self):
        shutil.rmtree(str(self.tmp), ignore_errors=True)

    def _run(self, extra_verdict=None):
        run = RunResult(run_id="r", started_at=datetime(2026, 8, 15, 10, 0, 0),
                        finished_at=datetime(2026, 8, 15, 10, 1, 0))
        case = CaseResult("TC001", "確認ケース")
        case.checks.append(CheckResult("DB確認", "db", REVIEW, "目視確認"))
        if extra_verdict:
            case.checks.append(CheckResult("自動NG", "file", extra_verdict, "不一致"))
        run.cases.append(case)
        return run

    def _filled_excel(self, run, result="OK", reviewer="山田", confirmed="2026-08-15"):
        path = self.tmp / "review.xlsx"
        build_workbook(run, path, {})
        wb = load_workbook(str(path))
        ws = wb["TC001"]
        header = next(row for row in range(1, ws.max_row + 1)
                      if ws.cell(row, 1).value == "No" and ws.cell(row, 6).value == "確認結果")
        ws.cell(header + 1, 6, result)
        ws.cell(header + 1, 7, reviewer)
        ws.cell(header + 1, 8, confirmed)
        wb.save(str(path))
        return path

    def test_ok_confirmation_finalizes_without_overwriting_auto_verdict(self):
        run = self._run()
        confirmations = review.apply_excel_confirmations(run, self._filled_excel(run))
        check = run.cases[0].checks[0]
        self.assertEqual(check.verdict, REVIEW, "自動判定は監査情報として残すこと")
        self.assertEqual(check.confirmation_result, OK)
        self.assertEqual(check.confirmation_by, "山田")
        self.assertEqual(check.confirmation_at, "2026-08-15")
        self.assertEqual(run.verdict, OK)
        self.assertEqual(len(confirmations), 1)

    def test_human_ng_makes_final_verdict_ng(self):
        run = self._run()
        review.apply_excel_confirmations(run, self._filled_excel(run, result="NG"))
        self.assertEqual(run.verdict, NG)

    def test_automatic_ng_cannot_be_overridden(self):
        run = self._run(extra_verdict=NG)
        review.apply_excel_confirmations(run, self._filled_excel(run, result="OK"))
        self.assertEqual(run.verdict, NG)
        self.assertEqual(run.cases[0].checks[1].verdict, NG)

    def test_missing_reviewer_is_rejected(self):
        run = self._run()
        with self.assertRaises(ConfigError) as ctx:
            review.apply_excel_confirmations(run, self._filled_excel(run, reviewer=""))
        self.assertIn("確認者", str(ctx.exception))

    def test_invalid_result_is_rejected(self):
        run = self._run()
        with self.assertRaises(ConfigError) as ctx:
            review.apply_excel_confirmations(run, self._filled_excel(run, result="PASS"))
        self.assertIn("OK または NG", str(ctx.exception))

    def test_uncollected_manual_case_blocks_finalize(self):
        run = self._run()
        path = self._filled_excel(run)
        run.manual_pending = ["TC008"]
        with self.assertRaises(ConfigError) as ctx:
            review.apply_excel_confirmations(run, path)
        self.assertIn("未採取", str(ctx.exception))

    def test_excel_from_another_run_is_rejected(self):
        run = self._run()
        path = self._filled_excel(run)
        run.run_id = "different-run"
        with self.assertRaises(ConfigError) as ctx:
            review.apply_excel_confirmations(run, path)
        self.assertIn("実行ID", str(ctx.exception))

    def test_single_run_report_can_be_finalized(self):
        source = self._run()
        report_run = result_store.merge_runs([source], [source.run_id])
        path = self._filled_excel(report_run)
        review.apply_excel_confirmations(source, path)
        self.assertEqual(source.verdict, OK)

    def test_final_workbook_keeps_confirmation_columns(self):
        run = self._run()
        review.apply_excel_confirmations(run, self._filled_excel(run))
        out = self.tmp / "final.xlsx"
        build_workbook(run, out, {})
        ws = load_workbook(str(out))["TC001"]
        values = [cell.value for row in ws.iter_rows() for cell in row]
        self.assertIn("要確認", values, "元の自動判定を残すこと")
        self.assertIn("山田", values)
        self.assertIn("人工確認済み", " ".join(str(v) for v in values if v))

    def test_cli_finalize_writes_excel_and_audit_json(self):
        from autotest.cli import _finalize

        out_dir = self.tmp / "output"
        run_dir = out_dir / "r"
        run_dir.mkdir(parents=True)
        run = self._run()
        result_store.save_case_result(run.cases[0], run_dir)
        result_store.save_run_meta(run, run_dir)
        filled = self._filled_excel(run)
        final = self.tmp / "final.xlsx"
        settings = Settings(
            raw={"batch": {"exe_path": "dummy.exe"},
                 "database": {"server": "s", "database": "d"},
                 "paths": {}, "excel": {}},
            source=self.tmp / "settings.yaml", project_root=self.tmp)
        args = SimpleNamespace(runs=["r"], excel=str(filled), out=str(final))
        self.assertEqual(_finalize(args, settings, out_dir), 0)
        self.assertTrue(final.exists())
        self.assertTrue(final.with_suffix(".review.json").exists())


if __name__ == "__main__":
    unittest.main()
