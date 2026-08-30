# -*- coding: utf-8 -*-
"""Excel -> Layout TXT 生成ツールの回帰テスト。"""

import csv
import io
import shutil
import sys
import tarfile
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from autotest.layout_txt import (  # noqa: E402
    LayoutTxtError,
    DEFAULT_EMPTY_VALUE,
    generate_layout_txt,
    generate_ocr_value,
    main as layout_main,
    read_layout_fields,
    render_form_tif_payload,
    render_form_txt_text,
    resolve_form_filename_stem,
    save_layout_default_values,
)
from autotest.layout_tar import PackageItem, base_name_from_front  # noqa: E402


class LayoutWorkbookCase(unittest.TestCase):
    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp(prefix="autotest_layout_txt_"))
        self.book = self.tmp / "layout.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "帳票定義"
        ws.append(["説明行"])
        headers = ["予備A", "FORM_ID", "LAYOUT_ID", "CYOUHYOU_NAME", "LAYOUT_ID",
                   "GROUP_ID", "GROUP_NAME", "ITEM_NAME", "ELEMENT_DATA_TYPE_NAME",
                   "ELEMENT_IME_NAME", "MAX_NUM_DIGITS", "ELEMENT_ID", "CONTROL_ID",
                   "入力属性", "入力規則", "補足", "出力例"]
        ws.append(headers)
        ws.append([None, 1001, 1, "帳票A", 1, 1, "基本", "傷病名1",
                   "文字列", "ひらがな", 100, 1001, 1001])
        ws["C3"].number_format = "00000"
        ws["N3"] = "必須"
        ws["O3"] = "全角100文字以内"
        ws["P3"] = "テスト補足"
        ws["Q3"] = "傷病名サンプル"
        ws.append([None, 1001, 1, "帳票A", 1, 2, "初診", "退院区分1",
                   "選択肢", "数値のみ", "NULL", 1002, 1002])
        ws["C4"].number_format = "00000"
        ws.append([None, 1001, 1, "帳票A", 1, 2, "初診", "複数選択",
                   "チェックボックス", "全タイプ", "NULL", 1003, 1003])
        ws["C5"].number_format = "00000"
        ws.append([None, 1001, 1, "帳票A", 1, 2, "初診", "生年月日",
                   "日付", "数値のみ", "NULL", 1004, 1004])
        ws["C6"].number_format = "00000"
        ws.append([None, 1001, 1, "帳票A", 1, 2, "初診", "入院期間1",
                   "日付 (From)", "数値のみ", "NULL", 1005, 1005])
        ws["C7"].number_format = "00000"
        ws.append([None, 1001, 1, "帳票A", 1, 2, "初診", "退院期間1",
                   "日付 (To)", "数値のみ", "NULL", 1006, 1006])
        ws["C8"].number_format = "00000"
        ws.append([None, 1001, 1, "帳票A", 1, 2, "初診", "処置日",
                   "カレンダー", "数値のみ", "NULL", 1007, 1007])
        ws["C9"].number_format = "00000"
        ws.append([None, 4001, 2, "全網羅帳票", 2, 1, "基本", "網羅日付1",
                   "日付", "数値のみ", "NULL", 2001, 2001])
        ws["C10"].number_format = "00000"
        ws.append([None, 4001, 2, "全網羅帳票", 2, 1, "基本", "網羅日付2",
                   "日付", "数値のみ", "NULL", 2002, 2002])
        ws["C11"].number_format = "00000"
        ws.append([None, 4001, 2, "全網羅帳票", 2, 1, "基本", "網羅日付3",
                   "日付", "数値のみ", "NULL", 2003, 2003])
        ws["C12"].number_format = "00000"
        ws.append([None, 4001, 2, "全網羅帳票", 2, 1, "基本", "網羅日付4",
                   "日付", "数値のみ", "NULL", 2004, 2004])
        ws["C13"].number_format = "00000"
        wb.save(str(self.book))

    def tearDown(self):
        shutil.rmtree(str(self.tmp), ignore_errors=True)


class TestReadLayoutFields(LayoutWorkbookCase):
    def test_auto_detects_header_and_identifier_columns(self):
        fields, sheet, header, columns = read_layout_fields(self.book)
        self.assertEqual(sheet, "帳票定義")
        self.assertEqual(header, 2)
        self.assertEqual(columns["form_id"], 2)
        self.assertEqual(columns["layout_id"], 3)
        self.assertEqual(columns["field_id"], 12)
        self.assertEqual(columns["data_type"], 9)
        self.assertEqual(columns["ime"], 10)
        self.assertEqual(columns["max_digits"], 11)
        self.assertEqual([field.form_id for field in fields],
                         ["1001"] * 52 + ["4001"] * 4)
        self.assertEqual([field.layout_id for field in fields],
                         ["00001"] * 52 + ["00002"] * 4)
        self.assertEqual([field.field_id for field in fields],
                         [str(index) for index in range(1, 53)] +
                         [str(index) for index in range(1, 5)])
        self.assertEqual([field.value for field in fields[:6]],
                         ["傷病名1", "1", "1", "5/8/6/1", "5/8/6/1",
                          "5/8/6/1"])
        self.assertEqual([field.value for field in fields[6:52]],
                         ["5/8/6/1"] * 46)
        self.assertEqual([field.value for field in fields[52:]],
                         ["5/8/6/1", "/2026/6/1", "5/2026/6/1",
                          "5/8/6/1"])
        self.assertEqual([field.attribute_flag for field in fields],
                         ["0"] * 56)
        self.assertEqual(
            [field.coordinates for field in fields],
            (["0,0,0,%d" % index for index in range(1, 53)] +
             ["0,0,0,%d" % index for index in range(1, 5)]))

    def test_column_can_be_selected_by_header_or_letter(self):
        fields, _sheet, _header, columns = read_layout_fields(
            self.book, form_column="B", layout_column="C", field_column="ELEMENT_ID",
            data_type_column="ELEMENT_DATA_TYPE_NAME", ime_column="J",
            max_digits_column="MAX_NUM_DIGITS")
        self.assertEqual(len(fields), 56)
        self.assertEqual(columns["field_id"], 12)

    def test_optional_columns_are_appended_when_present(self):
        fields, _sheet, _header, columns = read_layout_fields(self.book)
        self.assertEqual(columns["input_attribute"], 14)
        self.assertEqual(columns["input_rule"], 15)
        self.assertEqual(columns["notes"], 16)
        self.assertEqual(columns["output_example"], 17)
        self.assertEqual(fields[0].input_attribute, "必須")
        self.assertEqual(fields[0].input_rule, "全角100文字以内")
        self.assertEqual(fields[0].notes, "テスト補足")
        self.assertEqual(fields[0].output_example, "傷病名サンプル")
        self.assertEqual(fields[1].input_attribute, "")

    def test_existing_default_value_is_used_instead_of_generated_value(self):
        wb = load_workbook(str(self.book))
        ws = wb["帳票定義"]
        ws["R2"] = "默认值"
        ws["R3"] = "Excel保存値"
        wb.save(str(self.book))
        wb.close()

        fields, _sheet, _header, columns = read_layout_fields(self.book)

        self.assertEqual(columns["default_value"], 18)
        self.assertEqual(fields[0].value, "Excel保存値")
        self.assertEqual(fields[1].value, "1")

    def test_default_value_column_can_be_ignored(self):
        wb = load_workbook(str(self.book))
        ws = wb["帳票定義"]
        ws["R2"] = "OCR_DEFAULT_VALUE"
        ws["R3"] = "使用しない値"
        wb.save(str(self.book))
        wb.close()

        fields, _sheet, _header, columns = read_layout_fields(
            self.book, default_value_column="none")

        self.assertNotIn("default_value", columns)
        self.assertEqual(fields[0].value, "傷病名1")

    def test_saved_default_values_are_read_on_next_load(self):
        path, column, count = save_layout_default_values(
            self.book,
            {"3:1": "画面修正値", "4:1": "9", "5:1": ""},
            sheet_name="帳票定義")

        self.assertEqual(path, self.book)
        self.assertEqual(column, 18)
        self.assertEqual(count, 3)
        wb = load_workbook(str(self.book), data_only=True)
        ws = wb["帳票定義"]
        self.assertEqual(ws["R2"].value, "OCR_DEFAULT_VALUE")
        self.assertEqual(ws["R3"].value, "画面修正値")
        self.assertEqual(ws["R4"].value, "9")
        self.assertEqual(ws["R5"].value, DEFAULT_EMPTY_VALUE)
        wb.close()

        fields, _sheet, _header, _columns = read_layout_fields(self.book)
        self.assertEqual([field.value for field in fields[:3]], ["画面修正値", "9", ""])

    def test_calendar_defaults_share_one_multiline_excel_cell(self):
        expected = ["手動値%d" % index for index in range(1, 47)]
        values = {"9:%d" % index: value
                  for index, value in enumerate(expected, 1)}

        _path, column, count = save_layout_default_values(
            self.book, values, sheet_name="帳票定義")

        self.assertEqual(column, 18)
        self.assertEqual(count, 1)
        wb = load_workbook(str(self.book), data_only=True)
        ws = wb["帳票定義"]
        self.assertEqual(ws["R9"].value.split("\n"), expected)
        self.assertTrue(ws["R9"].alignment.wrap_text)
        wb.close()

        fields, _sheet, _header, _columns = read_layout_fields(self.book)
        self.assertEqual([field.value for field in fields[6:52]], expected)

    def test_calendar_default_count_mismatch_is_rejected(self):
        wb = load_workbook(str(self.book))
        ws = wb["帳票定義"]
        ws["R2"] = "OCR_DEFAULT_VALUE"
        ws["R9"] = "1件目\n2件目"
        wb.save(str(self.book))
        wb.close()

        with self.assertRaises(LayoutTxtError) as ctx:
            read_layout_fields(self.book)
        self.assertIn("1件または46件", str(ctx.exception))

    def test_optional_columns_can_be_disabled(self):
        fields, _sheet, _header, columns = read_layout_fields(
            self.book, input_attribute_column="none", input_rule_column="none",
            notes_column="none", output_example_column="none")
        self.assertNotIn("input_attribute", columns)
        self.assertEqual(fields[0].input_attribute, "")

    def test_one_calendar_row_expands_to_46_independent_items(self):
        calendar_book = self.tmp / "one_calendar.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append([
            "FORM_ID", "LAYOUT_ID", "CYOUHYOU_NAME", "予備1", "予備2",
            "予備3", "予備4", "ITEM_NAME", "ELEMENT_DATA_TYPE_NAME",
            "ELEMENT_IME_NAME", "MAX_NUM_DIGITS", "ELEMENT_ID",
        ])
        ws.append([
            4001, 1, "カレンダー帳票", "", "", "", "",
            "カレンダー", "カレンダー", "数値のみ", "NULL", 3001,
        ])
        wb.save(str(calendar_book))

        fields, _sheet, _header, _columns = read_layout_fields(calendar_book)

        self.assertEqual(len(fields), 46)
        self.assertEqual(
            [field.field_id for field in fields],
            [str(index) for index in range(1, 47)])
        self.assertEqual(set(field.item_name for field in fields), {"カレンダー"})
        self.assertEqual([field.occurrence_index for field in fields], list(range(1, 47)))
        self.assertEqual(fields[0].row_label, "2 (1/46)")
        self.assertEqual(fields[-1].row_label, "2 (46/46)")
        self.assertEqual(
            [field.value for field in fields[:4]],
            ["5/8/6/1", "/2026/6/1", "5/2026/6/1",
             "5/8/6/1|5/8/6/2"])
        self.assertEqual(sum("|" in field.value for field in fields), 11)
        self.assertEqual(len(set(field.coordinates for field in fields)), 46)
        self.assertEqual(fields[0].coordinates, "0,0,0,1")
        self.assertEqual(fields[-1].coordinates, "0,0,0,46")

        output = self.tmp / "46_calendar_output"
        generate_layout_txt(
            calendar_book, output, error_patterns="none", generate_tif=False)
        txt_values = next(csv.reader([
            (output / "4001.txt").read_text(encoding="cp932").strip()]))
        self.assertEqual((len(txt_values) - 2) // 4, 46)
        self.assertEqual(txt_values[3::4], [field.value for field in fields])

    def test_missing_sheet_is_reported_with_candidates(self):
        with self.assertRaises(LayoutTxtError) as ctx:
            read_layout_fields(self.book, sheet_name="存在しない")
        self.assertIn("帳票定義", str(ctx.exception))


class TestValueGeneration(unittest.TestCase):
    def test_supported_ime_rules(self):
        cases = (
            ("文字列", "ひらがな", 5, "あいうえお"),
            ("文字列", "半角カタカナ", 5, "ｱｲｳｴｵ"),
            ("文字列", "半角英数", 6, "Abc123"),
            ("文字列", "数値のみ", 4, "1234"),
            ("文字列", "小数点数値", 6, "1.0"),
            ("文字列", "全タイプ", 4, "あA1ｱ"),
            ("チェックボックス", "NULL", None, "1"),
            ("選択肢", "数値のみ", None, "1"),
        )
        for data_type, ime, digits, expected in cases:
            self.assertEqual(generate_ocr_value(data_type, ime, digits), expected)

    def test_checkbox_and_choice_values_cycle_within_allowed_values(self):
        self.assertEqual(
            [generate_ocr_value("チェックボックス", "NULL", None,
                                selection_index=index)
             for index in range(4)],
            ["1", "0", "1", "0"])
        self.assertEqual(
            [generate_ocr_value("選択肢", "数値のみ", None,
                                selection_index=index)
             for index in range(6)],
            ["1", "2", "1.2", "1", "2", "1.2"])

    def test_calendar_item_can_contain_multiple_dates(self):
        self.assertEqual(
            generate_ocr_value("カレンダー", "数値のみ", None,
                               date_mode="multiple"),
            "5/8/6/1|5/8/6/2")

    def test_name_value_obeys_ime_rule_instead_of_item_name(self):
        self.assertEqual(
            generate_ocr_value("氏名", "数値のみ", 10, item_name="患者氏名"),
            "1234567890")
        self.assertEqual(
            generate_ocr_value("氏名", "小数点数値", 10, item_name="患者氏名"),
            "1.0")
        self.assertEqual(
            generate_ocr_value("氏名", "ひらがな", 10, item_name="患者氏名"),
            "あいうえお")

    def test_max_and_over_profiles_hit_boundary(self):
        self.assertEqual(generate_ocr_value("文字列", "ひらがな", 7, "max"), "あいうえおあい")
        self.assertEqual(len(generate_ocr_value("文字列", "ひらがな", 7, "over")), 8)
        self.assertEqual(generate_ocr_value("文字列", "小数点数値", 5, "max"), "1.234")
        self.assertEqual(len(generate_ocr_value("文字列", "数値のみ", 5, "over")), 6)

    def test_string_field_uses_item_name_as_normal_value(self):
        self.assertEqual(
            generate_ocr_value("文字列", "ひらがな", 100, item_name="傷病名1"),
            "傷病名1")
        self.assertEqual(
            generate_ocr_value("文字列", "ひらがな", 3, item_name="医療機関名称1"),
            "医療機")
        self.assertEqual(
            generate_ocr_value("文字列", "ひらがな", 7, "max", item_name="傷病名1"),
            "傷病名1傷病名")

    def test_four_date_notations(self):
        expected = ("5/8/6/1", "/2026/6/1", "5/2026/6/1", "5/8/6/1")
        for index, value in enumerate(expected):
            self.assertEqual(
                generate_ocr_value("日付", "数値のみ", None,
                                   date_mode="cycle", date_index=index),
                value)
        self.assertEqual(
            generate_ocr_value("日付", "数値のみ", None, date_mode="multiple"),
            "5/8/6/1")
        self.assertEqual(
            generate_ocr_value("日付 (From)", "数値のみ", None,
                               date_mode="multiple"),
            "5/8/6/1")
        self.assertEqual(
            generate_ocr_value("日付 (To)", "数値のみ", None,
                               date_mode="multiple"),
            "5/8/6/1")


class TestGenerateLayoutTxt(LayoutWorkbookCase):
    def test_package_tif_renderer_creates_front_and_optional_back_payloads(self):
        from PIL import Image

        fields, _sheet, _header, _columns = read_layout_fields(self.book)
        front = render_form_tif_payload(
            fields, "1001", selected_rows=[3],
            field_overrides={3: {"value": "正面テスト"}}, side="front")
        back = render_form_tif_payload(
            fields, "1001", selected_rows=[3], side="back")
        front_text = render_form_txt_text(
            fields, "1001", selected_rows=[3],
            field_overrides={3: {"value": "正面テスト"}})

        self.assertNotEqual(front, back)
        for payload in (front, back):
            with Image.open(io.BytesIO(payload)) as image:
                self.assertEqual(image.format, "TIFF")
                self.assertEqual(image.size, (1654, 2339))
        self.assertEqual(
            next(csv.reader([front_text.strip()])),
            ["1001", "1", "1", "正面テスト", "0", "0,0,0,1"])

        with self.assertRaises(LayoutTxtError):
            render_form_tif_payload(fields, "1001", side="unknown")

    def test_selected_form_id_and_custom_filename(self):
        out = self.tmp / "selected"
        result = generate_layout_txt(
            self.book, out, selected_form_ids=["4001"],
            error_patterns="none", filename_template="CUSTOM_FORM",
            generate_tif=False)
        self.assertEqual(result.form_count, 1)
        self.assertEqual(result.field_count, 4)
        self.assertEqual([path.name for path in result.txt_files], ["CUSTOM_FORM.txt"])
        self.assertTrue((out / "CUSTOM_FORM.txt").is_file())

    def test_package_filename_uses_the_same_configured_template(self):
        self.assertEqual(
            resolve_form_filename_stem(
                "DATA_{source}_{form_id}_{pattern}_{seq:02d}F.txt",
                "1001", self.book),
            "DATA_layout_1001_normal_01F")

        base = base_name_from_front(Path(
            resolve_form_filename_stem("DATA_{form_id}F", "1001", self.book)
            + ".txt"))
        item = PackageItem(
            base_name=base, form_id="1001", front_image_bytes=b"F",
            back_image_bytes=b"R", front_recognition_text="FRONT")
        self.assertEqual(item.front_recognition_name, "DATA_1001F.txt")
        self.assertEqual(item.back_recognition_name, "DATA_1001R.txt")

    def test_selected_rows_and_screen_edits_are_written(self):
        out = self.tmp / "edited"
        result = generate_layout_txt(
            self.book, out, selected_form_ids=["1001"],
            selected_rows=[3, 4],
            field_overrides={
                3: {"field_id": "9901", "value": "画面修正値",
                    "attribute_flag": "2", "coordinates": "1,2,3,4"},
                4: {"value": "9"},
            },
            error_patterns="none", filename_template="EDITED",
            generate_tif=False)
        self.assertEqual(result.field_count, 2)
        text = (out / "EDITED.txt").read_text(encoding="cp932")
        self.assertEqual(
            text,
            '"1001","1","9901","画面修正値","2","1,2,3,4",'
            '"2","9","0","0,0,0,2"\n')

    def test_one_expanded_calendar_occurrence_can_be_edited(self):
        out = self.tmp / "calendar_edit"
        result = generate_layout_txt(
            self.book, out, selected_form_ids=["1001"],
            selected_rows=["9:2"],
            field_overrides={"9:2": {"value": "5/8/6/10|5/8/6/11"}},
            error_patterns="none", filename_template="CALENDAR_EDIT",
            generate_tif=False)

        self.assertEqual(result.field_count, 1)
        values = next(csv.reader([
            (out / "CALENDAR_EDIT.txt").read_text(encoding="cp932").strip()]))
        self.assertEqual(
            values,
            ["1001", "1", "8", "5/8/6/10|5/8/6/11", "0", "0,0,0,8"])

    def test_tar_only_contains_txt_and_tif_without_loose_files(self):
        out = self.tmp / "tar_only"
        result = generate_layout_txt(
            self.book, out, selected_form_ids=["1001"],
            error_patterns="none", filename_template="FORM_{form_id}",
            create_tar=True, tar_name="PACKAGE_{form_id}", tar_only=True)
        self.assertEqual(result.txt_files, [])
        self.assertEqual(result.tif_files, [])
        self.assertEqual(result.tar_file.name, "PACKAGE_1001.tar")
        self.assertEqual(result.archive_members, ["FORM_1001.txt", "FORM_1001.tif"])
        self.assertFalse((out / "FORM_1001.txt").exists())
        self.assertFalse((out / "FORM_1001.tif").exists())
        with tarfile.open(str(result.tar_file), "r") as archive:
            self.assertEqual(archive.getnames(), result.archive_members)
            txt = archive.extractfile("FORM_1001.txt").read().decode("cp932")
            tif = archive.extractfile("FORM_1001.tif").read()
        self.assertTrue(txt.startswith('"1001","1"'))
        self.assertIn(tif[:4], (b"II*\x00", b"MM\x00*"))

    def test_unknown_selected_form_id_is_rejected(self):
        with self.assertRaises(LayoutTxtError) as ctx:
            generate_layout_txt(
                self.book, self.tmp / "missing_form",
                selected_form_ids=["9999"], generate_tif=False)
        self.assertIn("9999", str(ctx.exception))

    def test_cli_can_select_form_and_output_only_tar(self):
        out = self.tmp / "cli_tar"
        code = layout_main([
            str(self.book), "--out-dir", str(out),
            "--form-id", "1001", "--error-patterns", "none",
            "--filename-template", "CLI_{form_id}", "--no-tif",
            "--tar-only", "--tar-name", "CLI_PACKAGE",
        ])
        self.assertEqual(code, 0)
        self.assertEqual([path.name for path in out.iterdir()], ["CLI_PACKAGE.tar"])
        with tarfile.open(str(out / "CLI_PACKAGE.tar"), "r") as archive:
            self.assertEqual(archive.getnames(), ["CLI_1001.txt"])

    def test_splits_one_raw_file_per_form_with_crlf_and_cp932(self):
        out = self.tmp / "out"
        result = generate_layout_txt(self.book, out, generate_tif=False)
        self.assertEqual(result.form_count, 2)
        self.assertEqual(result.field_count, 56)
        self.assertEqual(result.pattern_count, 25)
        self.assertEqual(len(result.txt_files), 25)
        self.assertEqual(result.txt_files[0].name, "1001.txt")
        self.assertEqual(result.txt_files[1].name, "4001_01_normal.txt")
        self.assertEqual(result.txt_files[-1].name, "4001_24_form_id_empty.txt")
        self.assertEqual(result.tif_files, [])

        text = (out / "1001.txt").read_bytes().decode("cp932")
        values = next(csv.reader([text.strip()]))
        self.assertEqual(values[:2], ["1001", "1"])
        self.assertEqual(values[2::4],
                         [str(index) for index in range(1, 53)])
        self.assertEqual(values[3:27:4],
                         ["傷病名1", "1", "1", "5/8/6/1", "5/8/6/1",
                          "5/8/6/1"])
        self.assertEqual(values[3::4][6:], ["5/8/6/1"] * 46)
        self.assertEqual(
            values[5::4],
            ["0,0,0,%d" % index for index in range(1, 53)])

        coverage = (out / "4001_01_normal.txt").read_bytes().decode("cp932")
        coverage_values = next(csv.reader([coverage.strip()]))
        self.assertEqual(coverage_values[:2], ["4001", "1"])
        self.assertEqual(coverage_values[2::4], ["1", "2", "3", "4"])
        self.assertEqual(coverage_values[3::4],
                         ["5/8/6/1", "/2026/6/1", "5/2026/6/1",
                          "5/8/6/1"])
        self.assertEqual(
            coverage_values[5::4],
            ["0,0,0,%d" % index for index in range(1, 5)])

    def test_count_and_element_id_patterns_change_actual_structure(self):
        out = self.tmp / "patterns"
        generate_layout_txt(self.book, out, generate_tif=False)

        zero = next(csv.reader([
            (out / "4001_02_count_zero.txt").read_text(encoding="cp932").strip()]))
        missing = next(csv.reader([
            (out / "4001_03_count_missing.txt").read_text(encoding="cp932").strip()]))
        extra = next(csv.reader([
            (out / "4001_04_count_extra.txt").read_text(encoding="cp932").strip()]))
        unknown = next(csv.reader([
            (out / "4001_05_element_id_unknown.txt").read_text(encoding="cp932").strip()]))
        duplicate = next(csv.reader([
            (out / "4001_06_element_id_duplicate.txt").read_text(encoding="cp932").strip()]))

        self.assertEqual(len(zero), 2)
        self.assertEqual((len(missing) - 2) // 4, 3)
        self.assertEqual((len(extra) - 2) // 4, 5)
        self.assertEqual(missing[-2], "1")
        self.assertEqual(extra[-2], "1")
        self.assertNotIn(unknown[2], ("1", "2", "3", "4"))
        self.assertEqual(duplicate[2], duplicate[6])
        extra_coordinates = extra[5::4]
        self.assertEqual(len(extra_coordinates), len(set(extra_coordinates)))

        duplicate_all = next(csv.reader([
            (out / "4001_09_count_duplicate_all.txt").read_text(
                encoding="cp932").strip()]))
        duplicate_all_coordinates = duplicate_all[5::4]
        self.assertEqual(
            len(duplicate_all_coordinates), len(set(duplicate_all_coordinates)))

    def test_mixed_attribute_uses_quoted_comma_as_one_value(self):
        out = self.tmp / "mixed"
        generate_layout_txt(self.book, out, generate_tif=False)
        raw = (out / "4001_08_attribute_mixed_1_2.txt").read_text(encoding="cp932")
        values = next(csv.reader([raw.strip()]))
        self.assertEqual(values[4::4], ["0", "1", "2", "1,2"])
        self.assertIn('"1,2"', raw)

    def test_filename_template_and_matching_tif(self):
        from PIL import Image

        out = self.tmp / "named"
        result = generate_layout_txt(
            self.book, out, error_patterns="none",
            filename_template="TEST_{form_id}_{seq:02d}")
        self.assertEqual(
            [path.name for path in result.txt_files],
            ["TEST_1001_01.txt", "TEST_4001_01.txt"])
        self.assertEqual(
            [path.name for path in result.tif_files],
            ["TEST_1001_01.tif", "TEST_4001_01.tif"])
        for txt_path, tif_path in zip(result.txt_files, result.tif_files):
            self.assertEqual(txt_path.stem, tif_path.stem)
            with Image.open(str(tif_path)) as image:
                self.assertEqual(image.format, "TIFF")
                self.assertEqual(image.size, (1654, 2339))

    def test_single_labeled_file_contains_each_form(self):
        out = self.tmp / "single"
        result = generate_layout_txt(
            self.book, out, split_by_form=False, output_format="labeled", encoding="utf-8",
            error_patterns="none", generate_tif=False)
        self.assertEqual(len(result.files), 1)
        text = result.files[0].read_text(encoding="utf-8")
        self.assertIn("FormID=1001", text)
        self.assertIn('"FieldID=1"', text)
        self.assertIn("FormID=4001", text)
        self.assertIn('"OCRText=5/2026/6/1"', text)
        self.assertEqual(len(text.splitlines()), 2, "1 行が 1 帳票であること")

    def test_existing_file_is_not_overwritten_without_permission(self):
        out = self.tmp / "existing"
        out.mkdir()
        target = out / "1001.txt"
        target.write_text("保護", encoding="utf-8")
        with self.assertRaises(LayoutTxtError) as ctx:
            generate_layout_txt(self.book, out, generate_tif=False)
        self.assertIn("上書きしません", str(ctx.exception))
        self.assertEqual(target.read_text(encoding="utf-8"), "保護")
        self.assertFalse((out / "4001_01_normal.txt").exists(),
                         "失敗時に一部だけ生成してはいけない")

    def test_overwrite_replaces_existing_file(self):
        out = self.tmp / "overwrite"
        out.mkdir()
        target = out / "1001.txt"
        target.write_text("old", encoding="utf-8")
        generate_layout_txt(self.book, out, overwrite=True, generate_tif=False)
        self.assertNotEqual(target.read_text(encoding="cp932"), "old")

    def test_tsv_contains_one_row_per_field(self):
        out = self.tmp / "tsv"
        result = generate_layout_txt(
            self.book, out, output_format="tsv", encoding="utf-8", generate_tif=False)
        rows = result.files[0].read_text(encoding="utf-8").splitlines()
        self.assertEqual(rows[0], "FormID\tTargetPresence\tFieldID\tOCRText\tAttributeFlag\tCoordinates")
        self.assertEqual(len(rows), 53)


if __name__ == "__main__":
    unittest.main()
