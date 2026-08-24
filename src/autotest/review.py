"""Excel に記入された人工確認を検証し、最終判定へ反映する。

自動判定 REVIEW は上書きしない。確認結果を別フィールドとして保持し、
「機械の判定」と「人が確定した結果」の両方を最終証跡に残す。
"""

import hashlib
import json
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook

from .config import ConfigError
from .models import NG, OK, RunResult


REVIEW_HEADERS = ["No", "確認項目", "分類", "自動判定", "内容",
                  "確認結果", "確認者", "確認日"]


def _text(value) -> str:
    return "" if value is None else str(value).strip()


def _confirmation_date(value) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    raw = _text(value)
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    raise ConfigError("確認日は YYYY-MM-DD（または YYYY/MM/DD）で入力してください: %r" % raw)


def _case_sheet(workbook, case):
    matches = []
    prefix = case.case_id + "  "
    for ws in workbook.worksheets:
        if ws.title == "サマリ":
            continue
        title = _text(ws.cell(1, 1).value)
        if ws.title == case.case_id or title.startswith(prefix):
            matches.append(ws)
    if len(matches) != 1:
        raise ConfigError(
            "確認用 Excel にケース %s のシートが一意に見つかりません（候補 %d 件）。"
            "別の実行結果の Excel を指定していないか確認してください。"
            % (case.case_id, len(matches)))
    return matches[0]


def _header_row(ws) -> int:
    for row in range(1, ws.max_row + 1):
        values = [_text(ws.cell(row, col).value) for col in range(1, 9)]
        if values == REVIEW_HEADERS:
            return row
    raise ConfigError(
        "%s シートに人工確認欄がありません。report/finalize の順序と Excel を確認してください。"
        % ws.title)


def _verify_workbook_source(workbook, run: RunResult) -> None:
    """別実行の見た目が似た Excel を誤って確定しないよう、来源を照合する。"""
    if "サマリ" not in workbook.sheetnames:
        raise ConfigError("確認用 Excel にサマリシートがありません")
    ws = workbook["サマリ"]
    meta = {}
    for row in range(1, min(ws.max_row, 40) + 1):
        label = _text(ws.cell(row, 1).value)
        if label:
            meta[label] = _text(ws.cell(row, 2).value)

    if run.filter_description.startswith("統合レポート:"):
        if meta.get("実行対象") != run.filter_description:
            raise ConfigError(
                "確認用 Excel の結合元が --run 指定と一致しません（Excel: %r / 指定: %r）"
                % (meta.get("実行対象"), run.filter_description))
    elif meta.get("実行ID") != run.run_id:
        # report は --run が 1 件だけでも「統合レポート」として再構築する。
        # 元 run の Excel と、1 件 report の Excel のどちらからでも確定できるよう、
        # 後者は結合元 ID が完全一致するときだけ受け付ける。
        single_report = "統合レポート: " + run.run_id
        if meta.get("実行対象") != single_report:
            raise ConfigError(
                "確認用 Excel の実行IDが --run 指定と一致しません"
                "（Excel: %r / 結合元: %r / 指定: %r）"
                % (meta.get("実行ID"), meta.get("実行対象"), run.run_id))


def apply_excel_confirmations(run: RunResult, excel_path: Path) -> List[Dict[str, str]]:
    """黄色欄の入力を検証し、RunResult の REVIEW 項目へ反映する。"""
    excel_path = Path(excel_path)
    if not excel_path.is_file():
        raise ConfigError("確認結果を記入した Excel がありません: %s" % excel_path)
    if run.manual_pending:
        raise ConfigError(
            "未採取の手動実施ケースが残っています: %s。"
            "先に manual --phase before/after で採取し、report で一冊にまとめてください。"
            % ", ".join(run.manual_pending))

    try:
        workbook = load_workbook(str(excel_path), read_only=True, data_only=False)
    except Exception as exc:
        raise ConfigError("確認用 Excel を開けません: %s (%s)" % (excel_path, exc))

    confirmations = []  # type: List[Dict[str, str]]
    try:
        _verify_workbook_source(workbook, run)
        reviewable = sum(1 for case in run.cases for check in case.checks if check.is_reviewable)
        if not reviewable:
            raise ConfigError("この実行結果には人工確認が必要な項目がありません")

        for case in run.cases:
            targets = [(i, check) for i, check in enumerate(case.checks, start=1)
                       if check.is_reviewable]
            if not targets:
                continue
            ws = _case_sheet(workbook, case)
            first = _header_row(ws) + 1
            rows_by_no = {}
            for row in range(first, min(ws.max_row, first + len(case.checks) + 5) + 1):
                value = ws.cell(row, 1).value
                try:
                    rows_by_no[int(value)] = row
                except (TypeError, ValueError):
                    continue

            for number, check in targets:
                row = rows_by_no.get(number)
                if row is None:
                    raise ConfigError("%s / No.%d の確認行が Excel にありません" % (case.case_id, number))
                excel_name = _text(ws.cell(row, 2).value)
                if excel_name != check.name:
                    raise ConfigError(
                        "%s / No.%d の確認項目が実行結果と一致しません（Excel: %r / 結果: %r）"
                        % (case.case_id, number, excel_name, check.name))

                result = _text(ws.cell(row, 6).value).upper()
                reviewer = _text(ws.cell(row, 7).value)
                raw_date = ws.cell(row, 8).value
                if result not in (OK, NG):
                    raise ConfigError(
                        "%s / No.%d の確認結果は OK または NG を入力してください（現在: %r）"
                        % (case.case_id, number, result))
                if not reviewer:
                    raise ConfigError("%s / No.%d の確認者が未入力です" % (case.case_id, number))
                if raw_date is None or not _text(raw_date):
                    raise ConfigError("%s / No.%d の確認日が未入力です" % (case.case_id, number))
                confirmed_at = _confirmation_date(raw_date)

                check.confirmation_result = result
                check.confirmation_by = reviewer
                check.confirmation_at = confirmed_at
                confirmations.append({
                    "case_id": case.case_id,
                    "check_no": str(number),
                    "check_name": check.name,
                    "auto_verdict": check.verdict,
                    "confirmation_result": result,
                    "confirmation_by": reviewer,
                    "confirmation_at": confirmed_at,
                })
    finally:
        close = getattr(workbook, "close", None)
        if close:
            close()

    pending = ["%s/%s" % (case.case_id, check.name)
               for case in run.cases for check in case.checks if check.needs_review]
    if pending:
        raise ConfigError("人工確認が未完了です: %s" % pending)
    return confirmations


def write_audit(path: Path, run: RunResult, source_runs: List[str],
                input_excel: Path, confirmations: List[Dict[str, str]]) -> Path:
    """最終 Excel と対になる、機械可読の人工確認監査ログを保存する。"""
    input_excel = Path(input_excel)
    digest = hashlib.sha256(input_excel.read_bytes()).hexdigest()
    data = {
        "format_version": 1,
        "finalized_at": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
        "source_runs": list(source_runs),
        "input_excel": str(input_excel),
        "input_excel_sha256": digest,
        "final_verdict": run.verdict,
        "confirmations": confirmations,
    }
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2, sort_keys=True)
    return path
