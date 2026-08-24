# -*- coding: utf-8 -*-
"""Excel のレイアウト定義から取込テスト用 TXT を生成する。

画像の仕様にある並びを既定フォーマットとし、1 帳票を CSV 形式の 1 行にする。

  1. FormID 情報
  2. 対象有無情報
  3. FieldID 情報       ┐
  4. OCR 文字識別結果   ├ 対象項目分だけ繰り返す
  5. 属性フラグ         │
  6. 座標情報           ┘

OCR 値は Excel の I/J/K 列（データ型 / IME / 最大桁数）から決定する。
列位置や出力形式はコマンド引数で差し替えられるため、帳票ごとにコードを
複製しなくてよい。

  PYTHONPATH=src python -m autotest.layout_txt definition.xlsx --out-dir output/layout_txt
  PYTHONPATH=src python -m autotest.layout_txt --gui
"""

import argparse
import io
import os
import re
import sys
import tarfile
import tempfile
from collections import OrderedDict
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import column_index_from_string, get_column_letter


class LayoutTxtError(Exception):
    """入力定義または生成条件に問題がある。"""


HEADER_ALIASES = {
    "form_id": ("FORM_ID", "FORMID"),
    "layout_id": ("LAYOUT_ID", "LAYOUTID"),
    # FieldID は数字 ID の列だけを対象とする。ITEM_NAME を代用すると、取込形式は
    # 作れても別項目として解釈されるため、自動フォールバックさせない。
    "field_id": ("ELEMENT_ID", "ELEMENTID", "FIELD_ID", "FIELDID",
                 "ITEM_ID", "ITEMID", "SET_ID", "SETID"),
    "item_name": ("ITEM_NAME", "ITEMNAME", "項目名"),
    "data_type": ("ELEMENT_DATA_TYPE_NAME", "ELEMENTDATATYPENAME", "DATA_TYPE", "DATATYPE"),
    # 元資料の見出しは ELEMENT_IME_NAME。IME/IME の表記揺れも許容する。
    "ime": ("ELEMENT_IME_NAME", "ELEMENTIMENAME", "IME_NAME", "IMENAME"),
    "max_digits": ("MAX_NUM_DIGITS", "MAXNUMDIGITS", "MAX_DIGITS", "MAXDIGITS"),
    "input_attribute": ("入力属性", "INPUT_ATTRIBUTE", "INPUTATTRIBUTE"),
    "input_rule": ("入力規則", "INPUT_RULE", "INPUTRULE"),
    "notes": ("補足", "NOTES", "NOTE", "REMARKS"),
    "output_example": ("出力例", "OUTPUT_EXAMPLE", "OUTPUTEXAMPLE"),
    "default_value": (
        "OCR_DEFAULT_VALUE", "OCRDEFAULTVALUE", "DEFAULT_VALUE", "DEFAULTVALUE",
        "OCR既定値", "既定値", "デフォルト値", "默认值",
    ),
}

NULL_WORDS = ("", "NULL", "NONE", "N/A", "NA", "－", "-")
INVALID_FILENAME_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')

DATE_VALUES = (
    "5/8/6/1",           # 和暦記載
    "/2026/6/1",        # 西暦記載
    "5/2026/6/1",       # 元号 + 4 桁西暦
    "5/8/6/1|5/8/6/2",  # 帳票上の 1 行に複数記載
)

CHECKBOX_VALUES = ("1", "0")
CHOICE_VALUES = ("1", "2", "1.2")
CALENDAR_OCCURRENCES = 46
DEFAULT_VALUE_HEADER = "OCR_DEFAULT_VALUE"
DEFAULT_EMPTY_VALUE = "<EMPTY_OCR>"

# 4001 は全網羅データ用。1 と 2 が同じ項目に混在する場合はカンマ区切り。
COVERAGE_ATTRIBUTE_FLAGS = ("0", "1", "2", "1,2")

# core は取込仕様で特に重要な不整合、all は入力ファイルの境界・破損系まで含む。
CORE_ERROR_PATTERNS = (
    "normal",
    "count_zero",
    "count_missing",
    "count_extra",
    "element_id_unknown",
    "element_id_duplicate",
    "unrecognizable",
    "attribute_mixed_1_2",
)

ALL_ERROR_PATTERNS = CORE_ERROR_PATTERNS + (
    "count_duplicate_all",
    "element_id_empty",
    "element_order_reverse",
    "attribute_1_without_count_error",
    "attribute_invalid",
    "ocr_empty",
    "ocr_over_max",
    "ocr_invalid_date",
    "ocr_invalid_selection",
    "ocr_invalid_ime",
    "coordinates_empty",
    "coordinates_invalid",
    "target_absent",
    "target_empty",
    "form_id_mismatch",
    "form_id_empty",
)

PATTERN_DESCRIPTIONS = {
    "normal": "正常データ（属性0）",
    "count_zero": "項目数0（Fieldブロックなし）",
    "count_missing": "項目数不足（末尾Fieldを欠落、属性1）",
    "count_extra": "項目数超過（末尾Fieldを1件追加、属性1）",
    "element_id_unknown": "ELEMENT_IDが定義外",
    "element_id_duplicate": "ELEMENT_IDが重複",
    "unrecognizable": "OCR認識不可（空値、属性2）",
    "attribute_mixed_1_2": "個数不正と認識不可が混在（属性1,2）",
    "count_duplicate_all": "全Fieldブロックを二重化（属性1）",
    "element_id_empty": "ELEMENT_IDが空",
    "element_order_reverse": "ELEMENT_IDの順序が逆転",
    "attribute_1_without_count_error": "個数正常なのに属性1",
    "attribute_invalid": "未定義の属性値9",
    "ocr_empty": "属性0なのにOCR値が空",
    "ocr_over_max": "最大桁数超過",
    "ocr_invalid_date": "日付形式不正",
    "ocr_invalid_selection": "選択肢が数字以外／区切り不正",
    "ocr_invalid_ime": "IME制限外文字",
    "coordinates_empty": "座標情報が空",
    "coordinates_invalid": "座標形式不正",
    "target_absent": "対象有無=0",
    "target_empty": "対象有無が空",
    "form_id_mismatch": "TXT内FormIDがファイル対象と不一致",
    "form_id_empty": "TXT内FormIDが空",
}


def _normalize(value) -> str:
    if value is None:
        return ""
    return re.sub(r"[\s_\-　]+", "", str(value)).upper()


def _text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y%m%d%H%M%S")
    if isinstance(value, date):
        return value.strftime("%Y%m%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _cell_text(cell) -> str:
    """ID の 00001 のようなゼロ埋め表示を可能な範囲で保持する。"""
    value = cell.value
    if value is None:
        return ""
    fmt = str(getattr(cell, "number_format", "") or "")
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if re.match(r"^0+$", fmt) and float(value).is_integer():
            return ("%%0%dd" % len(fmt)) % int(value)
    return _text(value)


def _parse_max_digits(value, row_number: int) -> Optional[int]:
    raw = _text(value)
    if raw.upper() in NULL_WORDS:
        return None
    try:
        number = int(float(raw))
    except (TypeError, ValueError):
        raise LayoutTxtError(
            "%d 行目の最大桁数を数値として解釈できません: %r" % (row_number, raw))
    if number < 0:
        raise LayoutTxtError("%d 行目の最大桁数が負数です: %s" % (row_number, raw))
    return number


def _repeat_to_length(seed: str, length: int) -> str:
    if length <= 0:
        return ""
    if not seed:
        seed = "A"
    return (seed * ((length // len(seed)) + 1))[:length]


def _date_value(date_mode: str, date_index: int,
                allow_multiple: bool = False) -> str:
    modes = {
        "wareki": DATE_VALUES[0],
        "seireki": DATE_VALUES[1],
        "era-seireki": DATE_VALUES[2],
    }
    if date_mode == "cycle":
        values = DATE_VALUES if allow_multiple else DATE_VALUES[:3]
        return values[date_index % len(values)]
    if date_mode == "multiple":
        # 複数日付はカレンダーだけに許可する。日付/From/Toは常に単一値。
        return DATE_VALUES[3] if allow_multiple else DATE_VALUES[0]
    if date_mode not in modes:
        raise LayoutTxtError("未対応の日付生成形式です: %s" % date_mode)
    return modes[date_mode]


def _base_value(data_type: str, ime_name: str,
                date_mode: str = "cycle", date_index: int = 0,
                selection_index: int = 0) -> Tuple[str, str]:
    """代表値と値種別を返す。値種別は桁境界値の作り方に使う。"""
    dtype = _normalize(data_type)
    ime = _normalize(ime_name)

    if "チェックボックス" in dtype or "CHECKBOX" in dtype:
        return CHECKBOX_VALUES[selection_index % len(CHECKBOX_VALUES)], "fixed"
    if "ラジオ" in dtype or "選択肢" in dtype or "RADIO" in dtype or "SELECT" in dtype:
        return CHOICE_VALUES[selection_index % len(CHOICE_VALUES)], "fixed"
    if "カレンダー" in dtype or "CALENDAR" in dtype:
        return _date_value(
            date_mode, date_index, allow_multiple=True), "date"
    if "日付" in dtype or "DATE" in dtype:
        return _date_value(
            date_mode, date_index, allow_multiple=False), "date"

    if "小数" in ime or "DECIMAL" in ime or "FLOAT" in ime:
        return "1.0", "decimal"
    if "数値" in ime or "数字" in ime or "NUMERIC" in ime or "DIGIT" in ime:
        return "1234567890", "digits"
    if "半角カタカナ" in ime or "HANKAKUKATAKANA" in ime:
        return "ｱｲｳｴｵ", "text"
    if "半角英数" in ime or "ALPHANUMERIC" in ime or "ASCII" in ime:
        return "Abc123", "text"
    if "ひらがな" in ime or "HIRAGANA" in ime:
        return "あいうえお", "text"
    if "全タイプ" in ime or "ALLTYPE" in ime:
        return "あA1ｱ", "text"

    if "署名" in dtype or "SIGN" in dtype:
        return "山田太郎", "text"
    return "テストA1", "text"


def generate_ocr_value(data_type: str, ime_name: str, max_digits: Optional[int],
                       profile: str = "normal", item_name: str = "",
                       date_mode: str = "cycle", date_index: int = 0,
                       selection_index: int = 0) -> str:
    """I/J/K の値から再現可能な OCR テスト値を作る。

    normal: 代表的な妥当値（最大桁数以内）
    max:    最大桁数ちょうどの妥当値
    over:   最大桁数 + 1 の境界外値
    """
    if profile not in ("normal", "max", "over"):
        raise LayoutTxtError("未対応の生成パターンです: %s" % profile)

    base, kind = _base_value(
        data_type, ime_name, date_mode=date_mode, date_index=date_index,
        selection_index=selection_index)
    dtype = _normalize(data_type)
    is_name = "氏名" in dtype or "NAME" in dtype
    is_selection = ("チェックボックス" in dtype or "選択肢" in dtype or
                    "ラジオ" in dtype or "CHECKBOX" in dtype or
                    "SELECT" in dtype or "RADIO" in dtype)
    # 文字列項目は ITEM_NAME 自体が最も識別しやすいテスト値になる。
    # 最大桁テストでも同じ名称を種にすることで、どの項目かを見失わない。
    # 氏名は例外で、ELEMENT_IME_NAME の入力制約を優先する。
    if item_name and not is_name and (
            "文字列" in dtype or "STRING" in dtype or "TEXT" in dtype):
        base = item_name
        kind = "text"
    # 選択系はK列の桁境界より、取入仕様で定めた値集合を優先する。
    if is_selection:
        return base
    if max_digits is None:
        return base

    if profile == "normal":
        return base[:max_digits]

    target = max_digits if profile == "max" else max_digits + 1
    if kind == "decimal":
        if target <= 0:
            return ""
        if target == 1:
            return "1"
        if target == 2:
            return "12"
        return "1." + _repeat_to_length("234567890", target - 2)
    if kind == "digits" or kind == "date" or kind == "fixed":
        return _repeat_to_length("1234567890", target)
    return _repeat_to_length(base, target)


def _serial_coordinates(index: int) -> str:
    return "0,0,0,%d" % (index + 1)


def _coordinate_value(coordinates: str, index: int) -> str:
    raw = str(coordinates or "auto").strip()
    if raw.lower() in ("auto", "serial") or raw == "連番":
        return _serial_coordinates(index)
    return raw


class LayoutField(object):
    def __init__(self, form_id: str, layout_id: str, field_id: str, item_name: str,
                 data_type: str, ime_name: str, max_digits: Optional[int],
                 value: str, row_number: int,
                 attribute_flag: str = "0", coordinates: str = "0,0,0,1",
                 input_attribute: str = "", input_rule: str = "",
                 notes: str = "", output_example: str = "",
                 occurrence_index: int = 1, occurrence_count: int = 1) -> None:
        self.form_id = form_id
        self.layout_id = layout_id
        self.field_id = field_id
        self.item_name = item_name
        self.data_type = data_type
        self.ime_name = ime_name
        self.max_digits = max_digits
        self.value = value
        self.row_number = row_number
        self.attribute_flag = attribute_flag
        self.coordinates = coordinates
        self.input_attribute = input_attribute
        self.input_rule = input_rule
        self.notes = notes
        self.output_example = output_example
        self.occurrence_index = occurrence_index
        self.occurrence_count = occurrence_count

    @property
    def instance_key(self) -> str:
        return "%d:%d" % (self.row_number, self.occurrence_index)

    @property
    def row_label(self) -> str:
        if self.occurrence_count <= 1:
            return str(self.row_number)
        return "%d (%d/%d)" % (
            self.row_number, self.occurrence_index, self.occurrence_count)


class GenerationResult(object):
    def __init__(self, files: List[Path], field_count: int,
                 form_count: int, sheet_name: str, header_row: int,
                 columns: Dict[str, int], tif_files: Optional[List[Path]] = None,
                 pattern_count: int = 0, tar_file: Optional[Path] = None,
                 archive_members: Optional[List[str]] = None) -> None:
        # files は後方互換のため TXT 一覧のまま維持する。
        self.files = files
        self.txt_files = files
        self.tif_files = tif_files or []
        self.field_count = field_count
        self.form_count = form_count
        self.pattern_count = pattern_count
        self.tar_file = tar_file
        self.archive_members = archive_members or []
        self.sheet_name = sheet_name
        self.header_row = header_row
        self.columns = columns


class GeneratedCase(object):
    """1行のTXTと対応TIFを構成する1テストPattern。"""

    def __init__(self, source_form_id: str, form_id: str,
                 fields: List[LayoutField], target_presence: str,
                 pattern: str, sequence: int) -> None:
        self.source_form_id = source_form_id
        self.form_id = form_id
        self.fields = fields
        self.target_presence = target_presence
        self.pattern = pattern
        self.sequence = sequence


def _header_map(ws, header_row: int) -> Dict[str, List[int]]:
    found = {}  # type: Dict[str, List[int]]
    for cell in ws[header_row]:
        key = _normalize(cell.value)
        if key:
            column = cell.column
            # openpyxl 2.5 系ではセル種別によって列記号が返る場合がある。
            if isinstance(column, str):
                column = column_index_from_string(column)
            found.setdefault(key, []).append(column)
    return found


def _find_header_row(ws, max_scan_rows: int = 50) -> int:
    best_row = 0
    best_score = 0
    important = set()
    for key in ("data_type", "ime", "max_digits"):
        important.update(_normalize(v) for v in HEADER_ALIASES[key])

    upper = min(max_scan_rows, ws.max_row)
    for row_number in range(1, upper + 1):
        values = [_normalize(c.value) for c in ws[row_number]]
        score = sum(1 for value in values if value in important)
        # 写真の表は I/J/K が連続している。この並びに一致する行を優先する。
        if len(values) >= 11:
            expected = (
                _normalize(values[8]) in set(_normalize(v) for v in HEADER_ALIASES["data_type"]),
                _normalize(values[9]) in set(_normalize(v) for v in HEADER_ALIASES["ime"]),
                _normalize(values[10]) in set(_normalize(v) for v in HEADER_ALIASES["max_digits"]),
            )
            score += sum(2 for ok in expected if ok)
        if score > best_score:
            best_score = score
            best_row = row_number
    if best_score == 0:
        raise LayoutTxtError(
            "見出し行を自動検出できません。--header-row で行番号を指定してください。")
    return best_row


def _column_from_selector(selector: str, role: str,
                          headers: Dict[str, List[int]]) -> Optional[int]:
    raw = str(selector or "auto").strip()
    if raw.lower() == "none":
        return None
    if raw.lower() == "auto":
        for alias in HEADER_ALIASES[role]:
            positions = headers.get(_normalize(alias), [])
            if positions:
                return positions[0]
        return None
    if re.match(r"^[A-Za-z]{1,3}$", raw):
        try:
            return column_index_from_string(raw.upper())
        except ValueError:
            pass
    positions = headers.get(_normalize(raw), [])
    if not positions:
        raise LayoutTxtError("列 %r が見出し行にありません（用途: %s）" % (raw, role))
    return positions[0]


def _resolve_columns(ws, header_row: int, form_column: str, layout_column: str,
                     field_column: str,
                     item_column: str, data_type_column: str, ime_column: str,
                     max_digits_column: str,
                     input_attribute_column: str = "auto",
                     input_rule_column: str = "auto",
                     notes_column: str = "auto",
                     output_example_column: str = "auto",
                     default_value_column: str = "auto") -> Dict[str, Optional[int]]:
    headers = _header_map(ws, header_row)
    selectors = {
        "form_id": form_column,
        "layout_id": layout_column,
        "field_id": field_column,
        "item_name": item_column,
        "data_type": data_type_column,
        "ime": ime_column,
        "max_digits": max_digits_column,
        "input_attribute": input_attribute_column,
        "input_rule": input_rule_column,
        "notes": notes_column,
        "output_example": output_example_column,
        "default_value": default_value_column,
    }
    columns = {}
    for role, selector in selectors.items():
        columns[role] = _column_from_selector(selector, role, headers)

    required = ("form_id", "layout_id", "field_id", "data_type", "ime", "max_digits")
    missing = [role for role in required if columns.get(role) is None]
    if missing:
        raise LayoutTxtError("必須列を特定できません: %s" % ", ".join(missing))
    return columns


def _value_at(row: Sequence, column: Optional[int]) -> str:
    if column is None or column < 1 or column > len(row):
        return ""
    return _cell_text(row[column - 1])


def _raw_value_at(row: Sequence, column: Optional[int]):
    """Return a cell value without stripping meaningful default-value whitespace."""
    if column is None or column < 1 or column > len(row):
        return None
    value = row[column - 1].value
    if value is None:
        return None
    if isinstance(value, str):
        return value
    return _cell_text(row[column - 1])


def _default_values(raw_value, occurrence_count: int,
                    row_number: int) -> Optional[List[str]]:
    if raw_value is None or raw_value == "":
        return None
    normalized = str(raw_value).replace("\r\n", "\n").replace("\r", "\n")
    values = normalized.split("\n")
    values = ["" if value == DEFAULT_EMPTY_VALUE else value for value in values]
    if len(values) == 1:
        return values * occurrence_count
    if len(values) != occurrence_count:
        raise LayoutTxtError(
            "%d 行目の既定値は1件または%d件（セル内改行区切り）にしてください: %d件"
            % (row_number, occurrence_count, len(values)))
    return values


def read_layout_fields(excel_path: Path, sheet_name: Optional[str] = None,
                       header_row: Optional[int] = None,
                       form_column: str = "auto", layout_column: str = "auto",
                       field_column: str = "auto",
                       item_column: str = "auto", data_type_column: str = "I",
                       ime_column: str = "J", max_digits_column: str = "K",
                       input_attribute_column: str = "auto",
                       input_rule_column: str = "auto",
                       notes_column: str = "auto",
                       output_example_column: str = "auto",
                       profile: str = "normal", date_mode: str = "coverage",
                       coverage_form_id: str = "4001",
                       attribute_flag: str = "0",
                       coordinates: str = "auto",
                       default_value_column: str = "auto") -> Tuple[List[LayoutField], str, int, Dict[str, int]]:
    path = Path(excel_path)
    if not path.is_file():
        raise LayoutTxtError("Excel ファイルが見つかりません: %s" % path)
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        raise LayoutTxtError(".xlsx / .xlsm のみ対応しています: %s" % path.name)

    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
    except Exception as exc:
        raise LayoutTxtError("Excel を開けません: %s" % exc)
    try:
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise LayoutTxtError(
                    "シート %r がありません。候補: %s" % (sheet_name, ", ".join(wb.sheetnames)))
            ws = wb[sheet_name]
        else:
            ws = wb.active

        actual_header_row = header_row or _find_header_row(ws)
        if actual_header_row < 1 or actual_header_row > ws.max_row:
            raise LayoutTxtError("見出し行が範囲外です: %s" % actual_header_row)
        raw_columns = _resolve_columns(
            ws, actual_header_row, form_column, layout_column, field_column, item_column,
            data_type_column, ime_column, max_digits_column,
            input_attribute_column=input_attribute_column,
            input_rule_column=input_rule_column,
            notes_column=notes_column,
            output_example_column=output_example_column,
            default_value_column=default_value_column)

        fields = []  # type: List[LayoutField]
        last_form_id = ""
        last_layout_id = ""
        date_counts = {}  # type: Dict[Tuple[str, str], int]
        selection_counts = {}  # type: Dict[Tuple[str, str], int]
        coordinate_counts = {}  # type: Dict[str, int]
        coverage_id = str(coverage_form_id or "").strip()
        if coverage_id and not coverage_id.isdigit():
            raise LayoutTxtError("全網羅用 FormID は数字で指定してください: %r" % coverage_id)
        for row_number, row in enumerate(
                ws.iter_rows(min_row=actual_header_row + 1), actual_header_row + 1):
            raw_form_id = _value_at(row, raw_columns.get("form_id"))
            raw_layout_id = _value_at(row, raw_columns.get("layout_id"))
            field_id = _value_at(row, raw_columns.get("field_id"))
            item_name = _value_at(row, raw_columns.get("item_name"))
            data_type = _value_at(row, raw_columns.get("data_type"))
            ime_name = _value_at(row, raw_columns.get("ime"))
            max_raw = _value_at(row, raw_columns.get("max_digits"))
            input_attribute = _value_at(row, raw_columns.get("input_attribute"))
            input_rule = _value_at(row, raw_columns.get("input_rule"))
            notes = _value_at(row, raw_columns.get("notes"))
            output_example = _value_at(row, raw_columns.get("output_example"))
            default_raw = _raw_value_at(row, raw_columns.get("default_value"))

            # 値を引き継ぐ前に、完全な空行を表末尾として無視する。
            if not any((raw_form_id, raw_layout_id, field_id, item_name,
                        data_type, ime_name, max_raw)):
                continue
            # 結合セルや省略表記の ID は直前値を引き継ぐ。
            if raw_form_id:
                last_form_id = raw_form_id
            if raw_layout_id:
                last_layout_id = raw_layout_id
            form_id = raw_form_id or last_form_id
            layout_id = raw_layout_id or last_layout_id
            if not form_id:
                raise LayoutTxtError("%d 行目の FormID が空です。" % row_number)
            if not layout_id:
                raise LayoutTxtError("%d 行目の LayoutID が空です。" % row_number)
            if not field_id:
                raise LayoutTxtError("%d 行目の FieldID が空です。" % row_number)
            if not form_id.isdigit():
                raise LayoutTxtError(
                    "%d 行目の FormID は数字で指定してください: %r" % (row_number, form_id))
            if not layout_id.isdigit():
                raise LayoutTxtError(
                    "%d 行目の LayoutID は数字で指定してください: %r" % (row_number, layout_id))
            if not field_id.isdigit():
                raise LayoutTxtError(
                    "%d 行目の FieldID は数字で指定してください: %r" % (row_number, field_id))
            if not item_name:
                item_name = field_id

            max_digits = _parse_max_digits(max_raw, row_number)
            dtype = _normalize(data_type)
            is_date = ("カレンダー" in dtype or "日付" in dtype or
                       "DATE" in dtype or "CALENDAR" in dtype)
            is_checkbox = "チェックボックス" in dtype or "CHECKBOX" in dtype
            is_choice = ("ラジオ" in dtype or "選択肢" in dtype or
                         "RADIO" in dtype or "SELECT" in dtype)
            date_kind = "calendar" if (
                "カレンダー" in dtype or "CALENDAR" in dtype) else "date"
            date_key = (form_id, date_kind)
            selection_kind = "checkbox" if is_checkbox else "choice"
            selection_key = (form_id, selection_kind)
            selection_index = selection_counts.get(selection_key, 0)
            actual_date_mode = date_mode
            if date_mode == "coverage":
                actual_date_mode = "cycle" if form_id == coverage_id else "wareki"
            occurrence_count = CALENDAR_OCCURRENCES if date_kind == "calendar" else 1
            saved_defaults = _default_values(
                default_raw, occurrence_count, row_number)
            for occurrence_index in range(1, occurrence_count + 1):
                date_index = date_counts.get(date_key, 0)
                coordinate_index = coordinate_counts.get(form_id, 0)
                value = generate_ocr_value(
                    data_type, ime_name, max_digits, profile=profile, item_name=item_name,
                    date_mode=actual_date_mode, date_index=date_index,
                    selection_index=selection_index)
                if saved_defaults is not None:
                    value = saved_defaults[occurrence_index - 1]

                fields.append(LayoutField(
                    form_id=form_id, layout_id=layout_id,
                    field_id=field_id, item_name=item_name,
                    data_type=data_type, ime_name=ime_name, max_digits=max_digits,
                    value=value, row_number=row_number,
                    attribute_flag=attribute_flag,
                    coordinates=_coordinate_value(coordinates, coordinate_index),
                    input_attribute=input_attribute, input_rule=input_rule,
                    notes=notes, output_example=output_example,
                    occurrence_index=occurrence_index,
                    occurrence_count=occurrence_count))
                coordinate_counts[form_id] = coordinate_index + 1
                if is_date:
                    date_counts[date_key] = date_index + 1
            if is_checkbox or is_choice:
                selection_counts[selection_key] = selection_index + 1

        if not fields:
            raise LayoutTxtError("見出し行より下に生成対象のデータがありません。")

        columns = {}  # type: Dict[str, int]
        for key, value in raw_columns.items():
            if value is not None:
                columns[key] = value
        return fields, ws.title, actual_header_row, columns
    finally:
        wb.close()


def _default_column_for_write(ws, header_row: int, selector: str) -> int:
    headers = _header_map(ws, header_row)
    raw = str(selector or "auto").strip()
    if raw.lower() == "none":
        raise LayoutTxtError("既定値列が none のためExcelへ保存できません。")
    if raw.lower() == "auto":
        for alias in HEADER_ALIASES["default_value"]:
            positions = headers.get(_normalize(alias), [])
            if positions:
                return positions[0]
        column = ws.max_column + 1
        ws.cell(row=header_row, column=column, value=DEFAULT_VALUE_HEADER)
        return column
    if re.match(r"^[A-Za-z]{1,3}$", raw):
        try:
            column = column_index_from_string(raw.upper())
        except ValueError:
            column = 0
        if column:
            header_cell = ws.cell(row=header_row, column=column)
            if header_cell.value in (None, ""):
                header_cell.value = DEFAULT_VALUE_HEADER
            return column
    positions = headers.get(_normalize(raw), [])
    if positions:
        return positions[0]
    column = ws.max_column + 1
    ws.cell(row=header_row, column=column, value=raw)
    return column


def _instance_position(value) -> Tuple[int, int]:
    raw = str(value).strip()
    matched = re.match(r"^(\d+)(?::(\d+))?$", raw)
    if not matched:
        raise LayoutTxtError("Excel行キーが不正です: %r" % raw)
    return int(matched.group(1)), int(matched.group(2) or "1")


def save_layout_default_values(
        excel_path: Path, values_by_instance: Dict,
        sheet_name: Optional[str] = None,
        header_row: Optional[int] = None,
        default_value_column: str = "auto") -> Tuple[Path, int, int]:
    """Persist GUI OCR values into one default-value column in the source Excel."""
    path = Path(excel_path)
    if not path.is_file():
        raise LayoutTxtError("Excel ファイルが見つかりません: %s" % path)
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        raise LayoutTxtError(".xlsx / .xlsm のみ対応しています: %s" % path.name)
    if not values_by_instance:
        raise LayoutTxtError("Excelへ保存するOCR値がありません。")

    grouped = {}  # type: Dict[int, List[Tuple[int, str]]]
    for instance_key, value in values_by_instance.items():
        row_number, occurrence_index = _instance_position(instance_key)
        grouped.setdefault(row_number, []).append((occurrence_index, str(value)))

    try:
        wb = load_workbook(
            str(path), data_only=False, keep_vba=path.suffix.lower() == ".xlsm")
    except Exception as exc:
        raise LayoutTxtError("Excel を開けません: %s" % exc)

    temp_name = None
    try:
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise LayoutTxtError(
                    "シート %r がありません。候補: %s"
                    % (sheet_name, ", ".join(wb.sheetnames)))
            ws = wb[sheet_name]
        else:
            ws = wb.active
        actual_header_row = header_row or _find_header_row(ws)
        column = _default_column_for_write(
            ws, actual_header_row, default_value_column)

        for row_number, indexed_values in grouped.items():
            if row_number <= actual_header_row or row_number > ws.max_row:
                raise LayoutTxtError("既定値の保存先Excel行が範囲外です: %d" % row_number)
            indexed_values.sort(key=lambda item: item[0])
            indexes = [item[0] for item in indexed_values]
            if indexes != list(range(1, len(indexes) + 1)):
                raise LayoutTxtError(
                    "%d 行目の展開番号が連続していません: %s"
                    % (row_number, ", ".join(str(index) for index in indexes)))
            encoded = [value if value != "" else DEFAULT_EMPTY_VALUE
                       for _index, value in indexed_values]
            cell = ws.cell(row=row_number, column=column)
            cell.value = "\n".join(encoded)
            if len(encoded) > 1:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        handle, temp_name = tempfile.mkstemp(
            prefix=path.stem + ".", suffix=path.suffix, dir=str(path.parent))
        os.close(handle)
        wb.save(temp_name)
        wb.close()
        wb = None
        os.replace(temp_name, str(path))
        temp_name = None
        return path, column, len(grouped)
    except LayoutTxtError:
        raise
    except Exception as exc:
        raise LayoutTxtError(
            "既定値をExcelへ保存できません。Excelを閉じて再試行してください: %s" % exc)
    finally:
        if wb is not None:
            wb.close()
        if temp_name:
            try:
                os.unlink(temp_name)
            except OSError:
                pass


def _safe_filename(value: str) -> str:
    name = INVALID_FILENAME_CHARS.sub("_", value).strip(" .")
    if not name:
        name = "layout"
    # Windows の予約名を避ける。
    if name.upper() in ("CON", "PRN", "AUX", "NUL") or re.match(r"^(COM|LPT)[1-9]$", name.upper()):
        name = "_" + name
    return name[:120]


def _quoted(value: str) -> str:
    """CSV の規則で全値を必ずダブルクォートする。"""
    return '"' + str(value).replace('"', '""') + '"'


def _render_form(form_id: str, fields: Sequence[LayoutField], output_format: str,
                 target_presence: str, line_ending: str) -> str:
    lines = []  # type: List[str]
    if output_format == "raw":
        values = [form_id, target_presence]
        for field in fields:
            values.extend((field.field_id, field.value, field.attribute_flag, field.coordinates))
        # 1 行が 1 帳票。値にカンマが含まれる座標もあるため、全項目を引用する。
        lines.append(",".join(_quoted(value) for value in values))
    elif output_format == "labeled":
        values = ("FormID=" + form_id, "TargetPresence=" + target_presence)
        labeled = list(values)
        for field in fields:
            labeled.extend((
                "FieldID=" + field.field_id,
                "OCRText=" + field.value,
                "AttributeFlag=" + field.attribute_flag,
                "Coordinates=" + field.coordinates,
            ))
        lines.append(",".join(_quoted(value) for value in labeled))
    elif output_format == "tsv":
        lines.append("FormID\tTargetPresence\tFieldID\tOCRText\tAttributeFlag\tCoordinates")
        for field in fields:
            lines.append("\t".join((form_id, target_presence, field.field_id, field.value,
                                     field.attribute_flag, field.coordinates)))
    else:
        raise LayoutTxtError("未対応の出力形式です: %s" % output_format)
    return line_ending.join(lines) + line_ending


def _copy_field(field: LayoutField) -> LayoutField:
    return LayoutField(
        form_id=field.form_id, layout_id=field.layout_id,
        field_id=field.field_id, item_name=field.item_name,
        data_type=field.data_type, ime_name=field.ime_name,
        max_digits=field.max_digits, value=field.value,
        row_number=field.row_number, attribute_flag=field.attribute_flag,
        coordinates=field.coordinates,
        input_attribute=field.input_attribute, input_rule=field.input_rule,
        notes=field.notes, output_example=field.output_example,
        occurrence_index=field.occurrence_index,
        occurrence_count=field.occurrence_count)


def _copy_fields(fields: Sequence[LayoutField]) -> List[LayoutField]:
    return [_copy_field(field) for field in fields]


def _deduplicate_coordinates(fields: Sequence[LayoutField]) -> None:
    last_numbers = []
    for field in fields:
        try:
            parts = [int(value) for value in field.coordinates.split(",")]
            if len(parts) == 4:
                last_numbers.append(parts[-1])
        except ValueError:
            continue
    next_number = max(last_numbers) + 1 if last_numbers else 1
    seen = set()
    for field in fields:
        coordinate = field.coordinates
        if coordinate and coordinate in seen:
            field.coordinates = "0,0,0,%d" % next_number
            next_number += 1
        if field.coordinates:
            seen.add(field.coordinates)


def _is_date_field(field: LayoutField) -> bool:
    dtype = _normalize(field.data_type)
    return ("カレンダー" in dtype or "日付" in dtype or
            "DATE" in dtype or "CALENDAR" in dtype)


def _is_selection_field(field: LayoutField) -> bool:
    dtype = _normalize(field.data_type)
    return ("チェックボックス" in dtype or "選択肢" in dtype or
            "ラジオ" in dtype or "CHECKBOX" in dtype or
            "SELECT" in dtype or "RADIO" in dtype)


def _unknown_field_id(fields: Sequence[LayoutField]) -> str:
    numeric = [int(field.field_id) for field in fields if field.field_id.isdigit()]
    candidate = (max(numeric) + 900001) if numeric else 99999999
    used = set(field.field_id for field in fields)
    while str(candidate) in used:
        candidate += 1
    return str(candidate)


def _invalid_ime_value(field: LayoutField) -> str:
    ime = _normalize(field.ime_name)
    if "ひらがな" in ime or "HIRAGANA" in ime:
        return "ABC123"
    if "半角カタカナ" in ime or "HANKAKUKATAKANA" in ime:
        return "漢字ＡＢＣ"
    if "半角英数" in ime or "ALPHANUMERIC" in ime or "ASCII" in ime:
        return "全角ＡＢＣ"
    if ("数値" in ime or "数字" in ime or "NUMERIC" in ime or
            "DIGIT" in ime or "小数" in ime or "DECIMAL" in ime):
        return "ABC"
    return "制限外@文字"


def _make_pattern_case(source_form_id: str, base_fields: Sequence[LayoutField],
                       target_presence: str, pattern: str,
                       sequence: int) -> GeneratedCase:
    fields = _copy_fields(base_fields)
    form_id = source_form_id
    target = target_presence

    if pattern == "normal":
        # Excel/GUIで指定された属性をそのまま使う。通常読込の既定値は0。
        pass
    elif pattern == "count_zero":
        fields = []
    elif pattern == "count_missing":
        if fields:
            fields = fields[:-1]
        if fields:
            fields[-1].attribute_flag = "1"
    elif pattern == "count_extra":
        if fields:
            extra = _copy_field(fields[-1])
            extra.attribute_flag = "1"
            fields.append(extra)
        for field in fields:
            field.attribute_flag = "1"
    elif pattern == "element_id_unknown":
        if fields:
            fields[0].field_id = _unknown_field_id(fields)
    elif pattern == "element_id_duplicate":
        if len(fields) >= 2:
            fields[1].field_id = fields[0].field_id
        elif fields:
            fields.append(_copy_field(fields[0]))
    elif pattern == "unrecognizable":
        for field in fields:
            field.value = ""
            field.attribute_flag = "2"
    elif pattern == "attribute_mixed_1_2":
        for index, field in enumerate(fields):
            field.attribute_flag = COVERAGE_ATTRIBUTE_FLAGS[
                index % len(COVERAGE_ATTRIBUTE_FLAGS)]
            if "2" in field.attribute_flag:
                field.value = ""
    elif pattern == "count_duplicate_all":
        duplicated = _copy_fields(fields)
        fields.extend(duplicated)
        for field in fields:
            field.attribute_flag = "1"
    elif pattern == "element_id_empty":
        if fields:
            fields[0].field_id = ""
    elif pattern == "element_order_reverse":
        fields.reverse()
    elif pattern == "attribute_1_without_count_error":
        for field in fields:
            field.attribute_flag = "1"
    elif pattern == "attribute_invalid":
        for field in fields:
            field.attribute_flag = "9"
    elif pattern == "ocr_empty":
        if fields:
            fields[0].value = ""
            fields[0].attribute_flag = "0"
    elif pattern == "ocr_over_max":
        target_field = next(
            (field for field in fields if field.max_digits is not None),
            fields[0] if fields else None)
        if target_field is not None:
            if target_field.max_digits is None:
                target_field.value += "X"
            else:
                target_field.value = generate_ocr_value(
                    target_field.data_type, target_field.ime_name,
                    target_field.max_digits, profile="over",
                    item_name=target_field.item_name)
    elif pattern == "ocr_invalid_date":
        target_field = next(
            (field for field in fields if _is_date_field(field)),
            fields[0] if fields else None)
        if target_field is not None:
            target_field.value = "5/99/99/99"
    elif pattern == "ocr_invalid_selection":
        target_field = next(
            (field for field in fields if _is_selection_field(field)),
            fields[0] if fields else None)
        if target_field is not None:
            target_field.value = "A,B"
    elif pattern == "ocr_invalid_ime":
        target_field = next(
            (field for field in fields
             if not _is_date_field(field) and not _is_selection_field(field)),
            fields[0] if fields else None)
        if target_field is not None:
            target_field.value = _invalid_ime_value(target_field)
    elif pattern == "coordinates_empty":
        if fields:
            fields[0].coordinates = ""
    elif pattern == "coordinates_invalid":
        if fields:
            fields[0].coordinates = "X,Y,-1,ABC"
    elif pattern == "target_absent":
        target = "0"
    elif pattern == "target_empty":
        target = ""
    elif pattern == "form_id_mismatch":
        form_id = str(int(source_form_id) + 1) if source_form_id.isdigit() else source_form_id + "_NG"
    elif pattern == "form_id_empty":
        form_id = ""
    else:
        raise LayoutTxtError("未対応のエラーPatternです: %s" % pattern)

    _deduplicate_coordinates(fields)
    return GeneratedCase(
        source_form_id=source_form_id, form_id=form_id, fields=fields,
        target_presence=target, pattern=pattern, sequence=sequence)


def _build_cases(groups, coverage_form_id: str, error_patterns: str,
                 target_presence: str) -> List[GeneratedCase]:
    if error_patterns not in ("none", "core", "all"):
        raise LayoutTxtError("エラーPatternは none / core / all から指定してください: %s"
                             % error_patterns)
    coverage_id = str(coverage_form_id or "").strip()
    cases = []  # type: List[GeneratedCase]
    for form_id, form_fields in groups.items():
        if coverage_id and form_id == coverage_id and error_patterns != "none":
            names = CORE_ERROR_PATTERNS if error_patterns == "core" else ALL_ERROR_PATTERNS
        else:
            names = ("normal",)
        for sequence, pattern in enumerate(names, 1):
            cases.append(_make_pattern_case(
                form_id, form_fields, target_presence, pattern, sequence))
    return cases


def _case_stem(template: str, case: GeneratedCase, source_stem: str,
               multiple_patterns: bool) -> str:
    raw_template = str(template or "{form_id}").strip()
    if raw_template.lower().endswith(".txt") or raw_template.lower().endswith(".tif"):
        raw_template = raw_template[:-4]
    values = {
        "form_id": case.source_form_id,
        "output_form_id": case.form_id,
        "pattern": case.pattern,
        "seq": case.sequence,
        "source": source_stem,
    }
    try:
        stem = raw_template.format(**values)
    except (KeyError, ValueError, IndexError) as exc:
        raise LayoutTxtError("ファイル名テンプレートが不正です: %s" % exc)
    # Patternを複数出すのにテンプレート側で識別していない場合は安全に自動付与する。
    if multiple_patterns and "{pattern" not in raw_template and "{seq" not in raw_template:
        stem += "_%02d_%s" % (case.sequence, case.pattern)
    return _safe_filename(stem)


def _tif_font(size: int):
    from PIL import ImageFont

    candidates = (
        "C:/Windows/Fonts/meiryo.ttc",
        "C:/Windows/Fonts/YuGothM.ttc",
        "/System/Library/Fonts/ヒラギノ角ゴシック W3.ttc",
        "/System/Library/Fonts/ヒラギノ角ゴシック W4.ttc",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    )
    for candidate in candidates:
        if Path(candidate).exists():
            try:
                return ImageFont.truetype(candidate, size)
            except OSError:
                pass
    return ImageFont.load_default()


def _draw_tif_text(draw, xy, value: str, font, fill=(0, 0, 0)) -> None:
    text = str(value)
    try:
        draw.text(xy, text, font=font, fill=fill)
    except UnicodeEncodeError:
        draw.text(xy, text.encode("ascii", "replace").decode("ascii"),
                  font=font, fill=fill)


def _tif_payload(cases: Sequence[GeneratedCase]) -> bytes:
    """TXTと同じ内容を確認できるA4相当・複数ページTIFFを作る。"""
    try:
        from PIL import Image, ImageDraw
    except ImportError as exc:
        raise LayoutTxtError("TIF生成にはPillowが必要です: %s" % exc)
    try:
        from io import BytesIO
    except ImportError:  # pragma: no cover - Python 3.6 では到達しない
        from cStringIO import StringIO as BytesIO

    width, height = 1654, 2339
    margin, header_h, row_h = 80, 220, 70
    rows_per_page = max(1, (height - header_h - margin) // row_h)
    normal_font = _tif_font(28)
    small_font = _tif_font(23)
    pages = []

    for case in cases:
        chunks = [case.fields[index:index + rows_per_page]
                  for index in range(0, len(case.fields), rows_per_page)] or [[]]
        for page_number, chunk in enumerate(chunks, 1):
            image = Image.new("RGB", (width, height), (255, 255, 255))
            draw = ImageDraw.Draw(image)
            draw.rectangle((margin, margin, width - margin, height - margin),
                           outline=(40, 40, 40))
            title = "FORM %s / Pattern %02d %s" % (
                case.form_id or "<EMPTY>", case.sequence, case.pattern)
            _draw_tif_text(draw, (margin + 24, margin + 20), title, normal_font)
            description = PATTERN_DESCRIPTIONS.get(case.pattern, case.pattern)
            _draw_tif_text(draw, (margin + 24, margin + 70), description, small_font)
            _draw_tif_text(
                draw, (margin + 24, margin + 112),
                "Target=%s  Fields=%d  Page=%d/%d" % (
                    case.target_presence or "<EMPTY>", len(case.fields),
                    page_number, len(chunks)), small_font)

            y = margin + 165
            columns = (margin + 20, margin + 250, margin + 650, margin + 1320)
            for x, label in zip(columns, ("ELEMENT_ID", "ITEM_NAME / OCR", "", "ATTR")):
                if label:
                    _draw_tif_text(draw, (x, y), label, small_font)
            y += 45
            for field in chunk:
                draw.rectangle((margin + 15, y, width - margin - 15, y + row_h - 5),
                               outline=(170, 170, 170))
                _draw_tif_text(draw, (columns[0], y + 17), field.field_id or "<EMPTY>", small_font)
                label = "%s : %s" % (field.item_name, field.value or "<EMPTY>")
                if len(label) > 52:
                    label = label[:49] + "..."
                _draw_tif_text(draw, (columns[1], y + 17), label, small_font)
                _draw_tif_text(draw, (columns[3], y + 17), field.attribute_flag, small_font)
                y += row_h
            pages.append(image)

    stream = BytesIO()
    first = pages[0]
    options = {
        "format": "TIFF", "save_all": True,
        "append_images": pages[1:], "dpi": (200, 200),
        "compression": "tiff_deflate",
    }
    try:
        first.save(stream, **options)
    except (OSError, TypeError):
        stream = BytesIO()
        options.pop("compression", None)
        first.save(stream, **options)
    return stream.getvalue()


def _atomic_write(path: Path, payload: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    handle, temp_name = tempfile.mkstemp(prefix=path.name + ".", suffix=".tmp", dir=str(path.parent))
    try:
        with os.fdopen(handle, "wb") as stream:
            stream.write(payload)
            stream.flush()
            os.fsync(stream.fileno())
        os.replace(temp_name, str(path))
    except Exception:
        try:
            os.unlink(temp_name)
        except OSError:
            pass
        raise


def _tar_payload(rendered: Sequence[Tuple[Path, bytes, str]]) -> bytes:
    """生成済みTXT/TIFを再現可能な無圧縮TARにまとめる。"""
    stream = io.BytesIO()
    with tarfile.open(fileobj=stream, mode="w") as archive:
        for path, payload, _kind in rendered:
            info = tarfile.TarInfo(name=path.name)
            info.size = len(payload)
            info.mtime = 0
            info.mode = 0o644
            info.uid = 0
            info.gid = 0
            info.uname = ""
            info.gname = ""
            archive.addfile(info, io.BytesIO(payload))
    return stream.getvalue()


def _tar_stem(tar_name: str, source_stem: str,
              form_ids: Sequence[str]) -> str:
    raw = str(tar_name or "{source}_layout_data").strip()
    if raw.lower().endswith(".tar"):
        raw = raw[:-4]
    form_value = form_ids[0] if len(form_ids) == 1 else "all"
    try:
        value = raw.format(source=source_stem, form_id=form_value)
    except (KeyError, ValueError, IndexError) as exc:
        raise LayoutTxtError("TARファイル名が不正です: %s" % exc)
    return _safe_filename(value)


def _filter_and_override_fields(
        fields: Sequence[LayoutField],
        selected_form_ids: Optional[Sequence[str]],
        selected_rows: Optional[Sequence],
        field_overrides: Optional[Dict]) -> List[LayoutField]:
    selected = None  # type: Optional[List[str]]
    if selected_form_ids:
        selected = [str(value).strip() for value in selected_form_ids if str(value).strip()]
        available = set(field.form_id for field in fields)
        missing = [value for value in selected if value not in available]
        if missing:
            raise LayoutTxtError("指定FormIDがExcelにありません: %s" % ", ".join(missing))

    row_set = None if selected_rows is None else set(
        str(value).strip() for value in selected_rows)
    overrides = field_overrides or {}
    result = []  # type: List[LayoutField]
    for original in fields:
        if selected is not None and original.form_id not in selected:
            continue
        if (row_set is not None and
                str(original.row_number) not in row_set and
                original.instance_key not in row_set):
            continue
        field = _copy_field(original)
        values = overrides.get(
            field.instance_key,
            overrides.get(field.row_number, overrides.get(str(field.row_number), {})))
        if values:
            for name in ("field_id", "value", "attribute_flag", "coordinates"):
                if name in values:
                    setattr(field, name, str(values[name]))
        result.append(field)
    if not result:
        raise LayoutTxtError("指定条件に一致する出力項目がありません。")
    return result


def generate_layout_txt(excel_path: Path, output_dir: Path,
                        sheet_name: Optional[str] = None,
                        header_row: Optional[int] = None,
                        form_column: str = "auto", layout_column: str = "auto",
                        field_column: str = "auto",
                        item_column: str = "auto", data_type_column: str = "I",
                        ime_column: str = "J", max_digits_column: str = "K",
                        input_attribute_column: str = "auto",
                        input_rule_column: str = "auto",
                        notes_column: str = "auto",
                        output_example_column: str = "auto",
                        profile: str = "normal", date_mode: str = "coverage",
                        coverage_form_id: str = "4001",
                        output_format: str = "raw",
                        encoding: str = "cp932", split_by_form: bool = True,
                        overwrite: bool = False, target_presence: str = "1",
                        attribute_flag: str = "0", coordinates: str = "auto",
                        crlf: bool = True, error_patterns: str = "all",
                        filename_template: str = "{form_id}",
                        generate_tif: bool = True,
                        selected_form_ids: Optional[Sequence[str]] = None,
                        selected_rows: Optional[Sequence] = None,
                        field_overrides: Optional[Dict] = None,
                        create_tar: bool = False, tar_name: str = "",
                        tar_only: bool = False,
                        default_value_column: str = "auto") -> GenerationResult:
    fields, actual_sheet, actual_header, columns = read_layout_fields(
        excel_path=excel_path, sheet_name=sheet_name, header_row=header_row,
        form_column=form_column, layout_column=layout_column,
        field_column=field_column, item_column=item_column,
        data_type_column=data_type_column, ime_column=ime_column,
        max_digits_column=max_digits_column,
        input_attribute_column=input_attribute_column,
        input_rule_column=input_rule_column,
        notes_column=notes_column,
        output_example_column=output_example_column,
        default_value_column=default_value_column,
        profile=profile, date_mode=date_mode,
        coverage_form_id=coverage_form_id,
        attribute_flag=attribute_flag, coordinates=coordinates)

    fields = _filter_and_override_fields(
        fields, selected_form_ids=selected_form_ids,
        selected_rows=selected_rows, field_overrides=field_overrides)

    groups = OrderedDict()  # type: OrderedDict[str, List[LayoutField]]
    for field in fields:
        groups.setdefault(field.form_id, []).append(field)

    cases = _build_cases(
        groups, coverage_form_id=coverage_form_id,
        error_patterns=error_patterns, target_presence=target_presence)

    line_ending = "\r\n" if crlf else "\n"
    rendered = []  # type: List[Tuple[Path, bytes, str]]
    txt_files = []  # type: List[Path]
    tif_files = []  # type: List[Path]
    output_dir = Path(output_dir)
    source_stem = _safe_filename(Path(excel_path).stem)

    if split_by_form:
        seen_names = set()
        pattern_totals = {}  # type: Dict[str, int]
        for case in cases:
            pattern_totals[case.source_form_id] = pattern_totals.get(case.source_form_id, 0) + 1
        for case in cases:
            stem = _case_stem(
                filename_template, case, source_stem,
                multiple_patterns=pattern_totals[case.source_form_id] > 1)
            filename = stem + ".txt"
            folded = filename.lower()
            if folded in seen_names:
                raise LayoutTxtError(
                    "ファイル名テンプレートの結果が重複します: %s" % filename)
            seen_names.add(folded)
            content = _render_form(
                case.form_id, case.fields, output_format,
                case.target_presence, line_ending)
            try:
                payload = content.encode(encoding)
            except (LookupError, UnicodeEncodeError) as exc:
                raise LayoutTxtError("%s でエンコードできません: %s" % (encoding, exc))
            txt_path = output_dir / filename
            txt_files.append(txt_path)
            rendered.append((txt_path, payload, "TXT"))
            if generate_tif:
                tif_path = output_dir / (stem + ".tif")
                tif_files.append(tif_path)
                rendered.append((tif_path, _tif_payload([case]), "TIF"))
    else:
        chunks = []
        for case in cases:
            chunks.append(_render_form(
                case.form_id, case.fields, output_format,
                case.target_presence, line_ending).rstrip("\r\n"))
        # 1 ファイルにまとめる場合も「1 行 = 1 帳票」を維持する。
        content = line_ending.join(chunks) + line_ending
        try:
            payload = content.encode(encoding)
        except (LookupError, UnicodeEncodeError) as exc:
            raise LayoutTxtError("%s でエンコードできません: %s" % (encoding, exc))
        # 単一ファイル時、既定の{form_id}は入力Excel名に読み替える。
        combined_template = filename_template
        if str(combined_template or "").strip() == "{form_id}":
            combined_template = "{source}"
        representative = GeneratedCase(
            source_form_id="all", form_id="all", fields=[],
            target_presence=target_presence, pattern="all", sequence=1)
        stem = _case_stem(combined_template, representative, source_stem, False)
        txt_path = output_dir / (stem + ".txt")
        txt_files.append(txt_path)
        rendered.append((txt_path, payload, "TXT"))
        if generate_tif:
            tif_path = output_dir / (stem + ".tif")
            tif_files.append(tif_path)
            rendered.append((tif_path, _tif_payload(cases), "TIF"))

    archive_members = [path.name for path, _payload, _kind in rendered]
    tar_file = None  # type: Optional[Path]
    writes = list(rendered)
    if tar_only:
        create_tar = True
    if create_tar:
        tar_file = output_dir / (
            _tar_stem(tar_name, source_stem, list(groups.keys())) + ".tar")
        tar_item = (tar_file, _tar_payload(rendered), "TAR")
        writes = [tar_item] if tar_only else rendered + [tar_item]

    existing = [path for path, _payload, _kind in writes if path.exists()]
    if existing and not overwrite:
        raise LayoutTxtError(
            "既存TXT/TIFを上書きしません。--overwriteを付けるか出力先を変えてください: %s"
            % ", ".join(str(p) for p in existing))

    # 全件を先に検証・エンコードしてから書く。途中エラーで半端な一式を残さない。
    output_dir.mkdir(parents=True, exist_ok=True)
    for path, payload, _kind in writes:
        _atomic_write(path, payload)

    return GenerationResult(
        files=[] if tar_only else txt_files,
        tif_files=[] if tar_only else tif_files, field_count=len(fields),
        form_count=len(groups), sheet_name=actual_sheet,
        header_row=actual_header, columns=columns,
        pattern_count=len(cases), tar_file=tar_file,
        archive_members=archive_members)


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="layout-txt", description="Excel の I/J/K 列から Layout 取込用 TXT を生成します")
    parser.add_argument("excel", nargs="?", help="入力 Excel（.xlsx / .xlsm）")
    parser.add_argument("--gui", action="store_true", help="操作画面を開く")
    parser.add_argument("--out-dir", default="output/layout_txt", help="TXT/TIF 出力フォルダ")
    parser.add_argument("--sheet", default=None, help="対象シート名（省略時はアクティブシート）")
    parser.add_argument("--header-row", type=int, default=None, help="見出し行番号（省略時は自動検出）")
    parser.add_argument("--form-column", default="auto", help="FormID 列（見出し名 / 列記号 / auto）")
    parser.add_argument("--layout-column", default="auto", help="LayoutID 列（見出し名 / 列記号 / auto）")
    parser.add_argument("--field-column", default="auto", help="FieldID 列（見出し名 / 列記号 / auto）")
    parser.add_argument("--item-column", default="auto", help="項目名列（見出し名 / 列記号 / auto）")
    parser.add_argument("--data-type-column", default="I", help="データ型列（既定: I）")
    parser.add_argument("--ime-column", default="J", help="IME/入力制限列（既定: J）")
    parser.add_argument("--max-digits-column", default="K", help="最大桁数列（既定: K）")
    parser.add_argument("--input-attribute-column", default="auto",
                        help="入力属性列（見出し名 / 列記号 / auto / none）")
    parser.add_argument("--input-rule-column", default="auto",
                        help="入力規則列（見出し名 / 列記号 / auto / none）")
    parser.add_argument("--notes-column", default="auto",
                        help="補足列（見出し名 / 列記号 / auto / none）")
    parser.add_argument("--output-example-column", default="auto",
                        help="出力例列（見出し名 / 列記号 / auto / none）")
    parser.add_argument("--default-value-column", default="auto",
                        help="OCR既定値列（見出し名 / 列記号 / auto / none）")
    parser.add_argument("--profile", choices=["normal", "max", "over"], default="normal",
                        help="normal=代表値 / max=最大桁 / over=最大桁+1")
    parser.add_argument("--date-mode",
                        choices=["coverage", "cycle", "wareki", "seireki",
                                 "era-seireki", "multiple"],
                        default="coverage",
                        help="日付形式（既定: 通常Form=和暦 / 4001=4形式網羅）")
    parser.add_argument("--coverage-form-id", default="4001",
                        help="属性・日付の全網羅データを作る FormID（既定: 4001）")
    parser.add_argument(
        "--form-id", dest="form_ids", action="append", default=[],
        help="出力対象FormID。複数回またはカンマ区切りで指定（省略時は全Form）")
    parser.add_argument("--error-patterns", choices=["none", "core", "all"], default="all",
                        help="4001の異常系: none=なし / core=主要8種 / all=全24種")
    parser.add_argument(
        "--filename-template", default="{form_id}",
        help="拡張子なしファイル名。{form_id}/{output_form_id}/{pattern}/{seq}/{source}が使用可能")
    parser.add_argument("--format", dest="output_format", choices=["raw", "labeled", "tsv"],
                        default="raw", help="TXT 形式（既定: 1帳票1行・全値ダブルクォート）")
    parser.add_argument("--encoding", default="cp932", help="文字コード（既定: cp932）")
    parser.add_argument("--single-file", action="store_true", help="FormID ごとに分けず 1 ファイルにする")
    parser.add_argument("--no-tif", action="store_true", help="TXTと同名の対応TIFを生成しない")
    parser.add_argument("--tar", action="store_true", help="生成TXT/TIFをTARにもまとめる")
    parser.add_argument(
        "--tar-name", default="{source}_layout_data",
        help="TAR名（拡張子不要）。{source}/{form_id}が使用可能")
    parser.add_argument("--tar-only", action="store_true", help="散文件を残さずTARだけ出力する")
    parser.add_argument("--overwrite", action="store_true", help="既存 TXT/TIF/TAR を上書きする")
    parser.add_argument("--target-presence", default="1", help="対象有無情報（既定: 1）")
    parser.add_argument("--attribute-flag", default="0", help="項目毎の属性フラグ（既定: 0）")
    parser.add_argument(
        "--coordinates", default="auto",
        help="項目毎の座標情報（既定: autoでFORM毎に0,0,0,1から末尾を連番）")
    parser.add_argument("--lf", action="store_true", help="改行を CRLF ではなく LF にする")
    return parser


def _print_result(result: GenerationResult) -> None:
    mapping = ", ".join("%s=%s" % (name, get_column_letter(index))
                        for name, index in sorted(result.columns.items()))
    print("[OK] Layout TXT/TIF/TAR を生成しました")
    print("  シート   : %s" % result.sheet_name)
    print("  見出し行 : %d" % result.header_row)
    print("  使用列   : %s" % mapping)
    print("  FormID   : %d 件" % result.form_count)
    print("  FieldID  : %d 件" % result.field_count)
    print("  Pattern  : %d 件" % result.pattern_count)
    for path in result.txt_files:
        print("生成TXT: %s" % path)
    for path in result.tif_files:
        print("生成TIF: %s" % path)
    if result.tar_file is not None:
        print("生成TAR: %s（%d ファイル格納）" % (
            result.tar_file, len(result.archive_members)))


def main(argv: Optional[List[str]] = None) -> int:
    args = _build_parser().parse_args(argv)
    if args.gui or not args.excel:
        from .layout_txt_gui import main as gui_main
        return gui_main(initial_excel=Path(args.excel) if args.excel else None)
    try:
        selected_form_ids = []
        for value in args.form_ids:
            selected_form_ids.extend(
                item.strip() for item in value.split(",") if item.strip())
        result = generate_layout_txt(
            excel_path=Path(args.excel), output_dir=Path(args.out_dir),
            sheet_name=args.sheet, header_row=args.header_row,
            form_column=args.form_column, layout_column=args.layout_column,
            field_column=args.field_column,
            item_column=args.item_column, data_type_column=args.data_type_column,
            ime_column=args.ime_column, max_digits_column=args.max_digits_column,
            input_attribute_column=args.input_attribute_column,
            input_rule_column=args.input_rule_column,
            notes_column=args.notes_column,
            output_example_column=args.output_example_column,
            default_value_column=args.default_value_column,
            profile=args.profile, date_mode=args.date_mode,
            coverage_form_id=args.coverage_form_id,
            error_patterns=args.error_patterns,
            filename_template=args.filename_template,
            selected_form_ids=selected_form_ids or None,
            output_format=args.output_format,
            encoding=args.encoding, split_by_form=not args.single_file,
            generate_tif=not args.no_tif,
            create_tar=args.tar or args.tar_only,
            tar_name=args.tar_name, tar_only=args.tar_only,
            overwrite=args.overwrite, target_presence=args.target_presence,
            attribute_flag=args.attribute_flag, coordinates=args.coordinates,
            crlf=not args.lf)
    except LayoutTxtError as exc:
        print("[設定エラー] %s" % exc, file=sys.stderr)
        return 2
    except Exception as exc:  # noqa: BLE001 - CLI 境界で理由を表示する
        print("[エラー] Layoutデータ生成に失敗しました: %s" % exc, file=sys.stderr)
        return 1
    _print_result(result)
    return 0


if __name__ == "__main__":
    sys.exit(main())
