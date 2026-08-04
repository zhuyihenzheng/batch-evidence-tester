"""Excel 成果物（テスト証跡ブック）の生成。

シート構成:
  サマリ        … 実行環境・全ケースの判定一覧（各ケースシートへリンク）
  <ケースID>    … ケースごとの判定明細・DB スナップショット・差分表・証跡画像

DB データは画像ではなくネイティブのセルとして書く。Excel 上でフィルタ・検索・
再突合ができ、差分セルは色で示す。
"""

import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Set

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image as PILImage

from .models import ERROR, NG, OK, SKIP, CaseResult, RunResult, Table

# XML に書けない制御文字。DB 値やログに \x00 等が混ざっていると openpyxl が
# IllegalCharacterError を投げ、証跡簿全体が生成できなくなるため置換する
_ILLEGAL_XML_RE = re.compile("[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]")


def _safe_cell(ws: Worksheet, row: int, col: int, value: Any):
    """外部由来の値（DB / ログ / コンソール出力）を安全にセルへ書く。

    - 制御文字は U+FFFD に置換（IllegalCharacterError の防止）
    - 先頭が = の文字列は数式ではなく文字列として書く（数式インジェクション防止）
    """
    if isinstance(value, str):
        cleaned = _ILLEGAL_XML_RE.sub("�", value)
        cell = ws.cell(row, col, cleaned)
        if cleaned.startswith("="):
            cell.data_type = "s"  # openpyxl は "=" 始まりを数式として扱うため上書き
        return cell
    return ws.cell(row, col, value)

# --- スタイル ----------------------------------------------------------------
F_TITLE = Font(name="Meiryo UI", size=14, bold=True)
F_SECTION = Font(name="Meiryo UI", size=11, bold=True, color="FFFFFF")
F_HEADER = Font(name="Meiryo UI", size=9, bold=True)
F_BODY = Font(name="Meiryo UI", size=9)
F_MONO = Font(name="Consolas", size=9)
F_LINK = Font(name="Meiryo UI", size=9, color="0563C1", underline="single")

FILL_SECTION = PatternFill("solid", fgColor="44546A")
FILL_HEADER = PatternFill("solid", fgColor="D9E1F2")
FILL_OK = PatternFill("solid", fgColor="C6EFCE")
FILL_NG = PatternFill("solid", fgColor="FFC7CE")
FILL_SKIP = PatternFill("solid", fgColor="EDEDED")
FILL_DIFF = PatternFill("solid", fgColor="FFEB9C")
FILL_LABEL = PatternFill("solid", fgColor="F2F2F2")
FILL_ERROR = PatternFill("solid", fgColor="FFD6D6")
FILL_WARN = PatternFill("solid", fgColor="FFF0CC")
FILL_HIT = PatternFill("solid", fgColor="DDEBF7")

# Table.cell_marks の値 -> セル塗り
MARK_FILL = {
    "diff": FILL_DIFF,
    "extra": FILL_DIFF,
    "error": FILL_ERROR,
    "warn": FILL_WARN,
    "hit": FILL_HIT,
}

FONT_OK = Font(name="Meiryo UI", size=9, bold=True, color="006100")
FONT_NG = Font(name="Meiryo UI", size=9, bold=True, color="9C0006")
FONT_SKIP = Font(name="Meiryo UI", size=9, bold=True, color="808080")

_thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

VERDICT_FILL = {OK: FILL_OK, NG: FILL_NG, ERROR: FILL_NG, SKIP: FILL_SKIP}
VERDICT_FONT = {OK: FONT_OK, NG: FONT_NG, ERROR: FONT_NG, SKIP: FONT_SKIP}

MAX_COL_WIDTH = 60
PX_PER_ROW = 19  # 既定行高 (pt) をピクセル換算した概算。画像の占有行数計算に使う


def build_workbook(run: RunResult, output_path: Path, cfg: dict) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    summary = wb.create_sheet("サマリ")
    sheet_names: Dict[str, str] = {}
    for case in run.cases:
        name = _safe_sheet_name(case.case_id, set(sheet_names.values()))
        sheet_names[case.case_id] = name
        ws = wb.create_sheet(name)
        _write_case_sheet(ws, case, run, cfg)

    _write_summary_sheet(summary, run, sheet_names, cfg)
    wb.active = 0

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


# =============================================================================
# サマリ
# =============================================================================


def _case_count_text(run: RunResult) -> str:
    """実行件数の表示。絞り込みがあった場合は除外件数も明示する。

    「3 件すべて OK」とだけ書くと、実行しなかったケースの存在が
    証跡から読み取れず、全体が合格したかのように誤読される。
    """
    if run.total_available is None or run.total_available == len(run.cases):
        return str(len(run.cases))
    excluded = run.total_available - len(run.cases)
    return "%d 件（全 %d 件中。%d 件は今回実行していません）" % (
        len(run.cases), run.total_available, excluded)


def _write_summary_sheet(ws: Worksheet, run: RunResult, sheet_names: Dict[str, str], cfg: Optional[dict] = None) -> None:
    ws.sheet_view.showGridLines = False
    ws["A1"] = "バッチ自動テスト 実行結果"
    ws["A1"].font = F_TITLE
    ws.row_dimensions[1].height = 24

    meta = [
        ("実行ID", run.run_id),
        ("実行環境", run.env_name),
        ("実施者", run.tester),
        ("対象 exe", run.exe_path),
        ("DB", f"{run.db_server} / {run.db_name}"),
        ("開始日時", run.started_at.strftime("%Y-%m-%d %H:%M:%S")),
        ("終了日時", run.finished_at.strftime("%Y-%m-%d %H:%M:%S") if run.finished_at else "-"),
        ("実行対象", run.filter_description or "全ケース（絞り込みなし）"),
        ("実行ケース数", _case_count_text(run)),
        ("OK / NG", f"{run.ok_count} / {run.ng_count}"),
    ]
    row = 3
    for label, value in meta:
        ws.cell(row, 1, label).font = F_HEADER
        ws.cell(row, 1).fill = FILL_LABEL
        ws.cell(row, 1).border = BORDER
        c = ws.cell(row, 2, value)
        c.font = F_BODY
        c.border = BORDER
        row += 1

    row += 1
    ws.cell(row, 1, "総合判定").font = F_HEADER
    ws.cell(row, 1).fill = FILL_LABEL
    ws.cell(row, 1).border = BORDER
    v = ws.cell(row, 2, run.verdict)
    v.fill = VERDICT_FILL[run.verdict]
    v.font = VERDICT_FONT[run.verdict]
    v.alignment = Alignment(horizontal="center")
    v.border = BORDER

    row += 2
    headers = ["No", "ケースID", "ケース名", "区分", "判定", "NG件数", "終了コード", "所要(秒)", "実行日時", "証跡シート"]
    _write_header_row(ws, row, headers)
    row += 1

    for i, case in enumerate(run.cases, start=1):
        exec_info = case.execution
        values = [
            i,
            case.case_id,
            case.name,
            " / ".join(case.tags),
            case.verdict,
            case.ng_count,
            "TIMEOUT" if exec_info.timed_out else (exec_info.exit_code if exec_info.exit_code is not None else "-"),
            round(exec_info.elapsed_sec, 2),
            exec_info.started_at.strftime("%m-%d %H:%M:%S") if exec_info.started_at else "-",
            sheet_names.get(case.case_id, ""),
        ]
        for col, value in enumerate(values, start=1):
            c = ws.cell(row, col, value)
            c.font = F_BODY
            c.border = BORDER
            c.alignment = Alignment(vertical="center", wrap_text=(col == 3))
        verdict_cell = ws.cell(row, 5)
        verdict_cell.fill = VERDICT_FILL[case.verdict]
        verdict_cell.font = VERDICT_FONT[case.verdict]
        verdict_cell.alignment = Alignment(horizontal="center")

        link_cell = ws.cell(row, 10)
        sheet = sheet_names.get(case.case_id)
        if sheet:
            link_cell.hyperlink = f"#'{sheet}'!A1"
            link_cell.font = F_LINK
        row += 1

    ws.auto_filter.ref = f"A{row - len(run.cases) - 1}:J{row - 1}"
    _set_widths(ws, [5, 14, 38, 14, 8, 8, 11, 10, 16, 14])
    if (cfg or {}).get("freeze_header", True):
        ws.freeze_panes = ws.cell(row - len(run.cases), 1)


# =============================================================================
# ケースシート
# =============================================================================


def _write_case_sheet(ws: Worksheet, case: CaseResult, run: RunResult, cfg: dict) -> None:
    ws.sheet_view.showGridLines = False
    row = 1

    ws.cell(row, 1, f"{case.case_id}  {case.name}").font = F_TITLE
    ws.row_dimensions[row].height = 24
    row += 2

    # 出力する節だけを順に採番する（該当データが無い節は飛ばす）
    section_no = _Counter()

    # --- 実行情報 ---------------------------------------------------------
    row = _section(ws, row, section_no.next("実行情報"))
    exec_info = case.execution
    info = [
        ("目的 / 観点", case.description or "-"),
        ("区分", " / ".join(case.tags) or "-"),
        ("実行コマンド", exec_info.command or "-"),
        ("開始日時", exec_info.started_at.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3] if exec_info.started_at else "-"),
        ("終了日時", exec_info.finished_at.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3] if exec_info.finished_at else "-"),
        ("所要時間", f"{exec_info.elapsed_sec:.2f} 秒"),
        ("終了コード", "TIMEOUT" if exec_info.timed_out else str(exec_info.exit_code)),
        ("実施者 / 環境", f"{run.tester} / {run.env_name}"),
    ]
    for label, value in info:
        ws.cell(row, 1, label).font = F_HEADER
        ws.cell(row, 1).fill = FILL_LABEL
        ws.cell(row, 1).border = BORDER
        c = ws.cell(row, 2, value)
        c.font = F_MONO if label == "実行コマンド" else F_BODY
        c.border = BORDER
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        row += 1

    ws.cell(row, 1, "総合判定").font = F_HEADER
    ws.cell(row, 1).fill = FILL_LABEL
    ws.cell(row, 1).border = BORDER
    v = ws.cell(row, 2, case.verdict)
    v.fill = VERDICT_FILL[case.verdict]
    v.font = VERDICT_FONT[case.verdict]
    v.alignment = Alignment(horizontal="center")
    v.border = BORDER
    row += 2

    if case.fatal_error:
        ws.cell(row, 1, "実行時エラー").font = F_HEADER
        c = ws.cell(row, 2, case.fatal_error)
        c.font = F_BODY
        c.fill = FILL_NG
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        row += 2

    # --- 判定明細 ---------------------------------------------------------
    row = _section(ws, row, section_no.next("判定明細"))
    _write_header_row(ws, row, ["No", "確認項目", "分類", "判定", "内容"])
    row += 1
    for i, check in enumerate(case.checks, start=1):
        for col, value in enumerate([i, check.name, check.category, check.verdict, check.detail], start=1):
            c = _safe_cell(ws, row, col, value)
            c.font = F_BODY
            c.border = BORDER
            c.alignment = Alignment(vertical="center", wrap_text=(col == 5))
        vc = ws.cell(row, 4)
        vc.fill = VERDICT_FILL.get(check.verdict, FILL_SKIP)
        vc.font = VERDICT_FONT.get(check.verdict, FONT_SKIP)
        vc.alignment = Alignment(horizontal="center")
        row += 1
    if not case.checks:
        ws.cell(row, 1, "（判定項目なし: 証跡採取のみ）").font = F_BODY
        row += 1
    row += 1

    # --- DB スナップショット（ネイティブ表）--------------------------------
    # max_db_rows はここでの「表示」行数の上限。判定は全行に対して実施済み
    db_display_limit = int(cfg.get("max_db_rows", 500))
    if case.db_before or case.db_after:
        row = _section(ws, row, section_no.next("DB スナップショット（実行前 / 実行後）"))
        for table_name in sorted(set(case.db_before) | set(case.db_after)):
            before = case.db_before.get(table_name)
            after = case.db_after.get(table_name)
            if before:
                row = _write_table(ws, row, before, subtitle="実行前", display_limit=db_display_limit)
                row += 1
            if after:
                changed = _mark_changes(after, before)
                row = _write_table(ws, row, changed, subtitle="実行後", legend="※ 実行前から変化したセルを黄色で表示",
                                   display_limit=db_display_limit)
                row += 1
        row += 1

    # --- 実行ログ（ネイティブ表）------------------------------------------
    if case.log_tables:
        row = _section(ws, row, section_no.next("実行ログ（本実行で出力された分）"))
        for table in case.log_tables:
            row = _write_log_table(ws, row, table)
            row += 1
        row += 1

    # --- 差分 -------------------------------------------------------------
    diff_checks = [c for c in case.checks if c.diff_table is not None]
    if diff_checks:
        row = _section(ws, row, section_no.next("期待値との差分"))
        for check in diff_checks:
            row = _write_table(ws, row, check.diff_table)
            row += 1
        row += 1

    # --- 証跡画像 ---------------------------------------------------------
    if case.images:
        row = _section(ws, row, section_no.next("証跡画像（入出力フォルダ / ファイル内容）"))
        display_width = int(cfg.get("image_display_width", 720))
        for image in case.images:
            ws.cell(row, 1, image.title).font = F_HEADER
            ws.cell(row, 1).fill = FILL_LABEL
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            row += 1
            if image.caption:
                ws.cell(row, 1, image.caption).font = F_BODY
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
                row += 1
            row = _insert_image(ws, row, image.path, display_width)
            row += 1

    # --- コンソール出力 ---------------------------------------------------
    if exec_info.stdout or exec_info.stderr:
        row = _section(ws, row, section_no.next("標準出力 / 標準エラー"))
        for label, text in (("STDOUT", exec_info.stdout), ("STDERR", exec_info.stderr)):
            if not text.strip():
                continue
            ws.cell(row, 1, label).font = F_HEADER
            ws.cell(row, 1).fill = FILL_LABEL
            row += 1
            for line in text.splitlines()[:200]:
                c = _safe_cell(ws, row, 1, line)
                c.font = F_MONO
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
                row += 1
            row += 1

    _set_widths(ws, [22, 24, 20, 12, 46, 20, 20, 20])


# =============================================================================
# パーツ
# =============================================================================


class _Counter:
    """節見出しの連番。出力しない節を飛ばしても番号が飛ばないようにする。"""

    def __init__(self) -> None:
        self._n = 0

    def next(self, title: str) -> str:
        self._n += 1
        return f"{self._n}. {title}"


def _section(ws: Worksheet, row: int, title: str) -> int:
    c = ws.cell(row, 1, title)
    c.font = F_SECTION
    c.fill = FILL_SECTION
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    for col in range(1, 9):
        ws.cell(row, col).fill = FILL_SECTION
    ws.row_dimensions[row].height = 18
    return row + 1


def _write_header_row(ws: Worksheet, row: int, headers: List[str]) -> None:
    for col, text in enumerate(headers, start=1):
        c = ws.cell(row, col, text)
        c.font = F_HEADER
        c.fill = FILL_HEADER
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")


def _write_table(
    ws: Worksheet,
    row: int,
    table: Table,
    subtitle: str = "",
    legend: str = "",
    display_limit: Optional[int] = None,
) -> int:
    """Table をネイティブセルとして書き出す。差分セルは色付け。"""
    caption = f"{subtitle}  {table.title}".strip()
    ws.cell(row, 1, caption).font = F_HEADER
    ws.cell(row, 1).fill = FILL_LABEL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(2, len(table.columns)))
    row += 1

    note_parts = [p for p in (table.note, legend) if p]
    if table.truncated_from:
        note_parts.append(f"※ 全 {table.truncated_from} 行のうち先頭 {len(table.rows)} 行を表示")
    if note_parts:
        c = ws.cell(row, 1, "  ".join(note_parts))
        c.font = Font(name="Meiryo UI", size=8, color="808080")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(2, len(table.columns)))
        row += 1

    if not table.columns:
        ws.cell(row, 1, "（データなし）").font = F_BODY
        return row + 1

    _write_header_row(ws, row, table.columns)
    header_row = row
    row += 1

    mono = set(table.mono_columns)
    shown_rows = table.rows if display_limit is None else table.rows[:display_limit]
    for r, data_row in enumerate(shown_rows):
        for c_idx, value in enumerate(data_row):
            cell = _safe_cell(ws, row, c_idx + 1, value)
            cell.font = F_MONO if c_idx in mono else F_BODY
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")
            fill = MARK_FILL.get(table.cell_marks.get((r, c_idx), ""))
            if fill:
                cell.fill = fill
        row += 1

    if display_limit is not None and len(table.rows) > display_limit:
        c = ws.cell(row, 1,
                    f"…（表示上の省略: 全 {len(table.rows)} 行のうち先頭 {display_limit} 行のみ表示。判定は全行で実施済み）")
        c.font = Font(name="Meiryo UI", size=8, color="808080")
        ws.merge_cells(start_row=row, start_column=1, end_row=row,
                       end_column=max(2, len(table.columns)))
        row += 1

    if len(shown_rows) > 8:
        ws.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(len(table.columns))}{row - 1}"
        )
    return row


def _write_log_table(ws: Worksheet, row: int, table: Table) -> int:
    """ログを「行番号 + 本文」で書く。

    本文は B〜H を結合して表示幅を稼ぐ。列幅は他の表（DB スナップショット等）と
    共有されるため、B 列だけを広げるとレイアウトが崩れてしまう。
    """
    ws.cell(row, 1, f"ログ: {table.title}").font = F_HEADER
    ws.cell(row, 1).fill = FILL_LABEL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1

    if table.note:
        c = ws.cell(row, 1, table.note)
        c.font = Font(name="Meiryo UI", size=8, color="808080")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        row += 1

    _write_header_row(ws, row, ["行", "内容"])
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    for col in range(2, 9):
        ws.cell(row, col).fill = FILL_HEADER
        ws.cell(row, col).border = BORDER
    row += 1

    for r, (line_no, content) in enumerate(table.rows):
        fill = MARK_FILL.get(table.cell_marks.get((r, 1), ""))

        no_cell = ws.cell(row, 1, line_no)
        no_cell.font = F_BODY
        no_cell.alignment = Alignment(horizontal="right", vertical="center")
        no_cell.border = BORDER

        text_cell = _safe_cell(ws, row, 2, content)
        text_cell.font = F_MONO
        text_cell.alignment = Alignment(vertical="center")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        for col in range(2, 9):
            ws.cell(row, col).border = BORDER
            if fill:
                ws.cell(row, col).fill = fill
        if fill:
            no_cell.fill = fill
        row += 1

    return row


def _mark_changes(after: Table, before: Optional[Table]) -> Table:
    """実行後スナップショットのうち、実行前から変化したセルにマークを付ける。"""
    if before is None or before.columns != after.columns:
        return after
    marks = dict(after.cell_marks)
    for r, row in enumerate(after.rows):
        if r >= len(before.rows):
            for c in range(len(row)):
                marks[(r, c)] = "extra"
            continue
        for c, value in enumerate(row):
            if c < len(before.rows[r]) and str(before.rows[r][c]) != str(value):
                marks[(r, c)] = "diff"
    return Table(
        title=after.title,
        columns=after.columns,
        rows=after.rows,
        cell_marks=marks,
        note=after.note,
        truncated_from=after.truncated_from,
    )


def _insert_image(ws: Worksheet, row: int, path: Path, display_width: int) -> int:
    """画像を貼り、占有した行数分だけ進めた次の行番号を返す。"""
    if not path.exists():
        ws.cell(row, 1, f"（画像が見つかりません: {path}）").font = F_BODY
        return row + 1

    with PILImage.open(path) as im:
        w, h = im.size
    scale = min(1.0, display_width / w)
    disp_w, disp_h = int(w * scale), int(h * scale)

    img = XLImage(str(path))
    img.width, img.height = disp_w, disp_h
    ws.add_image(img, f"A{row}")

    return row + max(2, int(disp_h / PX_PER_ROW) + 2)


def _set_widths(ws: Worksheet, widths: List[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(w, MAX_COL_WIDTH)


def _safe_sheet_name(base: str, used: Set[str]) -> str:
    """Excel のシート名制約（31 文字 / 禁止文字）に収める。"""
    name = "".join("_" if ch in "[]:*?/\\" else ch for ch in base)[:31] or "SHEET"
    if name not in used:
        return name
    for i in range(2, 100):
        suffix = f"_{i}"
        candidate = name[: 31 - len(suffix)] + suffix
        if candidate not in used:
            return candidate
    return f"SHEET_{datetime.now():%H%M%S}"
