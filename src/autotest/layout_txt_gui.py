# -*- coding: utf-8 -*-
"""Excel -> Layout TXT/TIF/TAR 生成ツールの単独GUI。"""

import csv
import re
import sys
from collections import OrderedDict
from pathlib import Path
from typing import Optional

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
except ImportError as exc:  # pragma: no cover - GUI無し環境
    raise SystemExit("画面を開けません（tkinterが見つかりません）: %s" % exc)

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .layout_txt import (
    LayoutTxtError,
    _print_result,
    generate_layout_txt,
    read_layout_fields,
    render_form_tif_payload,
    render_form_txt_text,
    resolve_form_filename_stem,
    save_layout_default_values,
)
from .layout_tar import (
    IMAGE_EXTENSIONS,
    LayoutTarError,
    PackageItem,
    base_name_from_front,
    build_image_tar,
    format_package_tar_name,
    format_extra_fields,
    matching_back_image,
    matching_recognition_file,
    parse_extra_fields,
    parse_manifest_columns,
)
from .layout_txt_settings import (
    LayoutGuiSettingsError,
    default_settings_path,
    load_settings,
    save_settings,
)


PROFILE_LABELS = (
    ("通常値（推奨）", "normal"),
    ("最大桁数ちょうど", "max"),
    ("最大桁数 + 1（異常系）", "over"),
)

FORMAT_LABELS = (
    ("1帳票1行（全値ダブルクォート）", "raw"),
    ("ラベル付き（確認用）", "labeled"),
    ("TSV（1項目1行）", "tsv"),
)

DATE_MODE_LABELS = (
    ("通常Form=和暦 / 4001=日付3・カレンダー4形式（推奨）", "coverage"),
    ("日付3形式 / カレンダー4形式を順番使用", "cycle"),
    ("和暦: 5/8/6/1", "wareki"),
    ("西暦: /2026/6/1", "seireki"),
    ("元号+西暦: 5/2026/6/1", "era-seireki"),
    ("複数（カレンダーのみ）: 5/8/6/1|5/8/6/2", "multiple"),
)

ERROR_PATTERN_LABELS = (
    ("全24 Pattern（推奨）", "all"),
    ("主要8 Pattern", "core"),
    ("異常Patternなし", "none"),
)

CSV_STYLE_LABELS = (
    ("固定10列（画像ごと・ヘッダーなし）", "image_list"),
    ("カスタム列（1件ごと・ヘッダーあり）", "custom"),
)

TREE_COLUMNS = (
    "include", "row", "layout_id", "field_id", "item_name",
    "data_type", "ime_name", "max_digits", "value",
    "attribute_flag", "coordinates", "input_attribute", "input_rule",
    "notes", "output_example",
)

TREE_HEADINGS = {
    "include": "出力", "row": "Excel行", "layout_id": "LAYOUT_ID",
    "field_id": "ELEMENT_ID", "item_name": "ITEM_NAME",
    "data_type": "DATA_TYPE", "ime_name": "IME", "max_digits": "MAX",
    "value": "OCR生成値（編集可）", "attribute_flag": "属性",
    "coordinates": "座標", "input_attribute": "入力属性",
    "input_rule": "入力規則", "notes": "補足", "output_example": "出力例",
}

TREE_WIDTHS = {
    "include": 48, "row": 62, "layout_id": 90, "field_id": 100,
    "item_name": 170, "data_type": 120, "ime_name": 110,
    "max_digits": 65, "value": 230, "attribute_flag": 60,
    "coordinates": 120, "input_attribute": 120, "input_rule": 160,
    "notes": 180, "output_example": 180,
}

EDITABLE_COLUMNS = {
    "include", "field_id", "value", "attribute_flag", "coordinates",
}

PACKAGE_COLUMNS = (
    "include", "scan_batch_id", "image_sequence", "base_name",
    "front_image", "back_image", "arrival_date", "form_id",
    "application_number", "reception_number", "format_id", "delivery_date",
    "delivery_shot",
    "front_recognition", "back_recognition", "back_recognition_result",
    "related_file", "extra_fields", "source",
)

PACKAGE_HEADINGS = {
    "include": "TAR", "scan_batch_id": "CSV位置1（編集可）",
    "image_sequence": "CSV位置2（編集可）",
    "base_name": "CSV位置3・画像基礎名（編集可）",
    "front_image": "CSV位置3・正面画像", "front_recognition": "正面TXT（FORM生成）",
    "back_image": "CSV位置3・背面画像", "back_recognition": "背面TXT（1項目）",
    "back_recognition_result": "背面認識値（編集可）",
    "arrival_date": "CSV位置4（編集可）", "form_id": "CSV位置5（編集可）",
    "application_number": "CSV位置6（編集可）",
    "reception_number": "CSV位置7（編集可）", "format_id": "CSV位置8（編集可）",
    "delivery_date": "CSV位置9（編集可）", "delivery_shot": "CSV位置10（編集可）",
    "related_file": "関連ファイル名", "extra_fields": "CSV追加項目 key=value;...",
    "source": "画像元",
}

PACKAGE_WIDTHS = {
    "include": 45, "scan_batch_id": 125, "image_sequence": 90,
    "form_id": 80, "base_name": 150,
    "front_image": 160, "front_recognition": 165,
    "back_image": 160, "back_recognition": 145,
    "back_recognition_result": 125, "related_file": 160,
    "arrival_date": 95, "application_number": 100,
    "reception_number": 100, "format_id": 95,
    "delivery_date": 105, "delivery_shot": 125,
    "extra_fields": 230, "source": 260,
}

PACKAGE_EDITABLE_COLUMNS = {
    "include", "scan_batch_id", "image_sequence", "form_id", "base_name",
    "arrival_date", "application_number", "reception_number", "format_id",
    "delivery_date", "delivery_shot", "back_recognition_result",
    "related_file", "extra_fields",
}

SETTING_VARIABLE_NAMES = (
    "excel_var", "output_var", "sheet_var", "header_var",
    "form_col_var", "layout_col_var", "field_col_var", "item_col_var",
    "type_col_var", "ime_col_var", "max_col_var",
    "input_attribute_col_var", "input_rule_col_var", "notes_col_var",
    "output_example_col_var", "default_value_col_var", "use_default_value_var",
    "profile_var", "date_mode_var",
    "coverage_form_var", "error_pattern_var", "filename_template_var",
    "generate_tif_var", "create_tar_var", "tar_only_var", "tar_name_var",
    "format_var", "encoding_var", "split_var", "overwrite_var",
    "package_tar_name_var", "package_result_var", "package_include_back_var",
    "manifest_name_var", "manifest_columns_var", "manifest_encoding_var",
    "manifest_style_var",
    "package_scan_batch_id_var", "package_arrival_date_var",
    "package_application_number_var", "package_reception_number_var",
    "package_format_id_var", "package_delivery_date_var",
    "package_delivery_shot_var",
)


def _value_for_label(pairs, label):
    for item_label, value in pairs:
        if item_label == label:
            return value
    return pairs[0][1]


class LayoutTxtGui(object):
    def __init__(self, root, initial_excel: Optional[Path] = None) -> None:
        self.root = root
        root.title("Layout TXT / TIF / TAR 生成ツール")
        self._configure_window_for_screen()

        self.excel_var = tk.StringVar(value="")
        self.output_var = tk.StringVar(value="")
        self.sheet_var = tk.StringVar(value="")
        self.header_var = tk.StringVar(value="自動")
        self.form_col_var = tk.StringVar(value="auto")
        self.layout_col_var = tk.StringVar(value="auto")
        self.field_col_var = tk.StringVar(value="auto")
        self.item_col_var = tk.StringVar(value="auto")
        self.type_col_var = tk.StringVar(value="I")
        self.ime_col_var = tk.StringVar(value="J")
        self.max_col_var = tk.StringVar(value="K")
        self.input_attribute_col_var = tk.StringVar(value="auto")
        self.input_rule_col_var = tk.StringVar(value="auto")
        self.notes_col_var = tk.StringVar(value="auto")
        self.output_example_col_var = tk.StringVar(value="auto")
        self.default_value_col_var = tk.StringVar(value="auto")
        self.use_default_value_var = tk.BooleanVar(value=True)
        self.profile_var = tk.StringVar(value=PROFILE_LABELS[0][0])
        self.date_mode_var = tk.StringVar(value=DATE_MODE_LABELS[0][0])
        self.coverage_form_var = tk.StringVar(value="4001")
        self.error_pattern_var = tk.StringVar(value=ERROR_PATTERN_LABELS[0][0])
        self.filename_template_var = tk.StringVar(value="{form_id}")
        self.generate_tif_var = tk.BooleanVar(value=True)
        self.create_tar_var = tk.BooleanVar(value=True)
        self.tar_only_var = tk.BooleanVar(value=False)
        self.tar_name_var = tk.StringVar(value="{source}_layout_data")
        self.format_var = tk.StringVar(value=FORMAT_LABELS[0][0])
        self.encoding_var = tk.StringVar(value="cp932")
        self.split_var = tk.BooleanVar(value=True)
        self.overwrite_var = tk.BooleanVar(value=False)
        self.package_tar_name_var = tk.StringVar(value="image_package")
        self.package_result_var = tk.StringVar(value="1")
        self.package_include_back_var = tk.BooleanVar(value=False)
        self.manifest_name_var = tk.StringVar(value="file_list.csv")
        self.manifest_columns_var = tk.StringVar(
            value="正面画像=front_image_file,背面画像=back_image_file,"
                  "FORM_ID=form_id,背面認識結果=back_recognition_result,"
                  "関連ファイル=related_file")
        self.manifest_encoding_var = tk.StringVar(value="cp932")
        self.manifest_style_var = tk.StringVar(value=CSV_STYLE_LABELS[0][0])
        self.package_scan_batch_id_var = tk.StringVar(value="")
        self.package_arrival_date_var = tk.StringVar(value="")
        self.package_application_number_var = tk.StringVar(value="")
        self.package_reception_number_var = tk.StringVar(value="")
        self.package_format_id_var = tk.StringVar(value="")
        self.package_delivery_date_var = tk.StringVar(value="")
        self.package_delivery_shot_var = tk.StringVar(value="")
        self.form_var = tk.StringVar(value="")
        self.form_summary_var = tk.StringVar(value="Excel定義を読み込んでください。")
        self.status_var = tk.StringVar(value="Excelを選択してください。")
        self.settings_path = default_settings_path()
        self.settings_load_error = ""
        self.visible_columns = list(TREE_COLUMNS)
        self.visible_column_vars = OrderedDict()

        self.loaded_fields = []  # type: List[LayoutField]
        self.form_fields = OrderedDict()  # type: OrderedDict[str, List[LayoutField]]
        self.cell_editor = None
        self.editor_context = None
        self.package_items = OrderedDict()
        self.package_sequence = 0
        self.package_editor = None
        self.package_editor_context = None

        self._load_persisted_settings()
        if initial_excel:
            self.excel_var.set(str(initial_excel))
            if not self.output_var.get().strip():
                self.output_var.set(str(initial_excel.parent / "layout_txt"))

        self._build()
        root.protocol("WM_DELETE_WINDOW", self._close)
        if Path(self.excel_var.get().strip()).is_file():
            self._load_sheets()
            self._read_definitions(show_errors=False)
        if self.settings_load_error:
            self.status_var.set(self.settings_load_error)

    def _configure_window_for_screen(self) -> None:
        screen_width = max(320, int(self.root.winfo_screenwidth()))
        screen_height = max(300, int(self.root.winfo_screenheight()))
        available_width = max(320, screen_width - 40)
        available_height = max(300, screen_height - 100)
        width = min(1380, available_width)
        height = min(980, available_height)
        left = max(0, (screen_width - width) // 2)
        top = max(0, (screen_height - height) // 2)
        self.root.geometry("%dx%d+%d+%d" % (width, height, left, top))
        self.root.minsize(min(900, width), min(620, height))

    def _sync_content_scrollregion(self, _event=None) -> None:
        if self.content_canvas is not None:
            self.content_canvas.configure(
                scrollregion=self.content_canvas.bbox("all"))

    def _fit_content_to_canvas(self, event) -> None:
        if self.content_canvas is None or self.content_window is None:
            return
        requested_width = self.content_frame.winfo_reqwidth()
        requested_height = self.content_frame.winfo_reqheight()
        # Treeview の要求幅は全列幅の合計になるが、表には専用の横スクロールがある。
        # ページ全体をその幅（約2500px）まで広げず、従来の設計幅を上限にする。
        target_width = max(event.width, min(1380, requested_width))
        target_height = max(event.height, requested_height)
        current_width = float(self.content_canvas.itemcget(
            self.content_window, "width") or 0)
        current_height = float(self.content_canvas.itemcget(
            self.content_window, "height") or 0)
        changes = {}
        if int(current_width) != target_width:
            changes["width"] = target_width
        if int(current_height) != target_height:
            changes["height"] = target_height
        if changes:
            self.content_canvas.itemconfigure(self.content_window, **changes)
        self._sync_content_scrollregion()

    def _load_persisted_settings(self) -> None:
        try:
            settings = load_settings(self.settings_path)
        except LayoutGuiSettingsError as exc:
            self.settings_load_error = str(exc)
            return
        values = settings.get("values", {})
        if not isinstance(values, dict):
            values = {}
        for name in SETTING_VARIABLE_NAMES:
            if name in values:
                getattr(self, name).set(values[name])
        visible = settings.get("visible_columns", list(TREE_COLUMNS))
        if isinstance(visible, list):
            filtered = [name for name in TREE_COLUMNS if name in visible]
            self.visible_columns = filtered or ["item_name"]

    def _settings_payload(self) -> dict:
        values = {}
        for name in SETTING_VARIABLE_NAMES:
            values[name] = getattr(self, name).get()
        return {
            "version": 1,
            "values": values,
            "visible_columns": list(self.visible_columns),
        }

    def _save_persisted_settings(self, show_message=False) -> bool:
        self._apply_visible_columns(save=False)
        try:
            path = save_settings(self._settings_payload(), self.settings_path)
        except LayoutGuiSettingsError as exc:
            self.status_var.set(str(exc))
            if show_message:
                messagebox.showerror("設定保存エラー", str(exc), parent=self.root)
            return False
        self.status_var.set("設定を保存しました: %s" % path)
        if show_message:
            messagebox.showinfo(
                "設定保存", "現在の画面設定を保存しました。\n\n%s" % path,
                parent=self.root)
        return True

    def _close(self) -> None:
        if self._save_persisted_settings(show_message=False):
            self.root.destroy()
            return
        if messagebox.askyesno(
                "設定保存エラー", "設定を保存できませんでした。保存せず閉じますか？",
                parent=self.root):
            self.root.destroy()

    def _build(self) -> None:
        shell = ttk.Frame(self.root)
        shell.pack(fill="both", expand=True)
        shell.rowconfigure(0, weight=1)
        shell.columnconfigure(0, weight=1)

        self.content_canvas = tk.Canvas(shell, highlightthickness=0, borderwidth=0)
        page_ybar = ttk.Scrollbar(
            shell, orient="vertical", command=self.content_canvas.yview)
        page_xbar = ttk.Scrollbar(
            shell, orient="horizontal", command=self.content_canvas.xview)
        self.content_canvas.configure(
            yscrollcommand=page_ybar.set, xscrollcommand=page_xbar.set)
        self.content_canvas.grid(row=0, column=0, sticky="nsew")
        page_ybar.grid(row=0, column=1, sticky="ns")
        page_xbar.grid(row=1, column=0, sticky="we")

        outer = ttk.Frame(self.content_canvas, padding=10)
        self.content_frame = outer
        self.content_window = self.content_canvas.create_window(
            (0, 0), window=outer, anchor="nw")
        outer.bind("<Configure>", self._sync_content_scrollregion)
        self.content_canvas.bind("<Configure>", self._fit_content_to_canvas)
        outer.columnconfigure(0, weight=1)
        outer.rowconfigure(5, weight=1)

        ttk.Label(outer, text="Excel → Layout TXT / TIF / TAR",
                  font=("", 16, "bold")).grid(row=0, column=0, sticky="w")

        source = ttk.LabelFrame(outer, text="入力・列設定", padding=8)
        source.grid(row=1, column=0, sticky="we", pady=(8, 0))
        for column in range(8):
            source.columnconfigure(column, weight=1 if column in (1, 3, 5) else 0)

        ttk.Label(source, text="入力Excel:").grid(row=0, column=0, sticky="w")
        ttk.Entry(source, textvariable=self.excel_var).grid(
            row=0, column=1, columnspan=6, sticky="we", padx=(6, 6))
        ttk.Button(source, text="参照...", command=self._choose_excel).grid(row=0, column=7)

        ttk.Label(source, text="出力先:").grid(row=1, column=0, sticky="w", pady=(7, 0))
        ttk.Entry(source, textvariable=self.output_var).grid(
            row=1, column=1, columnspan=6, sticky="we", padx=(6, 6), pady=(7, 0))
        ttk.Button(source, text="参照...", command=self._choose_output).grid(
            row=1, column=7, pady=(7, 0))

        ttk.Label(source, text="シート:").grid(row=2, column=0, sticky="w", pady=(7, 0))
        self.sheet_box = ttk.Combobox(
            source, textvariable=self.sheet_var, state="readonly", width=24)
        self.sheet_box.grid(row=2, column=1, sticky="we", padx=(6, 14), pady=(7, 0))
        self.sheet_box.bind("<<ComboboxSelected>>", self._definition_setting_changed)
        ttk.Label(source, text="見出し行:").grid(row=2, column=2, sticky="e", pady=(7, 0))
        ttk.Entry(source, textvariable=self.header_var, width=9).grid(
            row=2, column=3, sticky="w", padx=(6, 14), pady=(7, 0))
        ttk.Button(source, text="Excel定義を読込", command=self._read_definitions).grid(
            row=2, column=6, columnspan=2, sticky="e", pady=(7, 0))

        required_column_items = (
            ("FormID", self.form_col_var), ("LayoutID", self.layout_col_var),
            ("FieldID", self.field_col_var), ("ITEM_NAME", self.item_col_var),
            ("DATA_TYPE", self.type_col_var), ("IME", self.ime_col_var),
            ("MAX", self.max_col_var),
        )
        optional_column_items = (
            ("入力属性", self.input_attribute_col_var),
            ("入力規則", self.input_rule_col_var),
            ("補足", self.notes_col_var),
            ("出力例", self.output_example_col_var),
            ("OCR既定値", self.default_value_col_var),
        )
        columns_frame = ttk.Frame(source)
        columns_frame.grid(row=3, column=0, columnspan=8, sticky="we", pady=(7, 0))
        for index, (label, variable) in enumerate(required_column_items):
            ttk.Label(columns_frame, text=label + ":").grid(row=0, column=index * 2, sticky="w")
            ttk.Entry(columns_frame, textvariable=variable, width=9).grid(
                row=0, column=index * 2 + 1, sticky="w", padx=(3, 10))
        for index, (label, variable) in enumerate(optional_column_items):
            ttk.Label(columns_frame, text=label + ":").grid(
                row=1, column=index * 2, sticky="w", pady=(6, 0))
            ttk.Entry(columns_frame, textvariable=variable, width=12).grid(
                row=1, column=index * 2 + 1, sticky="w", padx=(3, 10), pady=(6, 0))
        ttk.Label(
            columns_frame, text="任意列は auto / 列記号 / 見出し名 / none",
            foreground="#666").grid(
                row=1, column=10, columnspan=4, sticky="w", pady=(6, 0))
        ttk.Checkbutton(
            columns_frame, text="ExcelのOCR既定値を使用",
            variable=self.use_default_value_var,
            command=self._default_value_usage_changed).grid(
                row=2, column=0, columnspan=4, sticky="w", pady=(6, 0))
        ttk.Label(
            columns_frame,
            text="OFFにすると既定値列を残したまま自動生成値へ戻します。",
            foreground="#666").grid(
                row=2, column=4, columnspan=10, sticky="w", pady=(6, 0))

        generation = ttk.LabelFrame(outer, text="生成・ファイル設定", padding=8)
        generation.grid(row=2, column=0, sticky="we", pady=(8, 0))
        for column in range(10):
            generation.columnconfigure(
                column, weight=1 if column in (1, 3, 5, 7, 9) else 0)

        ttk.Label(generation, text="データ:").grid(row=0, column=0, sticky="w")
        ttk.Combobox(
            generation, textvariable=self.profile_var,
            values=[label for label, _value in PROFILE_LABELS],
            state="readonly", width=19).grid(row=0, column=1, sticky="we", padx=(5, 12))
        ttk.Label(generation, text="日付:").grid(row=0, column=2, sticky="w")
        ttk.Combobox(
            generation, textvariable=self.date_mode_var,
            values=[label for label, _value in DATE_MODE_LABELS],
            state="readonly", width=27).grid(row=0, column=3, sticky="we", padx=(5, 12))
        ttk.Label(generation, text="全網羅Form:").grid(row=0, column=4, sticky="w")
        ttk.Entry(generation, textvariable=self.coverage_form_var, width=10).grid(
            row=0, column=5, sticky="w", padx=(5, 12))
        ttk.Label(generation, text="全出力Pattern:").grid(row=0, column=6, sticky="w")
        ttk.Combobox(
            generation, textvariable=self.error_pattern_var,
            values=[label for label, _value in ERROR_PATTERN_LABELS],
            state="readonly", width=18).grid(row=0, column=7, sticky="we", padx=(5, 0))

        ttk.Label(generation, text="TXT名:").grid(row=1, column=0, sticky="w", pady=(7, 0))
        ttk.Entry(generation, textvariable=self.filename_template_var).grid(
            row=1, column=1, sticky="we", padx=(5, 12), pady=(7, 0))
        ttk.Label(generation, text="画像＋TXT TAR名:").grid(
            row=1, column=2, sticky="w", pady=(7, 0))
        ttk.Entry(generation, textvariable=self.tar_name_var).grid(
            row=1, column=3, sticky="we", padx=(5, 12), pady=(7, 0))
        ttk.Label(generation, text="画像＋CSV TAR名:").grid(
            row=1, column=4, sticky="w", pady=(7, 0))
        ttk.Entry(generation, textvariable=self.package_tar_name_var).grid(
            row=1, column=5, sticky="we", padx=(5, 12), pady=(7, 0))
        ttk.Label(generation, text="形式:").grid(row=1, column=6, sticky="w", pady=(7, 0))
        ttk.Combobox(
            generation, textvariable=self.format_var,
            values=[label for label, _value in FORMAT_LABELS],
            state="readonly", width=16).grid(
                row=1, column=7, sticky="we", padx=(5, 12), pady=(7, 0))
        ttk.Label(generation, text="文字コード:").grid(row=1, column=8, sticky="w", pady=(7, 0))
        ttk.Combobox(
            generation, textvariable=self.encoding_var,
            values=("cp932", "utf-8-sig", "utf-8"),
            state="readonly", width=12).grid(
                row=1, column=9, sticky="we", padx=(5, 0), pady=(7, 0))

        flags = ttk.Frame(generation)
        flags.grid(row=2, column=0, columnspan=10, sticky="w", pady=(7, 0))
        ttk.Checkbutton(flags, text="FormIDごとに分割", variable=self.split_var).pack(
            side="left", padx=(0, 14))
        ttk.Checkbutton(flags, text="同名TIFを生成", variable=self.generate_tif_var).pack(
            side="left", padx=(0, 14))
        ttk.Checkbutton(flags, text="TARを生成", variable=self.create_tar_var).pack(
            side="left", padx=(0, 14))
        ttk.Checkbutton(flags, text="TARだけ残す", variable=self.tar_only_var,
                        command=self._tar_only_changed).pack(side="left", padx=(0, 14))
        ttk.Checkbutton(flags, text="既存ファイルを上書き",
                        variable=self.overwrite_var).pack(side="left")
        ttk.Button(
            flags, text="設定を保存",
            command=lambda: self._save_persisted_settings(show_message=True)).pack(
                side="left", padx=(16, 0))

        ttk.Label(
            generation,
            text="TXT名: {form_id}/{pattern}/{seq:02d}/{source}  両TAR名: {form_id}/{source}",
            foreground="#666").grid(row=3, column=0, columnspan=10, sticky="w", pady=(5, 0))
        ttk.Label(
            generation, text="設定保存先: %s" % self.settings_path,
            foreground="#666").grid(
                row=4, column=0, columnspan=10, sticky="w", pady=(3, 0))

        selector = ttk.LabelFrame(outer, text="FORM_IDを選択して内容を編集", padding=8)
        selector.grid(row=3, column=0, sticky="we", pady=(8, 0))
        selector.columnconfigure(6, weight=1)
        ttk.Label(selector, text="FORM_ID:").grid(row=0, column=0, sticky="w")
        self.form_box = ttk.Combobox(
            selector, textvariable=self.form_var, state="readonly", width=18)
        self.form_box.grid(row=0, column=1, sticky="w", padx=(6, 10))
        self.form_box.bind("<<ComboboxSelected>>", self._show_selected_form)
        ttk.Button(selector, text="表示", command=self._show_selected_form).grid(row=0, column=2)
        column_button = ttk.Menubutton(selector, text="表示列...")
        column_button.grid(row=0, column=3, padx=(10, 10))
        column_menu = tk.Menu(column_button, tearoff=False)
        for name in TREE_COLUMNS:
            variable = tk.BooleanVar(value=name in self.visible_columns)
            self.visible_column_vars[name] = variable
            column_menu.add_checkbutton(
                label=TREE_HEADINGS[name], variable=variable,
                command=self._apply_visible_columns)
        column_menu.add_separator()
        column_menu.add_command(label="全列を表示", command=self._show_all_columns)
        column_button.configure(menu=column_menu)
        self.defaults_button = ttk.Button(
            selector, text="編集OCR値をExcelへ既定値保存",
            command=self._save_defaults_to_excel)
        self.defaults_button.grid(row=0, column=4, padx=(0, 10))
        self.add_form_button = ttk.Button(
            selector, text="表示中FORMを出力リストへ追加",
            command=self._add_current_form_to_package)
        self.add_form_button.grid(row=0, column=5, padx=(0, 10))
        ttk.Label(selector, textvariable=self.form_summary_var, foreground="#444").grid(
            row=0, column=6, sticky="w")
        ttk.Label(
            selector,
            text="出力/OCR値/属性/ELEMENT_ID/座標はダブルクリックで編集できます。"
                 " 保存ボタンは表示中FORMのOCR値を入力Excelへ書き戻します。",
            foreground="#666").grid(row=1, column=0, columnspan=7, sticky="w", pady=(5, 0))

        package = ttk.LabelFrame(
            outer, text="TAR出力リスト（複数FORM・外部画像を追加して最後にまとめて梱包）",
            padding=8)
        package.grid(row=4, column=0, sticky="we", pady=(8, 0))
        package.columnconfigure(0, weight=1)

        package_toolbar = ttk.Frame(package)
        package_toolbar.grid(row=0, column=0, sticky="we")
        ttk.Label(package_toolbar, text="背面認識値（1項目）:").pack(side="left")
        ttk.Entry(
            package_toolbar, textvariable=self.package_result_var, width=8).pack(
                side="left", padx=(5, 12))
        ttk.Checkbutton(
            package_toolbar, text="背面画像あり（Fと同じ基礎名のRを追加）",
            variable=self.package_include_back_var).pack(side="left", padx=(0, 12))
        ttk.Button(
            package_toolbar, text="外部の正面画像を追加...",
            command=self._add_external_front_images).pack(side="left", padx=(0, 8))
        ttk.Button(
            package_toolbar, text="選択行の背面画像を設定...",
            command=self._set_back_image_for_selected).pack(side="left", padx=(0, 8))
        ttk.Button(
            package_toolbar, text="選択行を削除",
            command=self._remove_package_items).pack(side="left", padx=(0, 8))
        ttk.Button(
            package_toolbar, text="全消去",
            command=self._clear_package_items).pack(side="left")

        csv_defaults = ttk.Frame(package)
        csv_defaults.grid(row=1, column=0, sticky="we", pady=(6, 0))
        ttk.Label(csv_defaults, text="CSV初期値\n（追加時に反映）").pack(
            side="left", padx=(0, 10))
        default_fields = (
            ("CSV位置1", self.package_scan_batch_id_var, 14),
            ("CSV位置4", self.package_arrival_date_var, 9),
            ("CSV位置6", self.package_application_number_var, 11),
            ("CSV位置7", self.package_reception_number_var, 16),
            ("CSV位置8", self.package_format_id_var, 5),
            ("CSV位置9", self.package_delivery_date_var, 9),
            ("CSV位置10", self.package_delivery_shot_var, 5),
        )
        for label, variable, width in default_fields:
            field = ttk.Frame(csv_defaults)
            field.pack(side="left", padx=(0, 8))
            ttk.Label(field, text=label).pack(anchor="w")
            ttk.Entry(field, textvariable=variable, width=width).pack(anchor="w")
        ttk.Label(
            csv_defaults, text="位置2・3・5は追加内容から設定",
            foreground="#666").pack(side="left", padx=(0, 8))
        ttk.Button(
            csv_defaults, text="選択行／全行へ反映",
            command=self._apply_package_csv_defaults).pack(side="right")

        package_table = ttk.Frame(package)
        package_table.grid(row=2, column=0, sticky="we", pady=(6, 0))
        package_table.columnconfigure(0, weight=1)
        self.package_tree = ttk.Treeview(
            package_table, columns=PACKAGE_COLUMNS, show="headings",
            selectmode="extended", height=3)
        for name in PACKAGE_COLUMNS:
            self.package_tree.heading(name, text=PACKAGE_HEADINGS[name])
            self.package_tree.column(
                name, width=PACKAGE_WIDTHS[name], minwidth=42,
                stretch=name in ("extra_fields", "source"))
        package_ybar = ttk.Scrollbar(
            package_table, orient="vertical", command=self.package_tree.yview)
        package_xbar = ttk.Scrollbar(
            package_table, orient="horizontal", command=self.package_tree.xview)
        self.package_tree.configure(
            yscrollcommand=package_ybar.set, xscrollcommand=package_xbar.set)
        self.package_tree.grid(row=0, column=0, sticky="we")
        package_ybar.grid(row=0, column=1, sticky="ns")
        package_xbar.grid(row=1, column=0, sticky="we")
        self.package_tree.bind("<Double-1>", self._begin_package_cell_edit)

        package_options = ttk.Frame(package)
        package_options.grid(row=3, column=0, sticky="we", pady=(7, 0))
        package_options.columnconfigure(1, weight=1)
        package_options.columnconfigure(3, weight=1)
        package_options.columnconfigure(5, weight=1)
        ttk.Label(package_options, text="一覧CSV名:").grid(row=0, column=0, sticky="w")
        ttk.Entry(
            package_options, textvariable=self.manifest_name_var, width=18).grid(
                row=0, column=1, sticky="we", padx=(5, 12))
        ttk.Label(package_options, text="CSV文字コード:").grid(row=0, column=2, sticky="w")
        ttk.Combobox(
            package_options, textvariable=self.manifest_encoding_var,
            values=("cp932", "utf-8-sig", "utf-8"), state="readonly", width=11).grid(
                row=0, column=3, sticky="w", padx=(5, 12))
        ttk.Label(
            package_options, text="TARと同名の確認用フォルダも生成",
            foreground="#666").grid(row=0, column=4, columnspan=4, sticky="e")

        ttk.Label(package_options, text="CSV形式:").grid(
            row=1, column=0, sticky="w", pady=(6, 0))
        ttk.Combobox(
            package_options, textvariable=self.manifest_style_var,
            values=[label for label, _value in CSV_STYLE_LABELS],
            state="readonly", width=27).grid(
                row=1, column=1, sticky="w", padx=(5, 12), pady=(6, 0))
        ttk.Label(package_options, text="カスタムCSV列:").grid(
            row=1, column=2, sticky="w", pady=(6, 0))
        ttk.Entry(
            package_options, textvariable=self.manifest_columns_var).grid(
                row=1, column=3, columnspan=3, sticky="we", padx=(5, 12), pady=(6, 0))
        ttk.Button(
            package_options, text="選択画像＋認識TXTをTAR化",
            command=lambda: self._run_package(include_manifest=False)).grid(
                row=1, column=6, sticky="e", padx=(0, 8), pady=(6, 0))
        ttk.Button(
            package_options, text="選択画像＋一覧CSVをTAR化",
            command=lambda: self._run_package(include_manifest=True)).grid(
                row=1, column=7, sticky="e", pady=(6, 0))
        ttk.Label(
            package_options,
            text="固定10列: CSV位置1～10は一覧でダブルクリック編集できます。"
                 "位置3は画像基礎名を編集すると正面・背面へ反映されます。"
                 "画像ごとに1行、ヘッダーなし、全項目を引用します。",
            foreground="#666").grid(
                row=2, column=0, columnspan=8, sticky="w", pady=(4, 0))

        table_frame = ttk.Frame(outer)
        table_frame.grid(row=5, column=0, sticky="nsew", pady=(8, 0))
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)
        self.tree = ttk.Treeview(
            table_frame, columns=TREE_COLUMNS, show="headings", selectmode="browse")
        for name in TREE_COLUMNS:
            self.tree.heading(name, text=TREE_HEADINGS[name])
            self.tree.column(name, width=TREE_WIDTHS[name], minwidth=45,
                             stretch=name in ("item_name", "value"))
        self.tree.configure(displaycolumns=tuple(self.visible_columns))
        ybar = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        xbar = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=ybar.set, xscrollcommand=xbar.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        ybar.grid(row=0, column=1, sticky="ns")
        xbar.grid(row=1, column=0, sticky="we")
        self.tree.bind("<Double-1>", self._begin_cell_edit)

        status = tk.Label(outer, textvariable=self.status_var, anchor="w", justify="left",
                          fg="#333", wraplength=1160)
        status.grid(row=6, column=0, sticky="we", pady=(8, 4))

        buttons = ttk.Frame(outer)
        buttons.grid(row=7, column=0, sticky="e")
        ttk.Button(buttons, text="閉じる", command=self._close).pack(
            side="right", padx=(8, 0))
        self.all_button = ttk.Button(
            buttons, text="全FORM_IDを直接出力", command=self._generate_all)
        self.all_button.pack(side="right", padx=(8, 0))
        self.selected_button = ttk.Button(
            buttons, text="表示中FORM_IDを出力", command=self._generate_selected)
        self.selected_button.pack(side="right")

    def _choose_excel(self) -> None:
        path = filedialog.askopenfilename(
            parent=self.root, title="レイアウト定義Excelを選択",
            filetypes=(("Excel", "*.xlsx *.xlsm"), ("すべて", "*.*")))
        if not path:
            return
        self.excel_var.set(path)
        if not self.output_var.get():
            self.output_var.set(str(Path(path).parent / "layout_txt"))
        self._load_sheets()
        self._read_definitions(show_errors=False)

    def _choose_output(self) -> None:
        path = filedialog.askdirectory(parent=self.root, title="出力フォルダを選択")
        if path:
            self.output_var.set(path)

    def _load_sheets(self) -> None:
        path = Path(self.excel_var.get().strip())
        if not path.is_file():
            return
        try:
            wb = load_workbook(str(path), read_only=True, data_only=True)
            try:
                names = list(wb.sheetnames)
                active = wb.active.title
            finally:
                wb.close()
        except Exception as exc:
            messagebox.showerror("Excel読込エラー", str(exc), parent=self.root)
            return
        self.sheet_box.config(values=names)
        selected = self.sheet_var.get()
        self.sheet_var.set(selected if selected in names else active)

    def _definition_setting_changed(self, _event=None) -> None:
        self.loaded_fields = []
        self.form_fields = OrderedDict()
        self.form_box.config(values=[])
        self.form_var.set("")
        self._clear_tree()
        self.status_var.set("設定が変わりました。「Excel定義を読込」を押してください。")

    def _default_value_usage_changed(self) -> None:
        self._save_persisted_settings(show_message=False)
        if Path(self.excel_var.get().strip()).is_file():
            self._read_definitions(show_errors=False)
        else:
            self._definition_setting_changed()

    def _header_row(self):
        raw = self.header_var.get().strip()
        if not raw or raw in ("自動", "auto", "AUTO"):
            return None
        try:
            value = int(raw)
        except ValueError:
            raise LayoutTxtError("見出し行は数値または「自動」にしてください。")
        if value < 1:
            raise LayoutTxtError("見出し行は1以上にしてください。")
        return value

    def _read_args(self) -> dict:
        return {
            "excel_path": Path(self.excel_var.get().strip()),
            "sheet_name": self.sheet_var.get() or None,
            "header_row": self._header_row(),
            "form_column": self.form_col_var.get(),
            "layout_column": self.layout_col_var.get(),
            "field_column": self.field_col_var.get(),
            "item_column": self.item_col_var.get(),
            "data_type_column": self.type_col_var.get(),
            "ime_column": self.ime_col_var.get(),
            "max_digits_column": self.max_col_var.get(),
            "input_attribute_column": self.input_attribute_col_var.get(),
            "input_rule_column": self.input_rule_col_var.get(),
            "notes_column": self.notes_col_var.get(),
            "output_example_column": self.output_example_col_var.get(),
            "default_value_column": (
                self.default_value_col_var.get()
                if self.use_default_value_var.get() else "none"),
            "profile": _value_for_label(PROFILE_LABELS, self.profile_var.get()),
            "date_mode": _value_for_label(DATE_MODE_LABELS, self.date_mode_var.get()),
            "coverage_form_id": self.coverage_form_var.get().strip(),
        }

    def _read_definitions(self, show_errors=True) -> None:
        excel = Path(self.excel_var.get().strip())
        if not excel.is_file():
            if show_errors:
                messagebox.showwarning("入力なし", "入力Excelを選択してください。", parent=self.root)
            return
        self.status_var.set("Excel定義を読み込み中...")
        self.root.update_idletasks()
        try:
            fields, sheet, header, _columns = read_layout_fields(**self._read_args())
        except Exception as exc:
            self.status_var.set("Excel定義を読めませんでした: %s" % exc)
            if show_errors:
                messagebox.showerror("Excel定義エラー", str(exc), parent=self.root)
            return
        self.loaded_fields = fields
        self.form_fields = OrderedDict()
        for field in fields:
            self.form_fields.setdefault(field.form_id, []).append(field)
        form_ids = list(self.form_fields.keys())
        self.form_box.config(values=form_ids)
        if self.form_var.get() not in self.form_fields:
            self.form_var.set(form_ids[0])
        self._show_selected_form()
        self.status_var.set(
            "読込完了: シート %s / 見出し行 %d / FORM_ID %d件 / 項目 %d件"
            % (sheet, header, len(form_ids), len(fields)))

    def _clear_tree(self) -> None:
        self._finish_cell_edit(save=False)
        for item in self.tree.get_children(""):
            self.tree.delete(item)

    def _show_selected_form(self, _event=None) -> None:
        form_id = self.form_var.get().strip()
        if not self.form_fields:
            return
        if form_id not in self.form_fields:
            messagebox.showwarning("FORM_ID", "Excelに存在するFORM_IDを選択してください。",
                                   parent=self.root)
            return
        self._clear_tree()
        fields = self.form_fields[form_id]
        for field in fields:
            values = (
                "1", field.row_label, field.layout_id, field.field_id,
                field.item_name, field.data_type, field.ime_name,
                "NULL" if field.max_digits is None else field.max_digits,
                field.value, field.attribute_flag, field.coordinates,
                field.input_attribute, field.input_rule,
                field.notes, field.output_example,
            )
            self.tree.insert("", "end", iid=field.instance_key, values=values)
        self.form_summary_var.set("%s: %d項目（ダブルクリックで編集）" % (form_id, len(fields)))

    def _apply_visible_columns(self, save=True) -> None:
        if self.visible_column_vars:
            visible = [name for name in TREE_COLUMNS
                       if self.visible_column_vars[name].get()]
            if not visible:
                visible = ["item_name"]
                self.visible_column_vars["item_name"].set(True)
            self.visible_columns = visible
        if hasattr(self, "tree"):
            self.tree.configure(displaycolumns=tuple(self.visible_columns))
        if save and hasattr(self, "tree"):
            self._save_persisted_settings(show_message=False)

    def _show_all_columns(self) -> None:
        for variable in self.visible_column_vars.values():
            variable.set(True)
        self._apply_visible_columns()

    def _displayed_columns(self):
        raw = self.tree.cget("displaycolumns")
        if raw == "#all" or raw == ("#all",):
            return tuple(TREE_COLUMNS)
        if isinstance(raw, str):
            return tuple(self.root.tk.splitlist(raw))
        return tuple(raw)

    def _begin_cell_edit(self, event) -> None:
        region = self.tree.identify_region(event.x, event.y)
        item = self.tree.identify_row(event.y)
        column_token = self.tree.identify_column(event.x)
        if region != "cell" or not item or not column_token:
            return
        index = int(column_token[1:]) - 1
        displayed_columns = self._displayed_columns()
        if index < 0 or index >= len(displayed_columns):
            return
        column_name = displayed_columns[index]
        if column_name not in EDITABLE_COLUMNS:
            return
        values = list(self.tree.item(item, "values"))
        value_index = TREE_COLUMNS.index(column_name)
        if column_name == "include":
            values[value_index] = "0" if str(values[value_index]) == "1" else "1"
            self.tree.item(item, values=values)
            return
        self._finish_cell_edit(save=True)
        bbox = self.tree.bbox(item, column_token)
        if not bbox:
            return
        x, y, width, height = bbox
        editor = ttk.Entry(self.tree)
        editor.insert(0, values[value_index])
        editor.select_range(0, "end")
        editor.place(x=x, y=y, width=width, height=height)
        editor.focus_set()
        self.cell_editor = editor
        self.editor_context = (item, value_index)
        editor.bind("<Return>", lambda _event: self._finish_cell_edit(save=True))
        editor.bind("<Escape>", lambda _event: self._finish_cell_edit(save=False))
        editor.bind("<FocusOut>", lambda _event: self._finish_cell_edit(save=True))

    def _finish_cell_edit(self, save: bool) -> None:
        if self.cell_editor is None:
            return
        editor = self.cell_editor
        context = self.editor_context
        self.cell_editor = None
        self.editor_context = None
        if save and context is not None:
            item, index = context
            if self.tree.exists(item):
                values = list(self.tree.item(item, "values"))
                values[index] = editor.get()
                self.tree.item(item, values=values)
        editor.destroy()

    def _screen_edits(self):
        self._finish_cell_edit(save=True)
        rows = []  # type: List[str]
        overrides = {}  # type: Dict[str, Dict[str, str]]
        for item in self.tree.get_children(""):
            values = list(self.tree.item(item, "values"))
            if str(values[0]).strip() not in ("1", "ON", "on", "yes", "true"):
                continue
            rows.append(item)
            overrides[item] = {
                "field_id": values[3],
                "value": values[8],
                "attribute_flag": values[9],
                "coordinates": values[10],
            }
        if not rows:
            raise LayoutTxtError("出力対象項目が0件です。出力列を1にしてください。")
        return rows, overrides

    # ------------------------------------------------------------------
    # 複数FORM / 外部画像を蓄積するTAR出力リスト
    def _unique_package_base(self, preferred: str) -> str:
        raw = str(preferred or "image").strip() or "image"
        used = set(item.front_image_name.lower() for item in self.package_items.values())
        candidate = raw
        sequence = 2
        while (candidate + "F.tif").lower() in used or any(
                item.safe_base_name.lower() == candidate.lower()
                for item in self.package_items.values()):
            candidate = "%s_%02d" % (raw, sequence)
            sequence += 1
        return candidate

    @staticmethod
    def _package_source(item: PackageItem) -> str:
        if item.front_image_path is not None:
            value = str(item.front_image_path)
            if item.back_image_path is not None:
                value += " / 背面: " + str(item.back_image_path)
            return value
        return item.source_label or "画面から生成"

    def _package_values(self, item: PackageItem, include="1"):
        return (
            include,
            item.scan_batch_id,
            item.image_sequence,
            item.safe_base_name,
            item.front_image_name,
            item.back_image_name if item.has_back_image else "（なし）",
            item.arrival_date,
            item.form_id,
            item.application_number,
            item.reception_number,
            item.format_id,
            item.delivery_date,
            item.delivery_shot,
            (item.front_recognition_name
             if item.has_front_recognition else "（未設定）"),
            item.back_recognition_name if item.has_back_image else "（なし）",
            item.back_recognition_result if item.has_back_image else "",
            item.related_file or item.front_image_name,
            format_extra_fields(item.extra_fields),
            self._package_source(item),
        )

    def _prepare_package_csv_fields(self, item: PackageItem) -> None:
        match = re.match(r"^(\d{13})_(\d{3})(?:_|$)", item.safe_base_name)
        if not item.scan_batch_id:
            item.scan_batch_id = (
                self.package_scan_batch_id_var.get().strip()
                or (match.group(1) if match else ""))
        if not item.image_sequence:
            item.image_sequence = (
                match.group(2) if match else "%03d" % (self.package_sequence + 1))
        defaults = (
            ("arrival_date", self.package_arrival_date_var),
            ("application_number", self.package_application_number_var),
            ("reception_number", self.package_reception_number_var),
            ("format_id", self.package_format_id_var),
            ("delivery_date", self.package_delivery_date_var),
            ("delivery_shot", self.package_delivery_shot_var),
        )
        for name, variable in defaults:
            if not getattr(item, name):
                setattr(item, name, variable.get().strip())

    def _insert_package_item(self, item: PackageItem) -> str:
        self._prepare_package_csv_fields(item)
        self.package_sequence += 1
        iid = "package_%05d" % self.package_sequence
        self.package_items[iid] = item
        self.package_tree.insert(
            "", "end", iid=iid, values=self._package_values(item))
        self.package_tree.selection_set(iid)
        self.package_tree.see(iid)
        return iid

    def _apply_package_csv_defaults(self) -> None:
        self._finish_package_cell_edit(save=True)
        targets = list(self.package_tree.selection())
        if not targets:
            targets = list(self.package_tree.get_children(""))
        if not targets:
            messagebox.showwarning(
                "CSV初期値", "出力リストへ画像を追加してください。", parent=self.root)
            return
        values = {
            "scan_batch_id": self.package_scan_batch_id_var.get().strip(),
            "arrival_date": self.package_arrival_date_var.get().strip(),
            "application_number": self.package_application_number_var.get().strip(),
            "reception_number": self.package_reception_number_var.get().strip(),
            "format_id": self.package_format_id_var.get().strip(),
            "delivery_date": self.package_delivery_date_var.get().strip(),
            "delivery_shot": self.package_delivery_shot_var.get().strip(),
        }
        for iid in targets:
            item = self.package_items[iid]
            for name, value in values.items():
                setattr(item, name, value)
            include = self.package_tree.set(iid, "include") or "1"
            self.package_tree.item(iid, values=self._package_values(item, include))
        self.status_var.set("CSV初期値を%d件へ反映しました。" % len(targets))

    def _add_current_form_to_package(self) -> None:
        if not self.form_fields:
            self._read_definitions()
            if not self.form_fields:
                return
        form_id = self.form_var.get().strip()
        try:
            rows, overrides = self._screen_edits()
            self.status_var.set("FORM_ID %s の正面画像を生成中..." % form_id)
            self.root.update_idletasks()
            front = render_form_tif_payload(
                self.loaded_fields, form_id, selected_rows=rows,
                field_overrides=overrides, side="front")
            front_text = render_form_txt_text(
                self.loaded_fields, form_id, selected_rows=rows,
                field_overrides=overrides,
                output_format=_value_for_label(FORMAT_LABELS, self.format_var.get()))
            back = None
            if self.package_include_back_var.get():
                back = render_form_tif_payload(
                    self.loaded_fields, form_id, selected_rows=rows,
                    field_overrides=overrides, side="back")
            configured_stem = resolve_form_filename_stem(
                self.filename_template_var.get().strip(), form_id,
                Path(self.excel_var.get().strip()))
            base = self._unique_package_base(
                base_name_from_front(Path(configured_stem + ".txt")))
            item = PackageItem(
                base_name=base,
                form_id=form_id,
                front_recognition_text=front_text,
                back_recognition_result=self.package_result_var.get(),
                front_image_bytes=front,
                back_image_bytes=back,
                front_extension=".TIF",
                source_label="Excel/画面 FORM_ID %s" % form_id)
            self._insert_package_item(item)
        except (LayoutTxtError, LayoutTarError) as exc:
            self.status_var.set("出力リストへ追加できませんでした: %s" % exc)
            messagebox.showerror("追加エラー", str(exc), parent=self.root)
            return
        self.status_var.set(
            "出力リストへ追加: FORM_ID %s / %s / 正面TXTは既存ロジック / %s"
            % (form_id, item.front_image_name,
               ("背面 %s / 背面認識値 %s" % (
                   item.back_image_name, item.back_recognition_result)
                if item.has_back_image else "背面なし")))

    def _read_text_file(self, path: Optional[Path]) -> Optional[str]:
        if path is None:
            return None
        encodings = [self.encoding_var.get(), "cp932", "utf-8-sig", "utf-8"]
        tried = set()
        for encoding in encodings:
            if not encoding or encoding in tried:
                continue
            tried.add(encoding)
            try:
                return path.read_text(encoding=encoding).rstrip("\r\n")
            except (UnicodeDecodeError, LookupError):
                continue
            except OSError as exc:
                raise LayoutTarError("認識結果TXTを読めません: %s: %s" % (path, exc))
        raise LayoutTarError("認識結果TXTの文字コードを判定できません: %s" % path)

    def _read_back_recognition_result(self, path: Optional[Path]) -> str:
        text = self._read_text_file(path)
        if text is None:
            return self.package_result_var.get()
        rows = list(csv.reader(text.splitlines()))
        if len(rows) != 1 or len(rows[0]) != 1:
            raise LayoutTarError("背面TXTは1行1フィールドにしてください: %s" % path)
        return rows[0][0]

    def _choose_back_for_front(self, front: Path) -> Optional[Path]:
        found = matching_back_image(front)
        if found is not None:
            return found
        chosen = filedialog.askopenfilename(
            parent=self.root,
            title="%s に対応する背面画像（末尾R）を選択" % front.name,
            initialdir=str(front.parent),
            filetypes=(("画像", "*.tif *.tiff *.png *.jpg *.jpeg *.bmp *.gif"),
                       ("すべて", "*.*")))
        return Path(chosen) if chosen else None

    def _add_external_front_images(self) -> None:
        paths = filedialog.askopenfilenames(
            parent=self.root, title="正面画像を選択（複数可）",
            filetypes=(("画像", "*.tif *.tiff *.png *.jpg *.jpeg *.bmp *.gif"),
                       ("すべて", "*.*")))
        if not paths:
            return
        added = 0
        skipped = []
        missing_front_txt = []
        form_front_text = None
        form_text_ready = False
        try:
            for raw in paths:
                front = Path(raw)
                if front.suffix.lower() not in IMAGE_EXTENSIONS:
                    skipped.append("%s（画像形式外）" % front.name)
                    continue
                if front.stem.upper().endswith("R"):
                    skipped.append("%s（背面Rが正面として選択されています）" % front.name)
                    continue
                back = None
                if self.package_include_back_var.get():
                    back = self._choose_back_for_front(front)
                    if back is None:
                        skipped.append("%s（背面未選択）" % front.name)
                        continue
                    if back.suffix.lower() not in IMAGE_EXTENSIONS:
                        skipped.append("%s（背面画像形式外）" % front.name)
                        continue
                base = self._unique_package_base(base_name_from_front(front))
                front_text = self._read_text_file(matching_recognition_file(front))
                if front_text is None and self.form_fields:
                    if not form_text_ready:
                        rows, overrides = self._screen_edits()
                        form_front_text = render_form_txt_text(
                            self.loaded_fields, self.form_var.get().strip(),
                            selected_rows=rows, field_overrides=overrides,
                            output_format=_value_for_label(
                                FORMAT_LABELS, self.format_var.get()))
                        form_text_ready = True
                    front_text = form_front_text
                if front_text is None:
                    missing_front_txt.append(front.name)
                back_result = self.package_result_var.get()
                if back is not None:
                    back_result = self._read_back_recognition_result(
                        matching_recognition_file(back))
                item = PackageItem(
                    base_name=base,
                    form_id=self.form_var.get().strip(),
                    front_recognition_text=front_text,
                    back_recognition_result=back_result,
                    front_image_path=front,
                    front_extension=front.suffix,
                    back_image_path=back,
                    back_extension=back.suffix if back is not None else front.suffix,
                    source_label=str(front))
                self._insert_package_item(item)
                added += 1
        except (LayoutTxtError, LayoutTarError) as exc:
            messagebox.showerror("画像追加エラー", str(exc), parent=self.root)
            self.status_var.set("外部画像を追加できませんでした: %s" % exc)
            return

        message = "外部画像を%d件、出力リストへ追加しました。" % added
        if missing_front_txt:
            message += " 正面TXT未設定: " + ", ".join(missing_front_txt)
        if skipped:
            message += " スキップ: " + ", ".join(skipped)
        if skipped or missing_front_txt:
            messagebox.showwarning("追加結果の確認", message, parent=self.root)
        self.status_var.set(message)

    def _set_back_image_for_selected(self) -> None:
        selected = list(self.package_tree.selection())
        if len(selected) != 1:
            messagebox.showwarning(
                "背面画像", "背面画像を設定する行を1件選択してください。",
                parent=self.root)
            return
        iid = selected[0]
        item = self.package_items[iid]
        if item.has_back_image:
            remove = messagebox.askyesnocancel(
                "背面画像",
                "現在の背面画像を外しますか？\n\n"
                "「はい」: 背面なしにする\n「いいえ」: 別の背面画像へ変更",
                parent=self.root)
            if remove is None:
                return
            if remove:
                item.back_image_path = None
                item.back_image_bytes = None
                self.package_tree.item(
                    iid, values=self._package_values(
                        item, self.package_tree.set(iid, "include")))
                return

        chosen = filedialog.askopenfilename(
            parent=self.root, title="背面画像を選択",
            filetypes=(("画像", "*.tif *.tiff *.png *.jpg *.jpeg *.bmp *.gif"),
                       ("すべて", "*.*")))
        if not chosen:
            return
        path = Path(chosen)
        if path.suffix.lower() not in IMAGE_EXTENSIONS:
            messagebox.showerror("背面画像", "対応していない画像形式です。", parent=self.root)
            return
        try:
            result_path = matching_recognition_file(path)
            if result_path is not None:
                back_result = self._read_back_recognition_result(result_path)
            else:
                back_result = item.back_recognition_result
        except LayoutTarError as exc:
            messagebox.showerror("背面TXT", str(exc), parent=self.root)
            return
        item.back_image_path = path
        item.back_image_bytes = None
        item.back_extension = path.suffix
        item.back_recognition_result = back_result
        self.package_tree.item(
            iid, values=self._package_values(item, self.package_tree.set(iid, "include")))

    def _remove_package_items(self) -> None:
        self._finish_package_cell_edit(save=True)
        for iid in list(self.package_tree.selection()):
            self.package_items.pop(iid, None)
            if self.package_tree.exists(iid):
                self.package_tree.delete(iid)
        self.status_var.set("出力リスト: %d件" % len(self.package_items))

    def _clear_package_items(self) -> None:
        if not self.package_items:
            return
        if not messagebox.askyesno(
                "出力リストを消去", "出力リストを全件消去しますか？",
                parent=self.root):
            return
        self._finish_package_cell_edit(save=False)
        for iid in self.package_tree.get_children(""):
            self.package_tree.delete(iid)
        self.package_items.clear()
        self.status_var.set("出力リストを消去しました。元画像ファイルは変更していません。")

    def _begin_package_cell_edit(self, event) -> None:
        region = self.package_tree.identify_region(event.x, event.y)
        item_id = self.package_tree.identify_row(event.y)
        column_token = self.package_tree.identify_column(event.x)
        if region != "cell" or not item_id or not column_token:
            return
        index = int(column_token[1:]) - 1
        if index < 0 or index >= len(PACKAGE_COLUMNS):
            return
        column_name = PACKAGE_COLUMNS[index]
        if column_name not in PACKAGE_EDITABLE_COLUMNS:
            return
        values = list(self.package_tree.item(item_id, "values"))
        if column_name == "include":
            values[index] = "0" if str(values[index]) == "1" else "1"
            self.package_tree.item(item_id, values=values)
            return
        self._finish_package_cell_edit(save=True)
        bbox = self.package_tree.bbox(item_id, column_token)
        if not bbox:
            return
        x, y, width, height = bbox
        editor = ttk.Entry(self.package_tree)
        editor.insert(0, values[index])
        editor.select_range(0, "end")
        editor.place(x=x, y=y, width=width, height=height)
        editor.focus_set()
        self.package_editor = editor
        self.package_editor_context = (item_id, column_name)
        editor.bind("<Return>", lambda _event: self._finish_package_cell_edit(save=True))
        editor.bind("<Escape>", lambda _event: self._finish_package_cell_edit(save=False))
        editor.bind("<FocusOut>", lambda _event: self._finish_package_cell_edit(save=True))

    def _finish_package_cell_edit(self, save: bool) -> None:
        if self.package_editor is None:
            return
        editor = self.package_editor
        context = self.package_editor_context
        self.package_editor = None
        self.package_editor_context = None
        try:
            if save and context is not None:
                iid, column_name = context
                if self.package_tree.exists(iid) and iid in self.package_items:
                    item = self.package_items[iid]
                    raw = editor.get()
                    if column_name in (
                            "scan_batch_id", "image_sequence", "arrival_date",
                            "application_number", "reception_number", "format_id",
                            "delivery_date", "delivery_shot"):
                        setattr(item, column_name, raw.strip())
                    elif column_name == "form_id":
                        item.form_id = raw.strip()
                    elif column_name == "base_name":
                        if not raw.strip():
                            raise LayoutTarError("基礎名は空にできません")
                        item.base_name = raw.strip()
                    elif column_name == "back_recognition_result":
                        item.back_recognition_result = raw
                    elif column_name == "related_file":
                        item.related_file = raw.strip()
                    elif column_name == "extra_fields":
                        item.extra_fields = parse_extra_fields(raw)
                    include = self.package_tree.set(iid, "include") or "1"
                    self.package_tree.item(
                        iid, values=self._package_values(item, include))
        except LayoutTarError as exc:
            self.status_var.set("出力リストの編集値が不正です: %s" % exc)
            messagebox.showerror("編集エラー", str(exc), parent=self.root)
        finally:
            editor.destroy()

    def _selected_package_items(self):
        self._finish_package_cell_edit(save=True)
        items = []
        for iid in self.package_tree.get_children(""):
            include = str(self.package_tree.set(iid, "include")).strip().lower()
            if include in ("1", "on", "yes", "true"):
                items.append(self.package_items[iid])
        if not items:
            raise LayoutTarError("TAR列が1の画像を1件以上追加してください")
        return items

    def _custom_package_tar_name(self, items, include_manifest: bool) -> str:
        excel_raw = self.excel_var.get().strip()
        source = Path(excel_raw).stem if excel_raw else "images"
        return format_package_tar_name(
            self.tar_name_var.get(), self.package_tar_name_var.get(),
            include_manifest, source, [item.form_id for item in items])

    def _run_package(self, include_manifest: bool) -> None:
        output_raw = self.output_var.get().strip()
        if not output_raw:
            messagebox.showwarning(
                "出力先なし", "出力フォルダを選択してください。", parent=self.root)
            return
        try:
            items = self._selected_package_items()
            manifest_style = _value_for_label(
                CSV_STYLE_LABELS, self.manifest_style_var.get())
            columns = None
            if include_manifest and manifest_style == "custom":
                columns = parse_manifest_columns(self.manifest_columns_var.get())
            result = build_image_tar(
                items=items,
                output_dir=Path(output_raw),
                tar_name=self._custom_package_tar_name(items, include_manifest),
                include_recognition_txt=not include_manifest,
                include_manifest_csv=include_manifest,
                manifest_name=self.manifest_name_var.get().strip(),
                manifest_columns=columns,
                manifest_style=manifest_style,
                text_encoding=self.encoding_var.get(),
                csv_encoding=self.manifest_encoding_var.get(),
                overwrite=self.overwrite_var.get(),
                create_view_folder=True)
        except (LayoutTarError, OSError) as exc:
            self.status_var.set("TARを生成できませんでした: %s" % exc)
            messagebox.showerror("TAR生成エラー", str(exc), parent=self.root)
            return

        self._save_persisted_settings(show_message=False)
        self.status_var.set(
            "梱包完了: 帳票%d件 / TAR内%dファイル / %s / %s"
            % (result.item_count, len(result.archive_members), result.tar_file,
               result.view_folder))
        messagebox.showinfo(
            "TAR生成完了",
            "画像TARを生成しました。\n\n帳票: %d件\n格納ファイル: %d件\n"
            "CSV: %s\n\nTAR: %s\n確認用フォルダ: %s"
            % (result.item_count, len(result.archive_members),
               result.manifest_name or "なし", result.tar_file,
               result.view_folder),
            parent=self.root)

    def _save_defaults_to_excel(self) -> None:
        if not self.form_fields:
            self._read_definitions()
            if not self.form_fields:
                return
        self._finish_cell_edit(save=True)
        form_id = self.form_var.get().strip()
        values_by_instance = {}
        value_index = TREE_COLUMNS.index("value")
        for item in self.tree.get_children(""):
            values = list(self.tree.item(item, "values"))
            values_by_instance[item] = values[value_index]
        if not values_by_instance:
            messagebox.showwarning(
                "既定値保存", "保存するOCR値がありません。", parent=self.root)
            return

        excel = Path(self.excel_var.get().strip())
        if not excel.is_file():
            messagebox.showwarning(
                "入力なし", "入力Excelを選択してください。", parent=self.root)
            return
        if not messagebox.askyesno(
                "既定値をExcelへ保存",
                "表示中FORM_ID %s のOCR値を入力Excelへ書き込みます。\n"
                "Excelを開いている場合は閉じてください。\n\n実行しますか？" % form_id,
                parent=self.root):
            return

        self.defaults_button.config(state="disabled")
        self.status_var.set("OCR既定値をExcelへ保存中...")
        self.root.update_idletasks()
        try:
            path, column, row_count = save_layout_default_values(
                excel_path=excel,
                values_by_instance=values_by_instance,
                sheet_name=self.sheet_var.get() or None,
                header_row=self._header_row(),
                default_value_column=self.default_value_col_var.get())
        except LayoutTxtError as exc:
            self.status_var.set("既定値を保存できませんでした: %s" % exc)
            messagebox.showerror("既定値保存エラー", str(exc), parent=self.root)
        except Exception as exc:  # noqa: BLE001 - GUI境界で理由を表示する
            self.status_var.set("既定値の保存に失敗しました: %s" % exc)
            messagebox.showerror("既定値保存エラー", str(exc), parent=self.root)
        else:
            self._read_definitions(show_errors=False)
            self.status_var.set(
                "既定値保存完了: FORM_ID %s / %d Excel行 / %s列"
                % (form_id, row_count, get_column_letter(column)))
            messagebox.showinfo(
                "既定値保存完了",
                "表示中FORM_IDのOCR値をExcelへ保存しました。\n\n"
                "ファイル: %s\n列: %s\n更新行: %d\n\n"
                "次回読込時はこの既定値を優先します。"
                % (path, get_column_letter(column), row_count),
                parent=self.root)
        finally:
            self.defaults_button.config(state="normal")

    def _tar_only_changed(self) -> None:
        if self.tar_only_var.get():
            self.create_tar_var.set(True)

    def _validate_paths(self):
        excel = Path(self.excel_var.get().strip())
        output_raw = self.output_var.get().strip()
        if not excel.is_file():
            raise LayoutTxtError("入力Excelを選択してください。")
        if not output_raw:
            raise LayoutTxtError("出力フォルダを選択してください。")
        return excel, Path(output_raw)

    def _generate_selected(self) -> None:
        if not self.form_fields:
            self._read_definitions()
            if not self.form_fields:
                return
        form_id = self.form_var.get().strip()
        try:
            rows, overrides = self._screen_edits()
        except LayoutTxtError as exc:
            messagebox.showerror("設定エラー", str(exc), parent=self.root)
            return
        self._run_generation(
            selected_form_ids=[form_id], selected_rows=rows,
            field_overrides=overrides, error_patterns="none")

    def _generate_all(self) -> None:
        self._run_generation(
            selected_form_ids=None, selected_rows=None, field_overrides=None,
            error_patterns=_value_for_label(
                ERROR_PATTERN_LABELS, self.error_pattern_var.get()))

    def _run_generation(self, selected_form_ids, selected_rows,
                        field_overrides, error_patterns) -> None:
        try:
            excel, output = self._validate_paths()
        except LayoutTxtError as exc:
            messagebox.showwarning("入力なし", str(exc), parent=self.root)
            return
        self.selected_button.config(state="disabled")
        self.all_button.config(state="disabled")
        self.status_var.set("生成中...")
        self.root.update_idletasks()
        try:
            result = generate_layout_txt(
                excel_path=excel, output_dir=output,
                sheet_name=self.sheet_var.get() or None, header_row=self._header_row(),
                form_column=self.form_col_var.get(), layout_column=self.layout_col_var.get(),
                field_column=self.field_col_var.get(), item_column=self.item_col_var.get(),
                data_type_column=self.type_col_var.get(), ime_column=self.ime_col_var.get(),
                max_digits_column=self.max_col_var.get(),
                input_attribute_column=self.input_attribute_col_var.get(),
                input_rule_column=self.input_rule_col_var.get(),
                notes_column=self.notes_col_var.get(),
                output_example_column=self.output_example_col_var.get(),
                default_value_column=(
                    self.default_value_col_var.get()
                    if self.use_default_value_var.get() else "none"),
                profile=_value_for_label(PROFILE_LABELS, self.profile_var.get()),
                date_mode=_value_for_label(DATE_MODE_LABELS, self.date_mode_var.get()),
                coverage_form_id=self.coverage_form_var.get().strip(),
                error_patterns=error_patterns,
                filename_template=self.filename_template_var.get().strip(),
                generate_tif=self.generate_tif_var.get(),
                selected_form_ids=selected_form_ids,
                selected_rows=selected_rows, field_overrides=field_overrides,
                create_tar=self.create_tar_var.get() or self.tar_only_var.get(),
                tar_name=self.tar_name_var.get().strip(), tar_only=self.tar_only_var.get(),
                output_format=_value_for_label(FORMAT_LABELS, self.format_var.get()),
                encoding=self.encoding_var.get(), split_by_form=self.split_var.get(),
                overwrite=self.overwrite_var.get())
        except LayoutTxtError as exc:
            self.status_var.set("生成できませんでした: %s" % exc)
            messagebox.showerror("設定エラー", str(exc), parent=self.root)
        except Exception as exc:  # noqa: BLE001 - GUI境界で理由を表示する
            self.status_var.set("生成に失敗しました: %s" % exc)
            messagebox.showerror("生成エラー", str(exc), parent=self.root)
        else:
            _print_result(result)
            self._save_persisted_settings(show_message=False)
            tar_text = str(result.tar_file) if result.tar_file is not None else "なし"
            self.status_var.set(
                "完了: FORM_ID %d件 / 項目%d件 / TXT%d件 / TIF%d件 / TAR %s"
                % (result.form_count, result.field_count, len(result.txt_files),
                   len(result.tif_files), tar_text))
            messagebox.showinfo(
                "生成完了",
                "Layoutデータを生成しました。\n\nFORM_ID: %d件\n項目: %d件\n"
                "TXT: %d件\nTIF: %d件\nTAR: %s\n\n%s"
                % (result.form_count, result.field_count, len(result.txt_files),
                   len(result.tif_files), tar_text, output), parent=self.root)
        finally:
            self.selected_button.config(state="normal")
            self.all_button.config(state="normal")


def main(initial_excel: Optional[Path] = None) -> int:
    root = tk.Tk()
    LayoutTxtGui(root, initial_excel=initial_excel)
    root.mainloop()
    return 0


if __name__ == "__main__":
    sys.exit(main())
